"""Benchmarking module — captures comprehensive metrics for IEEE-grade evaluation."""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


@dataclass
class PageMetrics:
    """Metrics collected for a single page."""

    source_file: str = ""
    page_number: int = 0
    image_width: int = 0
    image_height: int = 0
    page_time_s: float = 0.0
    ocr_init_time_s: float = 0.0
    ocr_inference_time_s: float = 0.0
    layout_detection_time_s: float = 0.0
    vlm_inference_time_s: float = 0.0
    extraction_time_s: float = 0.0
    validation_time_s: float = 0.0
    total_boxes_detected: int = 0
    rows_extracted: int = 0
    empty_rows_skipped: int = 0
    vlm_fallbacks: int = 0


@dataclass
class RowMetrics:
    """Metrics collected for a single extracted row."""

    source_file: str = ""
    page_number: int = 0
    row_index: int = 0
    employee_name: str = ""
    patient_name: str = ""

    # Raw OCR text
    raw_ocr_date: str = ""
    raw_ocr_time_in: str = ""
    raw_ocr_time_out: str = ""
    raw_ocr_hours: str = ""

    # Parsed values
    parsed_date: str = ""
    parsed_time_in: str = ""
    parsed_time_out: str = ""
    parsed_hours: str = ""

    # Confidence scores
    date_confidence: float = 0.0
    time_in_confidence: float = 0.0
    time_out_confidence: float = 0.0
    hours_confidence: float = 0.0

    # Source tracking
    date_source: str = ""
    time_in_source: str = ""
    time_out_source: str = ""
    hours_source: str = ""

    # Derived
    calculated_hours: Optional[float] = None
    total_hours_text: str = ""
    is_overnight: bool = False

    # Validation
    status: str = ""
    validation_errors: str = ""

    # Corrections analysis
    corrections_applied: int = 0
    corrections_detail: str = ""


@dataclass
class RunMetrics:
    """Aggregate metrics for an entire file run."""

    source_file: str = ""
    model_type: str = ""
    extraction_mode: str = "ppocr_grid"
    device: str = ""
    image_dpi: int = 0
    image_dimensions: str = ""
    file_size_kb: float = 0.0
    num_pages: int = 0
    total_time_s: float = 0.0
    avg_page_time_s: float = 0.0
    total_rows_extracted: int = 0
    accepted_count: int = 0
    flagged_count: int = 0
    failed_count: int = 0
    mean_overall_confidence: float = 0.0
    min_overall_confidence: float = 0.0
    max_overall_confidence: float = 0.0
    total_ocr_boxes: int = 0
    vlm_fallbacks_triggered: int = 0
    empty_rows_skipped: int = 0
    hours_mismatch_rate: float = 0.0
    field_missing_rate: float = 0.0
    mean_cer: float = 0.0  # Character Error Rate (if ground truth available)


class BenchmarkCollector:
    """Collects and exports benchmark metrics during pipeline execution."""

    def __init__(self) -> None:
        self.run: RunMetrics = RunMetrics()
        self.pages: list[PageMetrics] = []
        self.rows: list[RowMetrics] = []
        self._runs: list[tuple[RunMetrics, list[PageMetrics], list[RowMetrics]]] = []

    def start_run(
        self,
        source_file: str,
        model_type: str,
        device: str,
        image_dpi: int,
        file_size_kb: float,
        extraction_mode: str = "ppocr_grid",
    ) -> None:
        self.run = RunMetrics(
            source_file=source_file,
            model_type=model_type,
            extraction_mode=extraction_mode,
            device=device,
            image_dpi=image_dpi,
            file_size_kb=file_size_kb,
        )
        self.pages = []
        self.rows = []

    def add_page(self, page: PageMetrics) -> None:
        self.pages.append(page)

    def add_row(self, row: RowMetrics) -> None:
        self.rows.append(row)

    def snapshot_run(self, mode: str = "ppocr") -> None:
        """Snapshot the current run data for later combined export."""
        import copy

        self.run.extraction_mode = mode
        run_copy = copy.deepcopy(self.run)
        pages_copy = copy.deepcopy(self.pages)
        rows_copy = copy.deepcopy(self.rows)
        self._runs.append((run_copy, pages_copy, rows_copy))

    def finalize(self, total_time_s: float) -> None:
        """Compute aggregate statistics after all pages/rows are collected."""
        self.run.num_pages = len(self.pages)
        self.run.total_time_s = total_time_s
        self.run.avg_page_time_s = total_time_s / len(self.pages) if self.pages else 0.0
        self.run.total_rows_extracted = len(self.rows)
        self.run.total_ocr_boxes = sum(p.total_boxes_detected for p in self.pages)
        self.run.vlm_fallbacks_triggered = sum(p.vlm_fallbacks for p in self.pages)
        self.run.empty_rows_skipped = sum(p.empty_rows_skipped for p in self.pages)

        # Status counts
        self.run.accepted_count = sum(1 for r in self.rows if r.status == "accepted")
        self.run.flagged_count = sum(1 for r in self.rows if r.status == "flagged")
        self.run.failed_count = sum(1 for r in self.rows if r.status == "failed")

        # Confidence stats
        all_confs = []
        for r in self.rows:
            for c in [
                r.date_confidence,
                r.time_in_confidence,
                r.time_out_confidence,
                r.hours_confidence,
            ]:
                if c > 0:
                    all_confs.append(c)

        if all_confs:
            self.run.mean_overall_confidence = sum(all_confs) / len(all_confs)
            self.run.min_overall_confidence = min(all_confs)
            self.run.max_overall_confidence = max(all_confs)

        # Hours mismatch rate
        rows_with_hours = [
            r
            for r in self.rows
            if r.calculated_hours is not None and r.total_hours_text
        ]
        if rows_with_hours:
            mismatches = sum(
                1 for r in rows_with_hours if "hours_mismatch" in r.validation_errors
            )
            self.run.hours_mismatch_rate = mismatches / len(rows_with_hours)

        # Field missing rate
        if self.rows:
            missing = sum(
                1
                for r in self.rows
                if not r.parsed_date or not r.parsed_time_in or not r.parsed_time_out
            )
            self.run.field_missing_rate = missing / len(self.rows)

        # Mean CER (character error rate) — raw vs parsed comparison
        # This measures how much the parser had to "fix" the OCR output
        cer_values = []
        for r in self.rows:
            for raw, parsed in [
                (r.raw_ocr_time_in, r.parsed_time_in),
                (r.raw_ocr_time_out, r.parsed_time_out),
            ]:
                if raw and parsed:
                    raw_clean = raw.strip().lower()
                    parsed_clean = parsed.strip().lower()
                    if raw_clean != parsed_clean:
                        # Compute Levenshtein-based CER
                        cer = _levenshtein_cer(raw_clean, parsed_clean)
                        cer_values.append(cer)

        self.run.mean_cer = sum(cer_values) / len(cer_values) if cer_values else 0.0

    def export(self, output_dir: Path) -> Path:
        """Export benchmark results to a standalone Excel file."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ── Sheet 1: Run Summary ──
        ws_run = wb.active
        ws_run.title = "Run Summary"
        _write_run_summary(ws_run, self.run)

        # ── Sheet 2: Page Details ──
        ws_page = wb.create_sheet("Page Details")
        _write_page_details(ws_page, self.pages)

        # ── Sheet 3: Row-Level ──
        ws_row = wb.create_sheet("Row-Level")
        _write_row_level(ws_row, self.rows)

        # ── Sheet 4: Corrections Log ──
        ws_corr = wb.create_sheet("Corrections")
        _write_corrections(ws_corr, self.rows)

        output_dir.mkdir(parents=True, exist_ok=True)
        stem = Path(self.run.source_file).stem
        benchmark_path = output_dir / f"benchmark_{stem}.xlsx"
        wb.save(benchmark_path)
        logger.info(f"Benchmark exported: {benchmark_path}")
        return benchmark_path

    def export_combined(self, output_dir: Path, anonymizer=None) -> Path:
        """Export a combined benchmark across all runs with paper-ready table format."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        if not self._runs:
            raise ValueError("No runs to combine. Call snapshot_run() after each file.")

        wb = Workbook()

        # ── Sheet 1: Per-File Results (paper-ready table) ──
        ws = wb.active
        ws.title = "Per-File Results"
        _write_combined_summary(ws, self._runs)

        # ── Sheet 2: Page Details ──
        all_pages: list[PageMetrics] = []
        for _, pages, _ in self._runs:
            all_pages.extend(pages)
        _write_page_details(wb.create_sheet("Page Details"), all_pages)

        # ── Sheet 3: Row-Level ──
        all_rows: list[RowMetrics] = []
        for _, _, rows in self._runs:
            all_rows.extend(rows)
        _write_row_level(wb.create_sheet("Row-Level"), all_rows)

        # ── Sheet 4: Corrections ──
        _write_corrections(wb.create_sheet("Corrections"), all_rows)

        output_dir.mkdir(parents=True, exist_ok=True)
        combined_path = output_dir / "benchmark_combined.xlsx"
        wb.save(combined_path)
        logger.info(f"Combined benchmark exported: {combined_path}")
        return combined_path


def _levenshtein_cer(s1: str, s2: str) -> float:
    """Compute Character Error Rate using Levenshtein distance."""
    if not s1 and not s2:
        return 0.0
    m, n = len(s1), len(s2)
    dp = [[0] * (n + 1) for _ in range(m + 1)]
    for i in range(m + 1):
        dp[i][0] = i
    for j in range(n + 1):
        dp[0][j] = j
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            if s1[i - 1] == s2[j - 1]:
                dp[i][j] = dp[i - 1][j - 1]
            else:
                dp[i][j] = 1 + min(dp[i - 1][j], dp[i][j - 1], dp[i - 1][j - 1])
    return dp[m][n] / max(len(s1), len(s2))


def _style_header_row(ws, row: int, num_cols: int) -> None:
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(
        start_color="2F5496", end_color="2F5496", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def _auto_fit(ws, max_row: int, max_col: int) -> None:
    for col_idx in range(1, max_col + 1):
        max_len = 0
        for row_idx in range(1, max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 45)


def _write_run_summary(ws, run: RunMetrics) -> None:
    headers = ["Metric", "Value"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    _style_header_row(ws, 1, 2)

    metrics = [
        ("Source File", run.source_file),
        ("Model Type", run.model_type),
        ("Device", run.device),
        ("Image DPI", run.image_dpi),
        ("Image Dimensions", run.image_dimensions),
        ("File Size (KB)", round(run.file_size_kb, 1)),
        ("Number of Pages", run.num_pages),
        ("Total Processing Time (s)", round(run.total_time_s, 2)),
        ("Avg Page Time (s)", round(run.avg_page_time_s, 2)),
        ("Total Rows Extracted", run.total_rows_extracted),
        ("Accepted Rows", run.accepted_count),
        ("Flagged Rows", run.flagged_count),
        ("Failed Rows", run.failed_count),
        (
            "Acceptance Rate",
            f"{run.accepted_count / max(run.total_rows_extracted, 1) * 100:.1f}%",
        ),
        ("Mean Overall Confidence", round(run.mean_overall_confidence, 4)),
        ("Min Overall Confidence", round(run.min_overall_confidence, 4)),
        ("Max Overall Confidence", round(run.max_overall_confidence, 4)),
        ("Total OCR Boxes Detected", run.total_ocr_boxes),
        ("VLM Fallbacks Triggered", run.vlm_fallbacks_triggered),
        ("Empty Rows Skipped", run.empty_rows_skipped),
        ("Hours Mismatch Rate", f"{run.hours_mismatch_rate * 100:.1f}%"),
        ("Field Missing Rate", f"{run.field_missing_rate * 100:.1f}%"),
        ("Mean Character Error Rate", round(run.mean_cer, 4)),
    ]

    for row_idx, (metric, value) in enumerate(metrics, 2):
        ws.cell(row=row_idx, column=1, value=metric)
        ws.cell(row=row_idx, column=2, value=value)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 30


def _write_page_details(ws, pages: list[PageMetrics]) -> None:
    headers = [
        "Source File",
        "Page #",
        "Image Width",
        "Image Height",
        "Page Time (s)",
        "OCR Init Time (s)",
        "OCR Inference Time (s)",
        "Layout Detection Time (s)",
        "Extraction Time (s)",
        "Validation Time (s)",
        "Total Boxes Detected",
        "Rows Extracted",
        "Empty Rows Skipped",
        "VLM Fallbacks",
    ]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    _style_header_row(ws, 1, len(headers))

    for row_idx, p in enumerate(pages, 2):
        values = [
            p.source_file,
            p.page_number,
            p.image_width,
            p.image_height,
            round(p.page_time_s, 2),
            round(p.ocr_init_time_s, 2),
            round(p.ocr_inference_time_s, 2),
            round(p.layout_detection_time_s, 2),
            round(p.extraction_time_s, 2),
            round(p.validation_time_s, 2),
            p.total_boxes_detected,
            p.rows_extracted,
            p.empty_rows_skipped,
            p.vlm_fallbacks,
        ]
        for col_idx, v in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=v)

    _auto_fit(ws, len(pages) + 1, len(headers))


def _write_row_level(ws, rows: list[RowMetrics]) -> None:
    headers = [
        "Source File",
        "Page #",
        "Row Index",
        "Employee Name",
        "Patient Name",
        "Raw OCR Date",
        "Parsed Date",
        "Date Confidence",
        "Date Source",
        "Raw OCR Time In",
        "Parsed Time In",
        "Time In Confidence",
        "Time In Source",
        "Raw OCR Time Out",
        "Parsed Time Out",
        "Time Out Confidence",
        "Time Out Source",
        "Raw OCR Hours",
        "Parsed Hours",
        "Hours Confidence",
        "Hours Source",
        "Calculated Hours",
        "Written Hours",
        "Is Overnight",
        "Status",
        "Validation Errors",
        "Corrections Applied",
        "Corrections Detail",
    ]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    _style_header_row(ws, 1, len(headers))

    for row_idx, r in enumerate(rows, 2):
        values = [
            r.source_file,
            r.page_number,
            r.row_index,
            r.employee_name,
            r.patient_name,
            r.raw_ocr_date,
            r.parsed_date,
            r.date_confidence,
            r.date_source,
            r.raw_ocr_time_in,
            r.parsed_time_in,
            r.time_in_confidence,
            r.time_in_source,
            r.raw_ocr_time_out,
            r.parsed_time_out,
            r.time_out_confidence,
            r.time_out_source,
            r.raw_ocr_hours,
            r.parsed_hours,
            r.hours_confidence,
            r.hours_source,
            r.calculated_hours,
            r.total_hours_text,
            "Yes" if r.is_overnight else "",
            r.status,
            r.validation_errors,
            r.corrections_applied,
            r.corrections_detail,
        ]
        for col_idx, v in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=v)

    _auto_fit(ws, len(rows) + 1, len(headers))


def _write_corrections(ws, rows: list[RowMetrics]) -> None:
    """Detailed log of every correction the parser applied."""
    headers = [
        "Source File",
        "Page #",
        "Row Index",
        "Field",
        "Raw OCR Value",
        "Corrected/Parsed Value",
        "Correction Type",
    ]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    _style_header_row(ws, 1, len(headers))

    data_row = 2
    for r in rows:
        corrections = _parse_corrections_detail(r.corrections_detail)
        for field_name, raw_val, corrected_val, correction_type in corrections:
            values = [
                r.source_file,
                r.page_number,
                r.row_index,
                field_name,
                raw_val,
                corrected_val,
                correction_type,
            ]
            for col_idx, v in enumerate(values, 1):
                ws.cell(row=data_row, column=col_idx, value=v)
            data_row += 1

    if data_row == 2:
        ws.cell(row=2, column=1, value="No corrections applied")

    _auto_fit(ws, max(data_row, 2), len(headers))


def _parse_corrections_detail(detail: str) -> list[tuple[str, str, str, str]]:
    """Parse corrections_detail string into structured tuples."""
    if not detail:
        return []
    corrections = []
    for entry in detail.split("; "):
        parts = entry.split("|", 2)
        if len(parts) == 3:
            field_name = parts[0].strip()
            raw_val = parts[1].strip()
            corrected_val = parts[2].strip()
            correction_type = _classify_correction(raw_val, corrected_val)
            corrections.append((field_name, raw_val, corrected_val, correction_type))
    return corrections


def _classify_correction(raw: str, corrected: str) -> str:
    """Classify the type of correction applied."""
    if not raw or not corrected:
        return "missing_value"
    if raw.lower() == corrected.lower():
        return "no_change"
    if re.search(r"[0\.C]M", raw) and "PM" in corrected.upper():
        return "AM/PM hallucination fix"
    if re.search(r"\s*P$", raw) and "PM" in corrected.upper():
        return "truncated AM/PM fix"
    if re.search(r"\s*A$", raw) and "AM" in corrected.upper():
        return "truncated AM/PM fix"
    if raw.replace(" ", "") == corrected.replace(" ", ""):
        return "whitespace normalization"
    if raw.isdigit() and ":" in corrected:
        return "time format reconstruction"
    return "parser correction"


def _write_combined_summary(
    ws, runs: list[tuple[RunMetrics, list[PageMetrics], list[RowMetrics]]]
) -> None:
    """Write a paper-ready combined summary table.

    Rows = metrics, columns = files + Combined total.
    """
    # Build metric rows
    metrics: list[tuple[str, str, list[float | int | str]]] = []

    file_names = []
    for run, pages, rows in runs:
        file_names.append(run.source_file)

    # Header row
    headers = ["Metric"] + file_names + ["Combined"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    _style_header_row(ws, 1, len(headers))

    # 1. Model Type
    ws.cell(row=2, column=1, value="Model Type")
    for i, (run, _, _) in enumerate(runs, 2):
        ws.cell(row=2, column=i, value=run.model_type)
    ws.cell(row=2, column=len(runs) + 2, value="—")

    # 2. Device
    ws.cell(row=3, column=1, value="Device")
    for i, (run, _, _) in enumerate(runs, 2):
        ws.cell(row=3, column=i, value=run.device)
    ws.cell(row=3, column=len(runs) + 2, value="—")

    # 3. Pages
    ws.cell(row=4, column=1, value="Pages")
    for i, (run, _, _) in enumerate(runs, 2):
        ws.cell(row=4, column=i, value=run.num_pages)
    ws.cell(row=4, column=len(runs) + 2, value=sum(run.num_pages for run, _, _ in runs))

    # 4. Total Time (s)
    ws.cell(row=5, column=1, value="Total Time (s)")
    for i, (run, _, _) in enumerate(runs, 2):
        ws.cell(row=5, column=i, value=round(run.total_time_s, 2))
    ws.cell(
        row=5,
        column=len(runs) + 2,
        value=round(sum(run.total_time_s for run, _, _ in runs), 2),
    )

    # 5. Avg Page Time (s)
    ws.cell(row=6, column=1, value="Avg Page Time (s)")
    for i, (run, _, _) in enumerate(runs, 2):
        ws.cell(row=6, column=i, value=round(run.avg_page_time_s, 2))
    ws.cell(
        row=6,
        column=len(runs) + 2,
        value=round(
            sum(run.total_time_s for run, _, _ in runs)
            / sum(run.num_pages for run, _, _ in runs)
            if sum(run.num_pages for run, _, _ in runs)
            else 0,
            2,
        ),
    )

    # 6. OCR Init Time (s)
    ws.cell(row=7, column=1, value="OCR Init Time (s)")
    for i, (_, pages, _) in enumerate(runs, 2):
        init = next((p.ocr_init_time_s for p in pages if p.ocr_init_time_s > 0), 0)
        ws.cell(row=7, column=i, value=round(init, 2))
    ws.cell(
        row=7,
        column=len(runs) + 2,
        value=round(
            max(
                next((p.ocr_init_time_s for p in pages if p.ocr_init_time_s > 0), 0)
                for _, pages, _ in runs
            ),
            2,
        ),
    )

    # 7. OCR Inference Time (s)
    ws.cell(row=8, column=1, value="OCR Inference Time (s)")
    for i, (_, pages, _) in enumerate(runs, 2):
        ws.cell(
            row=8, column=i, value=round(sum(p.ocr_inference_time_s for p in pages), 2)
        )
    ws.cell(
        row=8,
        column=len(runs) + 2,
        value=round(
            sum(sum(p.ocr_inference_time_s for p in pages) for _, pages, _ in runs), 2
        ),
    )

    # 8. Layout Detection Time (s)
    ws.cell(row=9, column=1, value="Layout Detection Time (s)")
    for i, (_, pages, _) in enumerate(runs, 2):
        ws.cell(
            row=9,
            column=i,
            value=round(sum(p.layout_detection_time_s for p in pages), 2),
        )
    ws.cell(
        row=9,
        column=len(runs) + 2,
        value=round(
            sum(sum(p.layout_detection_time_s for p in pages) for _, pages, _ in runs),
            2,
        ),
    )

    # 9. Extraction Time (s)
    ws.cell(row=10, column=1, value="Extraction Time (s)")
    for i, (_, pages, _) in enumerate(runs, 2):
        ws.cell(
            row=10, column=i, value=round(sum(p.extraction_time_s for p in pages), 2)
        )
    ws.cell(
        row=10,
        column=len(runs) + 2,
        value=round(
            sum(sum(p.extraction_time_s for p in pages) for _, pages, _ in runs), 2
        ),
    )

    # 10. Validation Time (s)
    ws.cell(row=11, column=1, value="Validation Time (s)")
    for i, (_, pages, _) in enumerate(runs, 2):
        ws.cell(
            row=11, column=i, value=round(sum(p.validation_time_s for p in pages), 2)
        )
    ws.cell(
        row=11,
        column=len(runs) + 2,
        value=round(
            sum(sum(p.validation_time_s for p in pages) for _, pages, _ in runs), 2
        ),
    )

    # 11. Total OCR Boxes
    ws.cell(row=12, column=1, value="Total OCR Boxes Detected")
    for i, (_, pages, _) in enumerate(runs, 2):
        ws.cell(row=12, column=i, value=sum(p.total_boxes_detected for p in pages))
    ws.cell(
        row=12,
        column=len(runs) + 2,
        value=sum(sum(p.total_boxes_detected for p in pages) for _, pages, _ in runs),
    )

    # 12. VLM Fallbacks
    ws.cell(row=13, column=1, value="VLM Fallbacks Triggered")
    for i, (_, pages, _) in enumerate(runs, 2):
        ws.cell(row=13, column=i, value=sum(p.vlm_fallbacks for p in pages))
    ws.cell(
        row=13,
        column=len(runs) + 2,
        value=sum(sum(p.vlm_fallbacks for p in pages) for _, pages, _ in runs),
    )

    # 13. Rows Extracted
    ws.cell(row=14, column=1, value="Rows Extracted")
    for i, (run, _, _) in enumerate(runs, 2):
        ws.cell(row=14, column=i, value=run.total_rows_extracted)
    ws.cell(
        row=14,
        column=len(runs) + 2,
        value=sum(run.total_rows_extracted for run, _, _ in runs),
    )

    # 14. Accepted
    ws.cell(row=15, column=1, value="Accepted")
    for i, (run, _, _) in enumerate(runs, 2):
        ws.cell(row=15, column=i, value=run.accepted_count)
    ws.cell(
        row=15,
        column=len(runs) + 2,
        value=sum(run.accepted_count for run, _, _ in runs),
    )

    # 15. Flagged
    ws.cell(row=16, column=1, value="Flagged")
    for i, (run, _, _) in enumerate(runs, 2):
        ws.cell(row=16, column=i, value=run.flagged_count)
    ws.cell(
        row=16, column=len(runs) + 2, value=sum(run.flagged_count for run, _, _ in runs)
    )

    # 16. Failed
    ws.cell(row=17, column=1, value="Failed")
    for i, (run, _, _) in enumerate(runs, 2):
        ws.cell(row=17, column=i, value=run.failed_count)
    ws.cell(
        row=17, column=len(runs) + 2, value=sum(run.failed_count for run, _, _ in runs)
    )

    # 17. Acceptance Rate
    ws.cell(row=18, column=1, value="Acceptance Rate")
    for i, (run, _, _) in enumerate(runs, 2):
        rate = run.accepted_count / max(run.total_rows_extracted, 1) * 100
        ws.cell(row=18, column=i, value=f"{rate:.1f}%")
    total_acc = sum(run.accepted_count for run, _, _ in runs)
    total_rows = sum(run.total_rows_extracted for run, _, _ in runs)
    ws.cell(
        row=18,
        column=len(runs) + 2,
        value=f"{total_acc / max(total_rows, 1) * 100:.1f}%",
    )

    # 18. Mean Confidence
    ws.cell(row=19, column=1, value="Mean Confidence")
    for i, (run, _, _) in enumerate(runs, 2):
        ws.cell(row=19, column=i, value=round(run.mean_overall_confidence, 4))
    ws.cell(
        row=19,
        column=len(runs) + 2,
        value=round(
            sum(run.mean_overall_confidence for run, _, _ in runs) / len(runs), 4
        ),
    )

    # 19. Min Confidence
    ws.cell(row=20, column=1, value="Min Confidence")
    for i, (run, _, _) in enumerate(runs, 2):
        ws.cell(row=20, column=i, value=round(run.min_overall_confidence, 4))
    ws.cell(
        row=20,
        column=len(runs) + 2,
        value=round(min(run.min_overall_confidence for run, _, _ in runs), 4),
    )

    # 20. Max Confidence
    ws.cell(row=21, column=1, value="Max Confidence")
    for i, (run, _, _) in enumerate(runs, 2):
        ws.cell(row=21, column=i, value=round(run.max_overall_confidence, 4))
    ws.cell(
        row=21,
        column=len(runs) + 2,
        value=round(max(run.max_overall_confidence for run, _, _ in runs), 4),
    )

    # 21. Mean CER
    ws.cell(row=22, column=1, value="Mean Character Error Rate")
    for i, (run, _, _) in enumerate(runs, 2):
        ws.cell(row=22, column=i, value=round(run.mean_cer, 4))
    ws.cell(
        row=22,
        column=len(runs) + 2,
        value=round(sum(run.mean_cer for run, _, _ in runs) / len(runs), 4),
    )

    # 22. Hours Mismatch Rate
    ws.cell(row=23, column=1, value="Hours Mismatch Rate")
    for i, (run, _, _) in enumerate(runs, 2):
        ws.cell(row=23, column=i, value=f"{run.hours_mismatch_rate * 100:.1f}%")
    # Weighted average
    total_with_hours = 0
    total_mismatches = 0
    for _, _, rows in runs:
        rows_with_hours = [
            r for r in rows if r.calculated_hours is not None and r.total_hours_text
        ]
        total_with_hours += len(rows_with_hours)
        total_mismatches += sum(
            1 for r in rows_with_hours if "hours_mismatch" in r.validation_errors
        )
    ws.cell(
        row=23,
        column=len(runs) + 2,
        value=f"{total_mismatches / max(total_with_hours, 1) * 100:.1f}%",
    )

    # 23. Field Missing Rate
    ws.cell(row=24, column=1, value="Field Missing Rate")
    for i, (run, _, _) in enumerate(runs, 2):
        ws.cell(row=24, column=i, value=f"{run.field_missing_rate * 100:.1f}%")
    total_missing = sum(
        sum(
            1
            for r in rows
            if not r.parsed_date or not r.parsed_time_in or not r.parsed_time_out
        )
        for _, _, rows in runs
    )
    total_rows_all = sum(len(rows) for _, _, rows in runs)
    ws.cell(
        row=24,
        column=len(runs) + 2,
        value=f"{total_missing / max(total_rows_all, 1) * 100:.1f}%",
    )

    # Style the metric column
    metric_font = Font(name="Calibri", bold=True, size=11)
    for row_idx in range(2, 25):
        ws.cell(row=row_idx, column=1).font = metric_font

    # Auto-fit
    ws.column_dimensions["A"].width = 35
    for col_idx in range(2, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(2, 25):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        from openpyxl.utils import get_column_letter

        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)
