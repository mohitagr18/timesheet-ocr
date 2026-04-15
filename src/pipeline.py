"""Pipeline orchestrator — ties all components together for end-to-end extraction."""

from __future__ import annotations

import json
import logging
import shutil
import subprocess
import sys
import time as time_module
from pathlib import Path
from typing import TYPE_CHECKING

import re
import numpy as np

from .benchmark import (
    BenchmarkCollector,
    PageMetrics,
    RowMetrics,
)
from .confidence import (
    Route,
    boxes_in_zone,
    route_by_confidence,
    should_fallback_entire_row,
)
from .debug_viz import VlmFallbackCell
from .exporter import export_results
from .layout import detect_layout
from .layout_model import DocLayoutDetector
from .models import (
    ExtractionResult,
    OcrSource,
    RowStatus,
    TimesheetRecord,
    TimesheetRow,
)
from .ocr_engine import OcrEngine
from .parser import (
    clean_name,
    disambiguate_times,
    extract_expected_year,
    extract_week_dates,
    parse_date,
    parse_hours,
    parse_time,
)
from .phi import PhiAnonymizer
from .preprocessing import load_image, pdf_to_images, preprocess_image
from .review_queue import build_review_queue
from .validation import validate_record
from .vlm_fallback import VlmFallback
from .vlm_cloud import CloudVlmExtractor
from .band_crop_extractor import BandCropExtractor

if TYPE_CHECKING:
    from .config import AppConfig

logger = logging.getLogger(__name__)


class Pipeline:
    """End-to-end extraction pipeline for timesheet documents.

    Orchestrates: load → preprocess → layout → OCR → fallback → parse → validate → export.
    """

    def __init__(self, config: AppConfig) -> None:
        self.config = config
        self.ocr_engine = OcrEngine(config)
        self.vlm = VlmFallback(config)
        self.layout_detector = DocLayoutDetector(config)
        self.cloud_vlm = CloudVlmExtractor(config)
        self.band_crop_extractor = BandCropExtractor(config)
        self.benchmark = BenchmarkCollector()
        self._ocr_init_time = 0.0
        self.name_db = None

    def _init_name_mapping(
        self, anonymizer: PhiAnonymizer, source_files: list[str]
    ) -> None:
        """Persist name mappings to local SQLite DB."""
        from .name_mapping import NameMappingDB

        output_dir = Path(self.config.output_path)
        output_dir.mkdir(parents=True, exist_ok=True)
        db_path = output_dir / "name_mapping.db"
        self.name_db = NameMappingDB(db_path)

        for sf in source_files:
            for real, anon_id in anonymizer._patient_map.items():
                self.name_db.upsert_patient(anon_id, real, sf)
            for real, anon_id in anonymizer._employee_map.items():
                self.name_db.upsert_employee(anon_id, real, sf)

    def cleanup(self) -> None:
        """Release memory-heavy model instances to prevent OOM on long runs."""
        import gc

        logger.info("Releasing model instances from memory...")

        if hasattr(self.ocr_engine, "_ocr"):
            del self.ocr_engine._ocr
            self.ocr_engine._initialized = False

        if hasattr(self.layout_detector, "_model"):
            del self.layout_detector._model
            self.layout_detector._initialized = False

        if hasattr(self.vlm, "_client"):
            del self.vlm._client
            self.vlm._initialized = False
            self.vlm._available = None

        gc.collect()
        logger.info("Model instances released.")

    def process_file(
        self, file_path: str | Path, anonymizer: PhiAnonymizer | None = None
    ) -> ExtractionResult:
        """Process a single file (PDF or image) through the full pipeline."""
        file_path = Path(file_path)
        start_time = time_module.time()

        if anonymizer is None:
            anonymizer = PhiAnonymizer([file_path.name])

        anon_filename = anonymizer.anonymize_filename(file_path.name)

        logger.info(f"{'=' * 60}")
        logger.info(f"Processing: {file_path.name}")
        logger.info(f"{'=' * 60}")

        # Initialize benchmark with anonymized filename
        file_size_kb = file_path.stat().st_size / 1024 if file_path.exists() else 0
        model_type = "PP-OCRv5_mobile"
        extraction_mode = getattr(self.config, "extraction_mode", "ppocr_grid")
        if extraction_mode == "vlm_full_page":
            model_type = "qwen3-vl:8b"
        elif extraction_mode == "band_crop_vlm_cloud":
            model_type = "gemini-band-crop"
        device = self.config.ppocr.device
        image_dpi = self.config.preprocessing.target_dpi
        self.benchmark.start_run(
            source_file=anon_filename,
            model_type=model_type,
            device=device,
            image_dpi=image_dpi,
            file_size_kb=file_size_kb,
            extraction_mode=extraction_mode,
        )
        self._ocr_init_time = 0.0
        self.vlm.reset_stats()

        # 1. Load file → list of images (one per page)
        images = self._load_file(file_path)

        # 2. Process each page
        records: list[TimesheetRecord] = []
        for page_idx, image in enumerate(images):
            logger.info(f"\n--- Page {page_idx + 1}/{len(images)} ---")

            page_start_time = time_module.time()
            record, page_bench = self._process_page(
                image, file_path.name, anon_filename, page_idx + 1, anonymizer
            )
            page_elapsed = time_module.time() - page_start_time
            page_bench.page_time_s = page_elapsed
            page_bench.image_width = image.shape[1]
            page_bench.image_height = image.shape[0]
            self.benchmark.add_page(page_bench)

            logger.info(
                f"Page {page_idx + 1} processing complete in {page_elapsed:.1f}s"
            )
            records.append(record)

        # 2.5 Aggregate Document-Level Metadata
        # (E.g. Page 1 has shifts/patient name, Page 2 has the employee signature)
        best_emp = (
            max(
                records,
                key=lambda r: (r.employee_name_confidence, len(r.employee_name)),
            )
            if records
            else None
        )
        best_pat = (
            max(records, key=lambda r: (r.patient_name_confidence, len(r.patient_name)))
            if records
            else None
        )

        for r in records:
            if not r.employee_name and best_emp and best_emp.employee_name:
                r.employee_name = best_emp.employee_name
                r.employee_name_confidence = best_emp.employee_name_confidence
            if not r.patient_name and best_pat and best_pat.patient_name:
                r.patient_name = best_pat.patient_name
                r.patient_name_confidence = best_pat.patient_name_confidence

        # Collect row-level benchmark metrics
        for record in records:
            for row in record.rows:
                self._collect_row_metrics(row, record)

        # 3. Build result
        elapsed = time_module.time() - start_time
        result = ExtractionResult(
            source_file=anon_filename,
            processing_time_seconds=elapsed,
            total_pages=len(images),
            records=records,
        )

        # 4. Build review queue
        result.review_items = build_review_queue(result, self.config)

        # 5. Export
        output_files = export_results(result, self.config)

        # 5.5 Export benchmark
        self.benchmark.finalize(elapsed)
        self.benchmark.run.image_dimensions = (
            f"{self.benchmark.pages[0].image_width}x{self.benchmark.pages[0].image_height}"
            if self.benchmark.pages
            else ""
        )
        bench_file = self.benchmark.export(self.config.output_path)
        output_files.append(bench_file)

        # Snapshot for combined export
        extraction_mode = getattr(self.config, "extraction_mode", "ppocr_grid")
        if extraction_mode == "vlm_full_page":
            mode_label = "vlm"
        elif extraction_mode == "ocr_only":
            mode_label = "ocr_only"
        elif extraction_mode == "band_crop_vlm_cloud":
            mode_label = "band_crop"
        else:
            mode_label = "ppocr"
        self.benchmark.snapshot_run(mode=mode_label)

        # 6. Summary
        logger.info(f"\n{'=' * 60}")
        logger.info(f"DONE: {file_path.name}")
        logger.info(f"  Pages: {result.total_pages}")
        logger.info(
            f"  Rows:  {result.total_rows} (accepted={result.accepted_count}, flagged={result.flagged_count}, failed={result.failed_count})"
        )
        logger.info(f"  Time:  {elapsed:.1f}s")
        logger.info(f"  Files: {[f.name for f in output_files]}")
        logger.info(f"{'=' * 60}")

        return result

    def process_directory(
        self, input_dir: str | Path | None = None, generate_combined: bool = True,
        files_to_process: list[Path] | None = None,
    ) -> list[ExtractionResult]:
        """Process all supported files in a directory.

        Args:
            input_dir: Directory containing input files.
            generate_combined: If True, generate combined comparison results
                after processing. Set to False when running as part of a
                multi-approach benchmark (run_all_approaches_safe.py).
            files_to_process: If provided, only process these specific Path objects
                instead of scanning the entire input directory. Used for resume logic.
        """
        if input_dir is None:
            input_dir = self.config.input_path
        else:
            input_dir = Path(input_dir)

        if not input_dir.exists():
            logger.error(f"Input directory not found: {input_dir}")
            return []

        # Use provided file list or scan the input directory
        if files_to_process is not None:
            files = sorted(files_to_process)
        else:
            # Find all supported files
            supported = {".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp"}
            files = sorted(
                f
                for f in input_dir.iterdir()
                if f.suffix.lower() in supported and not f.name.startswith(".")
            )

        if not files:
            logger.warning(f"No supported files found in {input_dir}")
            return []

        # Find already processed files in the merged Excel sheet to skip them.
        # Only used when files_to_process is NOT provided (i.e., full directory scan).
        processed_files: set[str] = set()
        if files_to_process is None:
            xlsx_path = self.config.output_path / "merged_results.xlsx"
            if xlsx_path.exists():
                try:
                    from openpyxl import load_workbook

                    wb = load_workbook(xlsx_path, read_only=True)
                    if self.config.export.excel_sheet_name in wb.sheetnames:
                        ws = wb[self.config.export.excel_sheet_name]
                    else:
                        ws = wb.active

                    # Source File is column A (1-indexed)
                    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                        if row[0]:
                            processed_files.add(str(row[0]))
                    if processed_files:
                        logger.info(
                            f"Found {len(processed_files)} previously processed file(s) in Excel."
                        )
                except Exception as e:
                    logger.warning(
                        f"Could not read existing Excel to find processed files: {e}"
                    )

        logger.info(f"Found {len(files)} file(s) in input directory")

        # Create PHI anonymizer with all filenames for deterministic mapping
        filenames = [f.name for f in files]
        anonymizer = PhiAnonymizer(filenames)

        # Persist name mappings to local SQLite DB
        self._init_name_mapping(anonymizer, filenames)

        # Create temp directory for subprocess results
        tmp_dir = self.config.output_path / ".tmp"
        tmp_dir.mkdir(parents=True, exist_ok=True)

        results = []
        for file_idx, file_path in enumerate(files):
            if file_path.name in processed_files:
                logger.info(
                    f"Skipping {file_path.name}: already processed and in Excel."
                )
                continue

            file_start = time_module.time()
            try:
                # Spawn subprocess to process this file in isolation
                # This ensures PaddleOCR memory is fully released after each file
                cmd = [
                    sys.executable, "-m", "src.process_single",
                    str(file_path),
                    str(self.config.output_path),
                ]
                # Pass config path if it's not the default
                config_file = Path("config.yaml")
                if config_file.exists():
                    cmd.append(str(config_file))

                # Pass all filenames for consistent anonymization mapping
                cmd.extend(["--filenames", json.dumps(filenames)])

                logger.info(f"▶ Processing {file_path.name} (subprocess #{file_idx + 1})...")
                proc = subprocess.run(
                    cmd,
                    capture_output=True,
                    text=True,
                    timeout=3600,  # 60 min timeout per file
                )

                # Read result from temp JSON
                result_file = tmp_dir / f"{file_path.stem}.result.json"
                error_file = tmp_dir / f"{file_path.stem}.error.json"

                if proc.returncode == 0 and result_file.exists():
                    result_data = json.loads(result_file.read_text())
                    # Reconstruct a minimal ExtractionResult-like dict
                    # (Full records are already exported by the subprocess)
                    results.append(result_data)
                    logger.info(
                        f"✓ {file_path.name}: {result_data['total_rows']} rows "
                        f"in {result_data['processing_time_seconds']:.1f}s"
                    )
                else:
                    error_msg = proc.stderr.strip()
                    if error_file.exists():
                        error_data = json.loads(error_file.read_text())
                        error_msg = error_data.get("error", proc.stderr)
                    logger.error(f"✗ {file_path.name}: {error_msg}")

            except subprocess.TimeoutExpired:
                logger.error(f"✗ {file_path.name}: Timed out after 60 minutes")
            except Exception as e:
                logger.error(f"✗ {file_path.name}: {e}", exc_info=True)

            # Rate-limit delay for cloud approach
            extraction_mode = getattr(self.config, "extraction_mode", "")
            if extraction_mode in ("layout_guided_vlm_cloud", "band_crop_vlm_cloud") and file_idx < len(files) - 1:
                delay = getattr(self.config.cloud_vlm, "inter_file_delay", 5)
                if delay > 0:
                    logger.info(f"⏳ Rate-limit delay: {delay}s before next file...")
                    time_module.sleep(delay)

        # Clean up temp directory
        if tmp_dir.exists():
            shutil.rmtree(tmp_dir)

        # Export combined benchmark across all files
        if results and len(self.benchmark._runs) > 1:
            combined_file = self.benchmark.export_combined(
                self.config.output_path, anonymizer
            )
            logger.info(f"Combined benchmark exported: {combined_file}")

        # Generate combined comparison and ground truth analysis
        if results and generate_combined:
            self._generate_combined_results()

        return results

    def _generate_combined_results(self) -> None:
        """Generate combined comparison and ground truth analysis automatically."""
        project_root = Path(__file__).resolve().parent.parent
        combined_script = project_root / "scripts" / "create_combined_results.py"
        consensus_script = project_root / "scripts" / "create_consensus_results.py"

        if combined_script.exists():
            logger.info("Generating combined comparison and ground truth analysis...")
            try:
                result = subprocess.run(
                    [sys.executable, str(combined_script)],
                    cwd=project_root,
                    capture_output=True,
                    text=True,
                    timeout=120,
                )
                if result.stdout:
                    for line in result.stdout.strip().split("\n"):
                        logger.info(f"  {line}")
                if result.stderr:
                    for line in result.stderr.strip().split("\n"):
                        logger.warning(f"  {line}")
                bench_file = self.config.output_path / "combined" / "benchmark_combined.xlsx"
                if bench_file.exists():
                    logger.info(f"Combined comparison exported: {bench_file}")
            except subprocess.TimeoutExpired:
                logger.error("Combined results generation timed out")
            except Exception as e:
                logger.error(f"Failed to generate combined results: {e}")
        else:
            logger.debug(f"Combined results script not found: {combined_script}")

        if consensus_script.exists():
            logger.info("Appending KPI/consensus sheets to benchmark_combined...")
            try:
                result = subprocess.run(
                    [sys.executable, str(consensus_script)],
                    cwd=project_root,
                    capture_output=True,
                    text=True,
                    timeout=120,
                )
                if result.stdout:
                    for line in result.stdout.strip().split("\n"):
                        logger.info(f"  {line}")
                if result.stderr:
                    for line in result.stderr.strip().split("\n"):
                        logger.warning(f"  {line}")
                bench_file = self.config.output_path / "combined" / "benchmark_combined.xlsx"
                if bench_file.exists():
                    logger.info(f"KPI/consensus sheets appended to: {bench_file}")
            except subprocess.TimeoutExpired:
                logger.error("Consensus results generation timed out")
            except Exception as e:
                logger.error(f"Failed to generate consensus results: {e}")
        else:
            logger.debug(f"Consensus results script not found: {consensus_script}")

    def _load_file(self, file_path: Path) -> list[np.ndarray]:
        """Load a file as a list of images (handles both PDFs and images)."""
        if file_path.suffix.lower() == ".pdf":
            return pdf_to_images(file_path, dpi=self.config.preprocessing.target_dpi)
        else:
            return [load_image(file_path)]

    def _process_page(
        self,
        image: np.ndarray,
        source_file: str,
        anon_filename: str,
        page_number: int,
        anonymizer: PhiAnonymizer,
    ) -> tuple[TimesheetRecord, PageMetrics]:
        """Process a single page image through OCR + validation."""
        if getattr(self.config, "debug", None):
            # Only generate debug images for this exact file
            self.config.debug.visualize_ocr = (source_file == "N.Rivera Timesheets -030426-031026.pdf")

        page_metrics = PageMetrics(source_file=anon_filename, page_number=page_number)

        preprocess_start = time_module.time()
        preprocessed = preprocess_image(image, self.config)
        preprocess_time = time_module.time() - preprocess_start

        # Extract patient name from the original filename
        patient_name = source_file.split("Timesheet")[0].strip(" -_")
        patient_conf = 1.0

        # Extract expected year from filename for date parsing
        expected_year = extract_expected_year(source_file)
        if expected_year:
            logger.info(f"Expected year from filename: {expected_year}")

        employee_name = ""
        employee_conf = 0.0
        rows = []

        if getattr(self.config, "extraction_mode", "ppocr_grid") == "vlm_full_page":
            # Quick OCR pass for page classification (grid vs signature)
            ocr_start = time_module.time()
            if not self.ocr_engine._initialized:
                init_start = time_module.time()
                self.ocr_engine._ensure_initialized()
                self._ocr_init_time = time_module.time() - init_start
                page_metrics.ocr_init_time_s = self._ocr_init_time
            ocr_result = self.ocr_engine.run(preprocessed)
            page_metrics.ocr_inference_time_s = time_module.time() - ocr_start
            page_metrics.total_boxes_detected = len(ocr_result.boxes)

            # Detect layout for zone coordinates
            layout_start = time_module.time()
            layout = detect_layout(image, self.config, ocr_result.boxes)
            page_metrics.layout_detection_time_s = time_module.time() - layout_start

            sig_threshold = self.config.debug.signature_ocr_threshold
            is_sig_page = PhiAnonymizer.is_signature_page(
                len(ocr_result.boxes), sig_threshold
            )

            if is_sig_page:
                # Signature page: extract employee name from footer zone, skip VLM
                logger.info(
                    f"Page {page_number} is a signature page ({len(ocr_result.boxes)} boxes < {sig_threshold} threshold). Skipping VLM extraction."
                )
                footer_boxes = boxes_in_zone(
                    ocr_result.boxes,
                    layout.footer_zone.x_start,
                    layout.footer_zone.y_start,
                    layout.footer_zone.x_end,
                    layout.footer_zone.y_end,
                )
                if footer_boxes:
                    sorted_footer = sorted(
                        footer_boxes, key=lambda b: (b.y_center, b.x_center)
                    )
                    for box in sorted_footer:
                        text = clean_name(box.text)
                        if text and len(text) > 2:
                            employee_name = text
                            employee_conf = box.confidence
                            break
            else:
                # Grid page: proceed with VLM full page extraction
                vlm_start = time_module.time()
                vlm_results = self.vlm.extract_full_page(image)
                page_metrics.vlm_inference_time_s = time_module.time() - vlm_start

                employee_name = vlm_results.get("rn_lpn_name", "")
                employee_conf = 0.9 if employee_name else 0.0

                valid_row_idx = 0
                shifts_data = vlm_results.get("shifts", [])
                if len(shifts_data) > 7:
                    logger.warning(
                        f"Page {page_number} VLM hallucinated {len(shifts_data)} rows! Discarding fake table."
                    )
                    shifts_data = []

                for row_data in shifts_data:
                    date_text = row_data.get("date", "").strip()
                    time_in_text = row_data.get("time_in", "").strip()
                    time_out_text = row_data.get("time_out", "").strip()
                    hours_text = row_data.get("total_hours", "").strip()

                    def is_noise(val: str) -> bool:
                        return len(re.sub(r"[^a-zA-Z0-9]", "", val)) == 0

                    if (
                        is_noise(time_in_text)
                        and is_noise(time_out_text)
                        and is_noise(hours_text)
                    ):
                        continue

                    time_in_parsed, time_out_parsed = disambiguate_times(
                        time_in_text, time_out_text, hours_text
                    )

                    rows.append(
                        TimesheetRow(
                            row_index=valid_row_idx,
                            date_text=date_text,
                            date_parsed=parse_date(date_text, expected_year),
                            time_in_text=time_in_text,
                            time_in_parsed=time_in_parsed,
                            time_out_text=time_out_text,
                            time_out_parsed=time_out_parsed,
                            total_hours_text=hours_text,
                            total_hours_parsed=parse_hours(hours_text),
                            date_confidence=0.9,
                            time_in_confidence=0.9,
                            time_out_confidence=0.9,
                            hours_confidence=0.9,
                            date_source=OcrSource.VLM,
                            time_in_source=OcrSource.VLM,
                            time_out_source=OcrSource.VLM,
                            hours_source=OcrSource.VLM,
                        )
                    )
                    valid_row_idx += 1

                # VLM debug visualization (if enabled, grid pages only)
                if (
                    getattr(self.config, "debug", None)
                    and self.config.debug.visualize_ocr
                ):
                    from . import vlm_debug_viz

                    current_record = TimesheetRecord(
                        source_file=anon_filename,
                        page_number=page_number,
                        employee_name=anonymizer.anonymize_employee(employee_name),
                        employee_name_confidence=employee_conf,
                        patient_name=anonymizer.anonymize_patient(patient_name),
                        patient_name_confidence=patient_conf,
                        rows=rows,
                    )
                    vlm_debug_viz.render_vlm_page(
                        image=image,
                        records=[current_record],
                        page_number=page_number,
                        source_file=anon_filename,
                        output_dir=self.config.debug_output_path,
                    )

        elif getattr(self.config, "extraction_mode", "ppocr_grid") in (
            "layout_guided_vlm_local",
            "layout_guided_vlm_cloud",
        ):
            # Quick OCR pass for page classification (grid vs signature)
            ocr_start = time_module.time()
            if not self.ocr_engine._initialized:
                init_start = time_module.time()
                self.ocr_engine._ensure_initialized()
                self._ocr_init_time = time_module.time() - init_start
                page_metrics.ocr_init_time_s = self._ocr_init_time
            ocr_result = self.ocr_engine.run(preprocessed)
            page_metrics.ocr_inference_time_s = time_module.time() - ocr_start
            page_metrics.total_boxes_detected = len(ocr_result.boxes)

            # Detect layout for zone coordinates
            layout_start = time_module.time()
            layout = detect_layout(image, self.config, ocr_result.boxes)
            page_metrics.layout_detection_time_s = time_module.time() - layout_start

            sig_threshold = self.config.debug.signature_ocr_threshold
            is_sig_page = PhiAnonymizer.is_signature_page(
                len(ocr_result.boxes), sig_threshold
            )

            if is_sig_page:
                logger.info(
                    f"Page {page_number} is a signature page ({len(ocr_result.boxes)} boxes < {sig_threshold} threshold). Skipping VLM extraction."
                )
                footer_boxes = boxes_in_zone(
                    ocr_result.boxes,
                    layout.footer_zone.x_start,
                    layout.footer_zone.y_start,
                    layout.footer_zone.x_end,
                    layout.footer_zone.y_end,
                )
                if footer_boxes:
                    sorted_footer = sorted(
                        footer_boxes, key=lambda b: (b.y_center, b.x_center)
                    )
                    for box in sorted_footer:
                        text = clean_name(box.text)
                        if text and len(text) > 2:
                            employee_name = text
                            employee_conf = box.confidence
                            break
            else:
                # Grid page: detect table with PP-DocLayoutV3, crop, send to VLM
                doclayout_start = time_module.time()
                table_zone = self.layout_detector.detect_table(image)
                page_metrics.layout_detection_time_s = (
                    time_module.time() - doclayout_start
                )

                if table_zone is not None:
                    table_crop = table_zone.crop(image)
                    logger.info(
                        f"Page {page_number}: Table detected, cropping with "
                        f"+{self.layout_detector.TABLE_PADDING}px padding"
                    )
                else:
                    table_crop = image
                    logger.warning(
                        f"Page {page_number}: No table detected by PP-DocLayoutV3, "
                        f"using full page"
                    )

                # VLM extraction on cropped table (local or cloud)
                vlm_start = time_module.time()
                if (
                    getattr(self.config, "extraction_mode", "ppocr_grid")
                    == "layout_guided_vlm_local"
                ):
                    vlm_results = self.vlm.extract_table_crop(table_crop)
                else:
                    vlm_results = self.cloud_vlm.extract_table_crop(table_crop)
                page_metrics.vlm_inference_time_s = time_module.time() - vlm_start

                employee_name = vlm_results.get("rn_lpn_name", "")
                employee_conf = 0.9 if employee_name else 0.0

                valid_row_idx = 0
                shifts_data = vlm_results.get("shifts", [])
                if len(shifts_data) > 7:
                    logger.warning(
                        f"Page {page_number} VLM hallucinated {len(shifts_data)} rows! Discarding fake table."
                    )
                    shifts_data = []

                for row_data in shifts_data:
                    date_text = row_data.get("date", "").strip()
                    time_in_text = row_data.get("time_in", "").strip()
                    time_out_text = row_data.get("time_out", "").strip()
                    hours_text = row_data.get("total_hours", "").strip()

                    def is_noise(val: str) -> bool:
                        return len(re.sub(r"[^a-zA-Z0-9]", "", val)) == 0

                    if (
                        is_noise(time_in_text)
                        and is_noise(time_out_text)
                        and is_noise(hours_text)
                    ):
                        continue

                    time_in_parsed, time_out_parsed = disambiguate_times(
                        time_in_text, time_out_text, hours_text
                    )

                    rows.append(
                        TimesheetRow(
                            row_index=valid_row_idx,
                            date_text=date_text,
                            date_parsed=parse_date(date_text, expected_year),
                            time_in_text=time_in_text,
                            time_in_parsed=time_in_parsed,
                            time_out_text=time_out_text,
                            time_out_parsed=time_out_parsed,
                            total_hours_text=hours_text,
                            total_hours_parsed=parse_hours(hours_text),
                            date_confidence=0.9,
                            time_in_confidence=0.9,
                            time_out_confidence=0.9,
                            hours_confidence=0.9,
                            date_source=OcrSource.VLM,
                            time_in_source=OcrSource.VLM,
                            time_out_source=OcrSource.VLM,
                            hours_source=OcrSource.VLM,
                        )
                    )
                    valid_row_idx += 1

                # VLM debug visualization (if enabled)
                if (
                    getattr(self.config, "debug", None)
                    and self.config.debug.visualize_ocr
                ):
                    from . import vlm_debug_viz

                    current_record = TimesheetRecord(
                        source_file=anon_filename,
                        page_number=page_number,
                        employee_name=anonymizer.anonymize_employee(employee_name),
                        employee_name_confidence=employee_conf,
                        patient_name=anonymizer.anonymize_patient(patient_name),
                        patient_name_confidence=patient_conf,
                        rows=rows,
                    )
                    vlm_debug_viz.render_vlm_page(
                        image=image,
                        records=[current_record],
                        page_number=page_number,
                        source_file=anon_filename,
                        output_dir=self.config.debug_output_path,
                    )

        elif getattr(self.config, "extraction_mode", "ppocr_grid") == "band_crop_vlm_cloud":
            # Band-crop approach: extract DATE row + footer block, send to Gemini
            band_start = time_module.time()
            payload_image, is_sig_page = self.band_crop_extractor.build_phi_safe_payload(
                image
            )
            page_metrics.layout_detection_time_s = time_module.time() - band_start

            if is_sig_page or payload_image is None:
                logger.info(
                    f"Page {page_number} identified as signature page by band_crop. "
                    f"Skipping extraction."
                )
            else:
                # Send stitched payload to Gemini
                vlm_start = time_module.time()
                vlm_results = self.cloud_vlm.extract_table_crop(payload_image)
                page_metrics.vlm_inference_time_s = time_module.time() - vlm_start

                employee_name = vlm_results.get("rn_lpn_name", "")
                employee_conf = 0.9 if employee_name else 0.0

                valid_row_idx = 0
                shifts_data = vlm_results.get("shifts", [])

                if len(shifts_data) > 7:
                    logger.warning(
                        f"Page {page_number} band_crop VLM hallucinated "
                        f"{len(shifts_data)} rows! Discarding."
                    )
                    shifts_data = []

                for row_data in shifts_data:
                    date_text = row_data.get("date", "").strip()
                    time_in_text = row_data.get("time_in", "").strip()
                    time_out_text = row_data.get("time_out", "").strip()
                    hours_text = row_data.get("total_hours", "").strip()

                    def is_noise(val: str) -> bool:
                        return len(re.sub(r"[^a-zA-Z0-9]", "", val)) == 0

                    if (
                        is_noise(time_in_text)
                        and is_noise(time_out_text)
                        and is_noise(hours_text)
                    ):
                        continue

                    time_in_parsed, time_out_parsed = disambiguate_times(
                        time_in_text, time_out_text, hours_text
                    )

                    rows.append(
                        TimesheetRow(
                            row_index=valid_row_idx,
                            date_text=date_text,
                            date_parsed=parse_date(date_text, expected_year),
                            time_in_text=time_in_text,
                            time_in_parsed=time_in_parsed,
                            time_out_text=time_out_text,
                            time_out_parsed=time_out_parsed,
                            total_hours_text=hours_text,
                            total_hours_parsed=parse_hours(hours_text),
                            date_confidence=0.9,
                            time_in_confidence=0.9,
                            time_out_confidence=0.9,
                            hours_confidence=0.9,
                            date_source=OcrSource.VLM,
                            time_in_source=OcrSource.VLM,
                            time_out_source=OcrSource.VLM,
                            hours_source=OcrSource.VLM,
                        )
                    )
                    valid_row_idx += 1

                # Debug visualization
                if (
                    getattr(self.config, "debug", None)
                    and self.config.debug.visualize_ocr
                ):
                    debug_path = (
                        self.config.debug_output_path
                        / f"band_crop_payload_page_{page_number}.jpg"
                    )
                    debug_path.parent.mkdir(parents=True, exist_ok=True)
                    import cv2 as cv2  # noqa: PLC0415
                    cv2.imwrite(str(debug_path), payload_image)

        elif self.config.extraction_mode == "ocr_only":
            ocr_start = time_module.time()
            if not self.ocr_engine._initialized:
                init_start = time_module.time()
                self.ocr_engine._ensure_initialized()
                self._ocr_init_time = time_module.time() - init_start
                page_metrics.ocr_init_time_s = self._ocr_init_time
            ocr_result = self.ocr_engine.run(preprocessed)
            page_metrics.ocr_inference_time_s = time_module.time() - ocr_start
            page_metrics.total_boxes_detected = len(ocr_result.boxes)

            layout_start = time_module.time()
            layout = detect_layout(image, self.config, ocr_result.boxes)
            page_metrics.layout_detection_time_s = time_module.time() - layout_start

            sig_threshold = self.config.debug.signature_ocr_threshold
            is_sig_page = PhiAnonymizer.is_signature_page(
                len(ocr_result.boxes), sig_threshold
            )

            if is_sig_page:
                logger.info(
                    f"Page {page_number} is a signature page ({len(ocr_result.boxes)} boxes < {sig_threshold} threshold). Skipping row extraction."
                )
                header_boxes = boxes_in_zone(
                    ocr_result.boxes,
                    layout.header_zone.x_start,
                    layout.header_zone.y_start,
                    layout.header_zone.x_end,
                    layout.header_zone.y_end,
                )
                if header_boxes:
                    sorted_header = sorted(
                        header_boxes, key=lambda b: (b.y_center, b.x_center)
                    )
                    if sorted_header:
                        employee_name = clean_name(sorted_header[0].text)
                        employee_conf = sorted_header[0].confidence
            else:
                generated_dates = None
                if self.config.layout.transposed:
                    generated_dates = extract_week_dates(
                        source_file,
                        self.config.validation.week_start_day,
                        self.config.validation.week_length,
                    )
                    if generated_dates:
                        logger.info(
                            f"Generated week dates: {generated_dates[0]} → {generated_dates[-1]}"
                        )

                header_boxes = boxes_in_zone(
                    ocr_result.boxes,
                    layout.header_zone.x_start,
                    layout.header_zone.y_start,
                    layout.header_zone.x_end,
                    layout.header_zone.y_end,
                )

                if header_boxes:
                    sorted_header = sorted(
                        header_boxes, key=lambda b: (b.y_center, b.x_center)
                    )
                    if sorted_header:
                        employee_name = clean_name(sorted_header[0].text)
                        employee_conf = sorted_header[0].confidence

                extraction_start = time_module.time()
                empty_rows = 0
                for row_idx, row_zone in enumerate(layout.row_zones):
                    row = self._extract_row_ocr_only(
                        all_boxes=ocr_result.boxes,
                        row_zone=row_zone,
                        row_idx=row_idx,
                        layout=layout,
                        expected_year=expected_year,
                        generated_dates=generated_dates,
                    )
                    if row is not None:
                        rows.append(row)
                    else:
                        empty_rows += 1
                page_metrics.extraction_time_s = time_module.time() - extraction_start
                page_metrics.rows_extracted = len(rows)
                page_metrics.empty_rows_skipped = empty_rows
                page_metrics.vlm_fallbacks = 0

                if (
                    getattr(self.config, "debug", None)
                    and self.config.debug.visualize_ocr
                    and not is_sig_page
                ):
                    from . import debug_viz

                    debug_viz.render_page(
                        image=image,
                        ocr_boxes=ocr_result.boxes,
                        layout=layout,
                        field_bands=layout.field_bands,
                        vlm_fallbacks=[],
                        page_number=page_number,
                        source_file=anon_filename,
                        output_dir=self.config.debug_output_path,
                        prefix="ocr_only_",
                    )

        else:
            ocr_start = time_module.time()
            if not self.ocr_engine._initialized:
                init_start = time_module.time()
                self.ocr_engine._ensure_initialized()
                self._ocr_init_time = time_module.time() - init_start
                page_metrics.ocr_init_time_s = self._ocr_init_time
            ocr_result = self.ocr_engine.run(preprocessed)
            page_metrics.ocr_inference_time_s = time_module.time() - ocr_start
            page_metrics.total_boxes_detected = len(ocr_result.boxes)

            layout_start = time_module.time()
            layout = detect_layout(image, self.config, ocr_result.boxes)
            page_metrics.layout_detection_time_s = time_module.time() - layout_start

            # Signature page detection — skip row extraction
            sig_threshold = self.config.debug.signature_ocr_threshold
            is_sig_page = PhiAnonymizer.is_signature_page(
                len(ocr_result.boxes), sig_threshold
            )

            if is_sig_page:
                logger.info(
                    f"Page {page_number} is a signature page ({len(ocr_result.boxes)} boxes < {sig_threshold} threshold). Skipping row extraction."
                )
                # Only extract employee name from header zone
                header_boxes = boxes_in_zone(
                    ocr_result.boxes,
                    layout.header_zone.x_start,
                    layout.header_zone.y_start,
                    layout.header_zone.x_end,
                    layout.header_zone.y_end,
                )
                if header_boxes:
                    sorted_header = sorted(
                        header_boxes, key=lambda b: (b.y_center, b.x_center)
                    )
                    if sorted_header:
                        employee_name = clean_name(sorted_header[0].text)
                        employee_conf = sorted_header[0].confidence
            else:
                # Grid page — extract rows
                # Generate week dates from filename for transposed layouts
                generated_dates = None
                if self.config.layout.transposed:
                    generated_dates = extract_week_dates(
                        source_file,
                        self.config.validation.week_start_day,
                        self.config.validation.week_length,
                    )
                    if generated_dates:
                        logger.info(
                            f"Generated week dates: {generated_dates[0]} → {generated_dates[-1]}"
                        )

                header_boxes = boxes_in_zone(
                    ocr_result.boxes,
                    layout.header_zone.x_start,
                    layout.header_zone.y_start,
                    layout.header_zone.x_end,
                    layout.header_zone.y_end,
                )

                if header_boxes:
                    sorted_header = sorted(
                        header_boxes, key=lambda b: (b.y_center, b.x_center)
                    )
                    if sorted_header:
                        employee_name = clean_name(sorted_header[0].text)
                        employee_conf = sorted_header[0].confidence

                extraction_start = time_module.time()
                empty_rows = 0
                total_vlm_fallbacks = 0
                all_vlm_fallbacks: list[VlmFallbackCell] = []
                for row_idx, row_zone in enumerate(layout.row_zones):
                    row, vlm_count, vlm_cells = self._extract_row(
                        image=image,
                        preprocessed=preprocessed,
                        all_boxes=ocr_result.boxes,
                        row_zone=row_zone,
                        row_idx=row_idx,
                        layout=layout,
                        expected_year=expected_year,
                        generated_dates=generated_dates,
                    )
                    total_vlm_fallbacks += vlm_count
                    all_vlm_fallbacks.extend(vlm_cells)
                    if row is not None:
                        rows.append(row)
                    else:
                        empty_rows += 1
                page_metrics.extraction_time_s = time_module.time() - extraction_start
                page_metrics.rows_extracted = len(rows)
                page_metrics.empty_rows_skipped = empty_rows
                page_metrics.vlm_fallbacks = total_vlm_fallbacks

                if (
                    getattr(self.config, "debug", None)
                    and self.config.debug.visualize_ocr
                    and not is_sig_page
                ):
                    from . import debug_viz

                    debug_viz.render_page(
                        image=image,
                        ocr_boxes=ocr_result.boxes,
                        layout=layout,
                        field_bands=layout.field_bands,
                        vlm_fallbacks=[],
                        page_number=page_number,
                        source_file=anon_filename,
                        output_dir=self.config.debug_output_path,
                    )

        # 6. Build record with anonymized names
        record = TimesheetRecord(
            source_file=anon_filename,
            page_number=page_number,
            employee_name=anonymizer.anonymize_employee(employee_name),
            employee_name_confidence=employee_conf,
            patient_name=anonymizer.anonymize_patient(patient_name),
            patient_name_confidence=patient_conf,
            rows=rows,
        )

        # 7. Validate
        validation_start = time_module.time()
        validate_record(record, self.config)
        page_metrics.validation_time_s = time_module.time() - validation_start

        return record, page_metrics

    def _extract_row(
        self,
        image,
        preprocessed,
        all_boxes,
        row_zone,
        row_idx,
        layout,
        expected_year=None,
        generated_dates=None,
    ) -> tuple[TimesheetRow | None, int, list[VlmFallbackCell]]:
        """Extract data from a single table row using OCR boxes + fallback.

        Returns (row, vlm_fallback_count, vlm_fallback_cells).
        """
        h = layout.image_height
        w = layout.image_width

        # Use dynamically detected field bands from layout
        field_bounds = layout.field_bands

        # For transposed layouts, use spatial matching with generated dates
        if self.config.layout.transposed:
            return self._extract_transposed_row(
                all_boxes=all_boxes,
                image=image,
                row_zone=row_zone,
                row_idx=row_idx,
                field_bounds=field_bounds,
                generated_dates=generated_dates,
                expected_year=expected_year,
            )

        # Non-transposed layout: use original OCR + fallback approach
        # Find OCR boxes in each column of this row
        cell_data: dict[str, tuple[str, float]] = {}
        cell_confidences: dict[str, float] = {}
        vlm_fallbacks = 0
        vlm_cells: list[VlmFallbackCell] = []

        for col_name, (start_pix, end_pix) in field_bounds.items():
            col_boxes = boxes_in_zone(
                all_boxes, start_pix, row_zone.y_start, end_pix, row_zone.y_end
            )

            if col_boxes:
                text = " ".join(
                    b.text for b in sorted(col_boxes, key=lambda b: b.x_center)
                )
                conf = min(b.confidence for b in col_boxes)
            else:
                text = ""
                conf = 0.0

            cell_data[col_name] = (text, conf)
            cell_confidences[col_name] = conf

        # Check if this row has any meaningful content
        if all(text == "" for text, _ in cell_data.values()):
            return None, 0, []  # Empty row, skip

        # Route through confidence check
        row_data = dict(cell_data)  # Copy for potential VLM updates
        sources: dict[str, OcrSource] = {k: OcrSource.PPOCR for k in field_bounds}

        # Per-cell fallback for low-confidence cells
        for col_name, conf in cell_confidences.items():
            route = route_by_confidence(conf, self.config)
            if route == Route.FALLBACK:
                vlm_fallbacks += 1
                start_pix, end_pix = field_bounds[col_name]
                cell_crop = image[row_zone.y_start : row_zone.y_end, start_pix:end_pix]

                vlm_value, vlm_conf = self.vlm.extract_cell_value(
                    cell_crop, col_name, expected_year
                )
                if vlm_value:
                    row_data[col_name] = (vlm_value, vlm_conf)
                    sources[col_name] = OcrSource.VLM
                    vlm_cells.append(
                        VlmFallbackCell(
                            row_idx=row_idx,
                            field_name=col_name,
                            x_start=start_pix,
                            y_start=row_zone.y_start,
                            x_end=end_pix,
                            y_end=row_zone.y_end,
                            vlm_text=vlm_value,
                            vlm_conf=vlm_conf,
                        )
                    )

        row = self._build_timesheet_row(
            row_idx, row_data, sources, field_bounds, expected_year
        )
        return row, vlm_fallbacks, vlm_cells

    def _extract_transposed_row(
        self,
        all_boxes,
        image,
        row_zone,
        row_idx,
        field_bounds,
        generated_dates,
        expected_year=None,
    ) -> tuple[TimesheetRow | None, int, list[VlmFallbackCell]]:
        """Extract data from a transposed row (column) using spatial matching.

        Strategy:
        1. DATE: Use generated_dates[row_idx] (source=GENERATED)
        2. TIME IN/TIME OUT/HOURS: Find OCR boxes whose centers fall within
           the column × field_band zone. Concatenate text from matching boxes.
        3. If a cell is empty after spatial matching, fall back to VLM on
           the cropped cell.

        Note: PaddleOCR often misses handwritten text. When spatial matching
        returns empty, VLM fallback handles it.

        Returns (row, vlm_fallback_count, vlm_fallback_cells).
        """
        x_left = row_zone.x_start
        x_right = row_zone.x_end
        vlm_fallbacks = 0
        vlm_cells: list[VlmFallbackCell] = []

        # DATE: Use generated date
        if generated_dates and row_idx < len(generated_dates):
            gen_date = generated_dates[row_idx]
            date_text = gen_date.strftime("%m/%d/%Y")
            date_conf = 1.0
            date_source = OcrSource.GENERATED
        else:
            date_text = ""
            date_conf = 0.0
            date_source = OcrSource.PPOCR

        row_data = {"date": (date_text, date_conf)}
        sources = {"date": date_source}

        for field_name in ["time_in", "time_out", "total_hours"]:
            y_start, y_end = field_bounds.get(field_name, (0, 0))

            # Find boxes in column × field_band zone
            cell_boxes = [
                b
                for b in all_boxes
                if x_left <= b.x_center <= x_right
                and y_start <= b.y_center <= y_end
                and b.text.strip()
            ]

            if cell_boxes:
                # Sort by Y position (top to bottom within the band)
                cell_boxes.sort(key=lambda b: b.y_center)
                text = " ".join(b.text for b in cell_boxes)
                conf = min(b.confidence for b in cell_boxes)
                source = OcrSource.PPOCR
                logger.debug(
                    f"Row {row_idx} {field_name}: PPOCR → '{text}' (conf={conf:.3f})"
                )
            else:
                # VLM fallback on cropped cell
                vlm_fallbacks += 1
                cell_crop = image[y_start:y_end, x_left:x_right]
                if cell_crop.size > 0:
                    vlm_value, vlm_conf = self.vlm.extract_cell_value(
                        cell_crop, field_name, expected_year
                    )
                    text = vlm_value
                    conf = vlm_conf
                    source = OcrSource.VLM
                    logger.debug(
                        f"Row {row_idx} {field_name}: VLM → '{text}' (conf={conf:.3f})"
                    )
                    vlm_cells.append(
                        VlmFallbackCell(
                            row_idx=row_idx,
                            field_name=field_name,
                            x_start=x_left,
                            y_start=y_start,
                            x_end=x_right,
                            y_end=y_end,
                            vlm_text=vlm_value,
                            vlm_conf=vlm_conf,
                        )
                    )
                else:
                    text = ""
                    conf = 0.0
                    source = OcrSource.PPOCR

            row_data[field_name] = (text, conf)
            sources[field_name] = source

        # Skip if no meaningful content
        if all(
            not row_data.get(k, ("", ""))[0].strip()
            for k in ["time_in", "time_out", "total_hours"]
        ):
            return None, vlm_fallbacks, vlm_cells

        row = self._build_timesheet_row(
            row_idx, row_data, sources, field_bounds, expected_year
        )
        return row, vlm_fallbacks, vlm_cells

    def _extract_row_ocr_only(
        self,
        all_boxes,
        row_zone,
        row_idx,
        layout,
        expected_year=None,
        generated_dates=None,
    ) -> TimesheetRow | None:
        """Extract data from a single row using OCR only — no VLM fallback."""
        field_bounds = layout.field_bands

        if self.config.layout.transposed:
            return self._extract_transposed_row_ocr_only(
                all_boxes=all_boxes,
                row_zone=row_zone,
                row_idx=row_idx,
                field_bounds=field_bounds,
                generated_dates=generated_dates,
                expected_year=expected_year,
            )

        cell_data: dict[str, tuple[str, float]] = {}
        sources: dict[str, OcrSource] = {}

        for col_name, (start_pix, end_pix) in field_bounds.items():
            col_boxes = boxes_in_zone(
                all_boxes, start_pix, row_zone.y_start, end_pix, row_zone.y_end
            )

            if col_boxes:
                text = " ".join(
                    b.text for b in sorted(col_boxes, key=lambda b: b.x_center)
                )
                conf = min(b.confidence for b in col_boxes)
            else:
                text = ""
                conf = 0.0

            cell_data[col_name] = (text, conf)
            sources[col_name] = OcrSource.PPOCR

        if all(text == "" for text, _ in cell_data.values()):
            return None

        return self._build_timesheet_row(
            row_idx, cell_data, sources, field_bounds, expected_year
        )

    def _extract_transposed_row_ocr_only(
        self,
        all_boxes,
        row_zone,
        row_idx,
        field_bounds,
        generated_dates,
        expected_year=None,
    ) -> TimesheetRow | None:
        """Extract data from a transposed row using OCR only — no VLM fallback."""
        x_left = row_zone.x_start
        x_right = row_zone.x_end

        if generated_dates and row_idx < len(generated_dates):
            gen_date = generated_dates[row_idx]
            date_text = gen_date.strftime("%m/%d/%Y")
            date_conf = 1.0
            date_source = OcrSource.GENERATED
        else:
            date_text = ""
            date_conf = 0.0
            date_source = OcrSource.PPOCR

        row_data = {"date": (date_text, date_conf)}
        sources = {"date": date_source}

        for field_name in ["time_in", "time_out", "total_hours"]:
            y_start, y_end = field_bounds.get(field_name, (0, 0))

            cell_boxes = [
                b
                for b in all_boxes
                if x_left <= b.x_center <= x_right
                and y_start <= b.y_center <= y_end
                and b.text.strip()
            ]

            if cell_boxes:
                cell_boxes.sort(key=lambda b: b.y_center)
                text = " ".join(b.text for b in cell_boxes)
                conf = min(b.confidence for b in cell_boxes)
            else:
                text = ""
                conf = 0.0

            row_data[field_name] = (text, conf)
            sources[field_name] = OcrSource.PPOCR

        if all(
            not row_data.get(k, ("", ""))[0].strip()
            for k in ["time_in", "time_out", "total_hours"]
        ):
            return None

        return self._build_timesheet_row(
            row_idx, row_data, sources, field_bounds, expected_year
        )

    def _build_timesheet_row(
        self, row_idx, row_data, sources, field_bounds, expected_year=None
    ) -> TimesheetRow:
        """Build a TimesheetRow from extracted data."""
        date_text, date_conf = row_data.get("date", ("", 0.0))
        time_in_text, time_in_conf = row_data.get("time_in", ("", 0.0))
        time_out_text, time_out_conf = row_data.get("time_out", ("", 0.0))
        hours_text, hours_conf = row_data.get("total_hours", ("", 0.0))

        row = TimesheetRow(
            row_index=row_idx,
            date_text=date_text,
            date_parsed=parse_date(date_text, expected_year),
            time_in_text=time_in_text,
            time_in_parsed=parse_time(time_in_text),
            time_out_text=time_out_text,
            time_out_parsed=parse_time(time_out_text),
            total_hours_text=hours_text,
            total_hours_parsed=parse_hours(hours_text),
            date_confidence=date_conf,
            time_in_confidence=time_in_conf,
            time_out_confidence=time_out_conf,
            hours_confidence=hours_conf,
            date_source=sources.get("date", OcrSource.PPOCR),
            time_in_source=sources.get("time_in", OcrSource.PPOCR),
            time_out_source=sources.get("time_out", OcrSource.PPOCR),
            hours_source=sources.get("total_hours", OcrSource.PPOCR),
        )

        return row

    def _collect_row_metrics(self, row: TimesheetRow, record: TimesheetRecord) -> None:
        """Collect per-row benchmark metrics after extraction and validation."""
        corrections = []
        corrections_detail_parts = []

        # Track corrections: raw OCR text vs parsed value
        for field_name, raw, parsed in [
            ("date", row.date_text, str(row.date_parsed) if row.date_parsed else ""),
            (
                "time_in",
                row.time_in_text,
                row.time_in_parsed.strftime("%H:%M") if row.time_in_parsed else "",
            ),
            (
                "time_out",
                row.time_out_text,
                row.time_out_parsed.strftime("%H:%M") if row.time_out_parsed else "",
            ),
            (
                "hours",
                row.total_hours_text,
                str(row.total_hours_parsed)
                if row.total_hours_parsed is not None
                else "",
            ),
        ]:
            raw_stripped = raw.strip()
            parsed_stripped = parsed.strip()
            if raw_stripped and parsed_stripped and raw_stripped != parsed_stripped:
                corrections.append((field_name, raw_stripped, parsed_stripped))
                corrections_detail_parts.append(
                    f"{field_name}|{raw_stripped}|{parsed_stripped}"
                )

        row_metrics = RowMetrics(
            source_file=record.source_file,
            page_number=record.page_number,
            row_index=row.row_index,
            employee_name=record.employee_name,
            patient_name=record.patient_name,
            raw_ocr_date=row.date_text,
            raw_ocr_time_in=row.time_in_text,
            raw_ocr_time_out=row.time_out_text,
            raw_ocr_hours=row.total_hours_text,
            parsed_date=str(row.date_parsed) if row.date_parsed else "",
            parsed_time_in=row.time_in_parsed.strftime("%H:%M")
            if row.time_in_parsed
            else "",
            parsed_time_out=row.time_out_parsed.strftime("%H:%M")
            if row.time_out_parsed
            else "",
            parsed_hours=str(row.total_hours_parsed)
            if row.total_hours_parsed is not None
            else "",
            date_confidence=row.date_confidence,
            time_in_confidence=row.time_in_confidence,
            time_out_confidence=row.time_out_confidence,
            hours_confidence=row.hours_confidence,
            date_source=row.date_source.value if hasattr(row, "date_source") else "",
            time_in_source=row.time_in_source.value
            if hasattr(row, "time_in_source")
            else "",
            time_out_source=row.time_out_source.value
            if hasattr(row, "time_out_source")
            else "",
            hours_source=row.hours_source.value if hasattr(row, "hours_source") else "",
            calculated_hours=row.calculated_hours(),
            total_hours_text=row.total_hours_text,
            is_overnight=row.is_overnight,
            status=row.status.value,
            validation_errors="; ".join(row.validation_errors),
            corrections_applied=len(corrections),
            corrections_detail="; ".join(corrections_detail_parts),
        )
        self.benchmark.add_row(row_metrics)
