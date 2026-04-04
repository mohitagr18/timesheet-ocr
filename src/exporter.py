"""Export module — CSV, JSON, and Excel output generation."""

from __future__ import annotations

import csv
import json
import logging
from datetime import datetime, date, time
from pathlib import Path
from typing import TYPE_CHECKING

from .models import ExtractionResult, RowStatus

if TYPE_CHECKING:
    from .config import AppConfig

logger = logging.getLogger(__name__)


def export_results(result: ExtractionResult, config: AppConfig) -> list[Path]:
    """Export extraction results in all configured formats.

    Returns list of output file paths created.
    """
    output_dir = config.output_path
    output_dir.mkdir(parents=True, exist_ok=True)

    stem = Path(result.source_file).stem
    created_files: list[Path] = []

    formats = config.export.formats

    if "csv" in formats:
        csv_path = output_dir / f"{stem}_results.csv"
        _export_csv(result, csv_path)
        created_files.append(csv_path)

    if "json" in formats:
        json_path = output_dir / f"{stem}_results.json"
        _export_json(result, json_path)
        created_files.append(json_path)

    if "xlsx" in formats:
        xlsx_path = output_dir / "merged_results.xlsx"
        _export_excel(result, xlsx_path, config)
        created_files.append(xlsx_path)

    # Always export review queue if there are flagged items and toggle is on
    if result.review_items and config.export.include_review_json:
        review_path = output_dir / f"{stem}_review.json"
        _export_review_queue(result, review_path)
        created_files.append(review_path)

    # Always export validation report if toggle is on
    if config.export.include_report_json:
        report_path = output_dir / f"{stem}_report.json"
        _export_report(result, report_path)
        created_files.append(report_path)

    logger.info(f"Exported {len(created_files)} files to {output_dir}")
    return created_files


def _sorted_rows(result: ExtractionResult):
    """Flatten and sort all rows by employee (alpha) → date (asc) → time_in (asc)."""
    all_rows = []
    for record in result.records:
        for row in record.rows:
            all_rows.append((record, row))

    all_rows.sort(
        key=lambda x: (
            x[0].employee_name.lower(),
            x[1].date_parsed or date.min,
            x[1].time_in_parsed or time.min,
        )
    )

    return all_rows


def _export_csv(result: ExtractionResult, path: Path) -> None:
    """Export results as CSV (one row per timesheet row)."""
    fieldnames = [
        "source_file",
        "page",
        "row_index",
        "employee_name",
        "patient_name",
        "date",
        "date_source",
        "time_in",
        "time_in_source",
        "time_out",
        "time_out_source",
        "total_hours",
        "hours_source",
        "calculated_hours",
        "is_overnight",
        "is_over_24h_limit",
        "confidence",
        "status",
    ]

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for record, row in _sorted_rows(result):
            writer.writerow(
                {
                    "source_file": record.source_file,
                    "page": record.page_number,
                    "row_index": row.row_index,
                    "employee_name": record.employee_name,
                    "patient_name": record.patient_name,
                    "date": str(row.date_parsed) if row.date_parsed else row.date_text,
                    "date_source": row.date_source.value
                    if hasattr(row, "date_source")
                    else "",
                    "time_in": row.time_in_parsed.strftime("%H:%M")
                    if row.time_in_parsed
                    else row.time_in_text,
                    "time_in_source": row.time_in_source.value
                    if hasattr(row, "time_in_source")
                    else "",
                    "time_out": row.time_out_parsed.strftime("%H:%M")
                    if row.time_out_parsed
                    else row.time_out_text,
                    "time_out_source": row.time_out_source.value
                    if hasattr(row, "time_out_source")
                    else "",
                    "total_hours": row.total_hours_parsed
                    if row.total_hours_parsed
                    else row.total_hours_text,
                    "hours_source": row.hours_source.value
                    if hasattr(row, "hours_source")
                    else "",
                    "calculated_hours": row.calculated_hours() or "",
                    "is_overnight": "Yes" if row.is_overnight else "",
                    "is_over_24h_limit": "Yes"
                    if getattr(row, "is_over_24h_limit", False)
                    else "",
                    "confidence": f"{row.min_confidence:.2f}",
                    "status": row.status.value,
                }
            )

    logger.info(f"CSV exported: {path}")


def _export_json(result: ExtractionResult, path: Path) -> None:
    """Export full structured results as JSON."""
    data = result.model_dump(mode="json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)
    logger.info(f"JSON exported: {path}")


def _export_excel(result: ExtractionResult, path: Path, config: AppConfig) -> None:
    """Export results as formatted Excel workbook. Appends to existing if it exists."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if path.exists():
        wb = load_workbook(path)
        if config.export.excel_sheet_name in wb.sheetnames:
            ws = wb[config.export.excel_sheet_name]
        else:
            ws = wb.active
            ws.title = config.export.excel_sheet_name
        data_row = ws.max_row + 1
        headers_needed = False
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = config.export.excel_sheet_name
        data_row = 2
        headers_needed = True

    # ── Styles ─────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(
        start_color="2F5496", end_color="2F5496", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    accepted_fill = PatternFill(
        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
    )
    flagged_fill = PatternFill(
        start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
    )
    failed_fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ── Headers ────────────────────────────────────────────────────
    headers = [
        "Source File",
        "Page",
        "Row #",
        "Employee Name",
        "Patient Name",
        "Date",
        "Date Source",
        "Time In",
        "Time In Source",
        "Time Out",
        "Time Out Source",
        "Total Hours",
        "Hours Source",
        "Calculated Hours",
        "Overnight",
        "Over 24h Limit",
        "Confidence",
        "Status",
        "Issues",
    ]

    if headers_needed:
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    # ── Data rows ──────────────────────────────────────────────────
    for record, row in _sorted_rows(result):
        values = [
            record.source_file,
            record.page_number,
            row.row_index,
            record.employee_name,
            record.patient_name,
            str(row.date_parsed) if row.date_parsed else row.date_text,
            row.date_source.value if hasattr(row, "date_source") else "",
            row.time_in_parsed.strftime("%H:%M")
            if row.time_in_parsed
            else row.time_in_text,
            row.time_in_source.value if hasattr(row, "time_in_source") else "",
            row.time_out_parsed.strftime("%H:%M")
            if row.time_out_parsed
            else row.time_out_text,
            row.time_out_source.value if hasattr(row, "time_out_source") else "",
            row.total_hours_parsed
            if row.total_hours_parsed is not None
            else row.total_hours_text,
            row.hours_source.value if hasattr(row, "hours_source") else "",
            row.calculated_hours() or "",
            "Yes" if row.is_overnight else "",
            "Yes" if getattr(row, "is_over_24h_limit", False) else "",
            round(row.min_confidence, 2),
            row.status.value,
            "; ".join(row.validation_errors) if row.validation_errors else "",
        ]

        # Choose row fill based on status
        if row.status == RowStatus.ACCEPTED:
            row_fill = accepted_fill
        elif row.status == RowStatus.FLAGGED:
            row_fill = flagged_fill
        else:
            row_fill = failed_fill

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=data_row, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

        data_row += 1

    # ── Auto-fit column widths ─────────────────────────────────────
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(2, data_row):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{data_row - 1}"

    wb.save(path)
    logger.info(f"Excel exported: {path}")


def _export_review_queue(result: ExtractionResult, path: Path) -> None:
    """Export review queue as JSON."""
    data = {
        "source_file": result.source_file,
        "generated_at": datetime.now().isoformat(),
        "total_rows": result.total_rows,
        "accepted_rows": result.accepted_count,
        "flagged_rows": result.flagged_count,
        "items": [item.model_dump(mode="json") for item in result.review_items],
    }

    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)
    logger.info(f"Review queue exported: {path}")


def _export_report(result: ExtractionResult, path: Path) -> None:
    """Export validation summary report as JSON."""
    # Collect all validation warnings
    all_warnings = []
    for record in result.records:
        for row in record.rows:
            for error in row.validation_errors:
                all_warnings.append(
                    {
                        "page": record.page_number,
                        "row": row.row_index,
                        "rule": error,
                        "detail": f"Row {row.row_index}: {error}",
                    }
                )

    data = {
        "source_file": result.source_file,
        "processing_time_seconds": round(result.processing_time_seconds, 2),
        "total_pages": result.total_pages,
        "total_rows_detected": result.total_rows,
        "accepted": result.accepted_count,
        "flagged_for_review": result.flagged_count,
        "failed": result.failed_count,
        "validation_warnings": all_warnings,
    }

    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)
    logger.info(f"Report exported: {path}")
