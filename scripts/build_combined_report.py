#!/usr/bin/env python
"""Build complete combined benchmark report from all 6 approaches.

Reads merged_results.xlsx from each approach folder + ground_truth.xlsx,
computes all metrics (hours mismatch, field missing, CER, GT accuracy,
consensus/KPIs), and writes a single benchmark_combined.xlsx with 4 sheets:
  1. Approach Comparison (with GT rows)
  2. Human-Verified Results
  3. IEEE Paper Results (incl. False/Missed Accepts)
    4. Per-Row Detail

Backs up existing file before overwriting.

Usage:
    uv run python scripts/build_combined_report.py
"""

import glob
import os
import re
import shutil
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_DIR = PROJECT_ROOT / "output"
COMBINED_DIR = OUTPUT_DIR / "combined"
GT_PATH = PROJECT_ROOT / "ground_truth.xlsx"
COMBINED_PATH = COMBINED_DIR / "benchmark_combined.xlsx"
BACKUP_PATH = COMBINED_DIR / "benchmark_combined_backup.xlsx"

APPROACHES = [
    ("ocr_only", "OCR Only", "E8E8E8"),
    ("ppocr_grid", "OCR+VLM", "E2EFDA"),
    ("vlm_full_page", "VLM Full", "D6E4F0"),
    ("layout_guided_vlm_local", "Layout Local", "FFF2CC"),
    ("layout_guided_vlm_cloud", "Layout Cloud", "FCE4EC"),
    ("band_crop_vlm_cloud", "Band-Crop VLM", "F3E5F5"),
]

FILLS = {}
for aid, _, color in APPROACHES:
    FILLS[aid] = PatternFill(start_color=color, end_color=color, fill_type="solid")

ANON_MAP = {
    "patient_a_week1": "File 1 (Week 1)",
    "patient_b_week2": "File 2 (Week 2)",
    "patient_c_week3": "File 3 (Week 3)",
}

HOURS_TOL = 0.25
TIME_TOL = 30

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MISMATCH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PARTIAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


# ── Parsing helpers ──────────────────────────────────────────────────

def _parse_date(val):
    if val is None:
        return None
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    val = str(val).strip()
    if not val:
        return None
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", val)
    if m:
        mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        return f"{y:04d}-{mo:02d}-{d:02d}"
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", val)
    if m:
        return val
    return val


def _parse_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _parse_time_min(val):
    if val is None:
        return None
    val = str(val).strip()
    if not val:
        return None
    m = re.match(r"(\d{1,2}):(\d{2})\s*(AM|PM|am|pm)?", val)
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        ap = m.group(3)
        if ap and ap.lower() == "pm" and h != 12:
            h += 12
        elif ap and ap.lower() == "am" and h == 12:
            h = 0
        if 0 <= h <= 23 and 0 <= mi <= 59:
            return h * 60 + mi
        return None
    m = re.match(r"(\d{1,2})(\d{2})(?!\d)", val)
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        if 0 <= h <= 23 and 0 <= mi <= 59:
            return h * 60 + mi
        return None
    return None


def _short_label(label, max_len=16):
    if len(label) <= max_len:
        return label
    suffix = label.split("(")[-1].rstrip(")") if "(" in label else label.split()[-1]
    prefix_max = max_len - len(suffix) - 3
    prefix = label[:prefix_max].rsplit(" ", 1)[0] if " " in label[:prefix_max] else label[:prefix_max]
    return f"{prefix} ({suffix})"


def _anon_source(source):
    for key, label in ANON_MAP.items():
        if key in str(source):
            return label
    return str(source)


def _levenshtein_cer(s1, s2):
    if not s1 and not s2:
        return 0.0
    m, n = len(s1), len(s2)
    dp = [[0] * (n + 1) for _ in range(m + 1)]
    for i in range(m + 1):
        dp[i][0] = i
    for j in range(n + 1):
        dp[0][j] = j
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            if s1[i - 1] == s2[j - 1]:
                dp[i][j] = dp[i - 1][j - 1]
            else:
                dp[i][j] = 1 + min(dp[i - 1][j], dp[i][j - 1], dp[i - 1][j - 1])
    return dp[m][n] / max(len(s1), len(s2))


def _style_cell(ws, row, col, fill=None):
    cell = ws.cell(row=row, column=col)
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    return cell


# ── Load data ────────────────────────────────────────────────────────

def load_approach_data(approach_id):
    merged_path = OUTPUT_DIR / approach_id / "merged_results.xlsx"
    rows = []
    summary = {}
    if not merged_path.exists():
        return rows, summary

    # Read benchmark for timing/CER stats
    bench_files = sorted(glob.glob(str(OUTPUT_DIR / approach_id / "benchmark_*.xlsx")))
    for bf in bench_files:
        try:
            wb = openpyxl.load_workbook(bf, read_only=True, data_only=True)
            if "Run Summary" in wb.sheetnames:
                ws = wb["Run Summary"]
                for row in ws.iter_rows(values_only=True):
                    if row[0]:
                        summary[str(row[0])] = row[1]
            wb.close()
        except Exception:
            pass

    wb = openpyxl.load_workbook(merged_path, read_only=True, data_only=True)
    ws = wb.active
    header = None
    for r in ws.iter_rows(values_only=True):
        if header is None:
            header = [str(c).strip() if c else "" for c in r]
            continue
        if not any(r):
            continue
        rec = dict(zip(header, r))
        rec["_date"] = _parse_date(rec.get("Date"))
        rec["_status"] = str(rec.get("Status", "")).strip().lower()
        rec["_source"] = str(rec.get("Source File", ""))
        rec["_time_in"] = rec.get("Time In", "")
        rec["_time_out"] = rec.get("Time Out", "")
        rec["_hours"] = rec.get("Total Hours")
        rows.append(rec)
    wb.close()

    # Fill summary
    summary.setdefault("Total Processing Time (s)", 0)
    summary.setdefault("Total Rows Extracted", len(rows))
    summary.setdefault("Number of Pages", len(set(r["_source"] for r in rows)))
    summary.setdefault("VLM Fallbacks Triggered", 0)
    accepted = sum(1 for r in rows if r["_status"] == "accepted")
    flagged = sum(1 for r in rows if r["_status"] == "flagged")
    failed = sum(1 for r in rows if r["_status"] == "failed")
    summary.setdefault("Accepted Rows", accepted)
    summary.setdefault("Flagged Rows", flagged)
    summary.setdefault("Failed Rows", failed)

    return rows, summary


def load_ground_truth():
    if not GT_PATH.exists():
        return []
    try:
        wb = openpyxl.load_workbook(GT_PATH, read_only=True, data_only=True)
        ws = wb.active
        header = None
        rows = []
        for r in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c).strip() if c else "" for c in r]
                continue
            if not any(r):
                continue
            rec = dict(zip(header, r))
            rec["_date"] = _parse_date(rec.get("date"))
            rec["_time_in_min"] = _parse_time_min(rec.get("time_in"))
            rec["_time_out_min"] = _parse_time_min(rec.get("time_out"))
            gt_h = _parse_float(rec.get("total_hours"))
            if gt_h is None:
                gt_h = _compute_hours(rec["_time_in_min"], rec["_time_out_min"])
            rec["_hours"] = gt_h
            rows.append(rec)
        wb.close()
        return rows
    except Exception as e:
        print(f"  WARNING: Failed to load ground truth: {e}")
        return []


def _compute_hours(ti_min, to_min):
    if ti_min is None or to_min is None:
        return None
    diff = to_min - ti_min
    if diff < 0:
        diff += 24 * 60
    return diff / 60.0


# ── GT accuracy computation ──────────────────────────────────────────

def compute_gt_metrics(rows, gt):
    if not gt:
        return {"gt_h": 0.0, "gt_ti": 0.0, "gt_to": 0.0, "gt_fc": 0.0,
                "gt_matched": 0, "gt_not_found": 0, "matched_list": [],
                "duplicates": [], "extra": [],
                "false_accepts": 0, "missed_accepts": 0}

    gt_by_key = {}
    for g in gt:
        source = str(g.get("source_file", ""))
        key = (source, g["_date"])
        gt_by_key[key] = g

    approach_by_key = {}
    for ar in rows:
        key = (ar["_source"], ar["_date"])
        approach_by_key.setdefault(key, []).append(ar)

    matched_list = []
    duplicates = []
    extra_list = []
    gt_matched = 0
    gt_not_found = 0
    hours_ok = 0
    ti_ok = 0
    to_ok = 0
    fc_ok = 0
    false_accepts = 0
    missed_accepts = 0

    for gt_key, gt_rec in gt_by_key.items():
        candidates = approach_by_key.get(gt_key, [])
        if not candidates:
            gt_not_found += 1
            matched_list.append({"_date": gt_rec["_date"], "_status": "not_extracted",
                                 "_fully_correct": False, "_not_extracted": True})
            continue

        scored = []
        for ar in candidates:
            ah = _parse_float(ar["_hours"])
            ati = _parse_time_min(ar["_time_in"])
            ato = _parse_time_min(ar["_time_out"])
            gh = _parse_float(gt_rec.get("total_hours"))
            gti = _parse_time_min(gt_rec.get("time_in"))
            gto = _parse_time_min(gt_rec.get("time_out"))
            hd = abs(ah - gh) if (ah is not None and gh is not None) else 999
            tid = abs(ati - gti) if (ati is not None and gti is not None) else 9999
            tod = abs(ato - gto) if (ato is not None and gto is not None) else 9999
            scored.append((hd * 1000 + tid + tod * 0.01, ar))
        scored.sort(key=lambda x: x[0])
        _, best = scored[0]

        for _, dup in scored[1:]:
            duplicates.append(dup)

        gt_matched += 1
        bh = _parse_float(best["_hours"])
        bti = _parse_time_min(best["_time_in"])
        bto = _parse_time_min(best["_time_out"])
        gh = _parse_float(gt_rec.get("total_hours"))
        gti = _parse_time_min(gt_rec.get("time_in"))
        gto = _parse_time_min(gt_rec.get("time_out"))

        ho = bh is not None and gh is not None and abs(bh - gh) <= HOURS_TOL
        tino = bti is not None and gti is not None and abs(bti - gti) <= TIME_TOL
        too = bto is not None and gto is not None and abs(bto - gto) <= TIME_TOL

        if ho:
            hours_ok += 1
        if tino:
            ti_ok += 1
        if too:
            to_ok += 1
        if ho and tino and too:
            fc_ok += 1

        matched_list.append({**best, "_fully_correct": ho and tino and too,
                             "_partially_correct": sum([ho, tino, too]) > 0,
                             "_not_extracted": False,
                             "_hours_ok": ho, "_time_in_ok": tino, "_time_out_ok": too})

        if best.get("_status") == "accepted" and not (ho and tino and too):
            false_accepts += 1
        if best.get("_status") != "accepted" and ho and tino and too:
            missed_accepts += 1

    # Extra rows
    gt_dates_all = set(gt_by_key.keys())
    for ar_key in approach_by_key:
        if ar_key not in gt_dates_all:
            extra_list.extend(approach_by_key[ar_key])

    m = gt_matched
    return {
        "gt_h": hours_ok / m if m > 0 else 0.0,
        "gt_ti": ti_ok / m if m > 0 else 0.0,
        "gt_to": to_ok / m if m > 0 else 0.0,
        "gt_fc": fc_ok / m if m > 0 else 0.0,
        "gt_matched": m,
        "gt_not_found": gt_not_found,
        "matched_list": matched_list,
        "duplicates": duplicates,
        "extra": extra_list,
        "false_accepts": false_accepts,
        "missed_accepts": missed_accepts,
        "fully_correct_count": fc_ok,
        "partial_count": m - fc_ok - (m - sum(1 for ml in matched_list if ml.get("_fully_correct"))),
    }


# ── Sheet writers ────────────────────────────────────────────────────

def _write_approach_comparison(ws, all_data):
    row = 1
    ws.cell(row=row, column=1, value="Approach Comparison: Handwritten Timesheet OCR").font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True, size=10)
    row += 2

    ws.cell(row=row, column=1, value="SUMMARY METRICS").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    headers = ["Metric"] + [d["label"] for d in all_data.values()]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    metrics_list = [
        ("Total Processing Time (s)", lambda a: f"{d[a]['summary'].get('Total Processing Time (s)', 0):.2f}"),
        ("Pages Processed", lambda a: str(d[a]['summary'].get('Number of Pages', 0))),
        ("Rows Extracted", lambda a: str(d[a]['summary'].get('Total Rows Extracted', 0))),
        ("Accepted Rows", lambda a: str(d[a]['summary'].get('Accepted Rows', 0))),
        ("Flagged Rows", lambda a: str(d[a]['summary'].get('Flagged Rows', 0))),
        ("Failed Rows", lambda a: str(d[a]['summary'].get('Failed Rows', 0))),
        ("Mean Confidence", lambda a: str(round(d[a]['summary'].get('Mean Overall Confidence', 0), 4))),
        ("Min Confidence", lambda a: str(round(d[a]['summary'].get('Min Overall Confidence', 0), 4))),
        ("VLM Fallbacks Triggered", lambda a: str(d[a]['summary'].get('VLM Fallbacks Triggered', 0))),
        ("Hours Mismatch Rate", lambda a: _compute_hours_mismatch(d[a])),
        ("GT Hours Accuracy (±15min)", lambda a: f"{d[a]['gt']['gt_h']*100:.1f}%" if d[a]['gt']['gt_h'] > 0 else "N/A"),
        ("GT Time-In Accuracy (±30min)", lambda a: f"{d[a]['gt']['gt_ti']*100:.1f}%" if d[a]['gt']['gt_ti'] > 0 else "N/A"),
        ("GT Time-Out Accuracy (±30min)", lambda a: f"{d[a]['gt']['gt_to']*100:.1f}%" if d[a]['gt']['gt_to'] > 0 else "N/A"),
        ("GT Fully Correct Rate", lambda a: f"{d[a]['gt']['gt_fc']*100:.1f}%" if d[a]['gt']['gt_fc'] > 0 else "N/A"),
        ("GT Rows Matched", lambda a: str(d[a]['gt']['gt_matched'])),
        ("GT Rows Not Found", lambda a: str(d[a]['gt']['gt_not_found'])),
        ("Field Missing Rate", lambda a: _compute_field_missing(d[a])),
        ("Mean CER", lambda a: _compute_mean_cer(d[a])),
                    ("False Accepts", lambda a: str(d[a]['gt']['false_accepts'])),
        ("Missed Accepts", lambda a: str(d[a]['gt']['missed_accepts'])),
]

    for label, fn in metrics_list:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        for i, (aid, d_) in enumerate(all_data.items()):
            fill = FILLS.get(aid, None)
            _style_cell(ws, row, i + 2, fill)
            ws.cell(row=row, column=i + 2, value=fn(aid))
        row += 1

    ws.column_dimensions["A"].width = 35
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22


def _compute_hours_mismatch(data):
    matched = data["gt"].get("matched_list", [])
    extracted = [m for m in matched if not m.get("_not_extracted")]
    if not extracted:
        return "N/A"
    mismatch = sum(1 for m in extracted if not m.get("_hours_ok", False))
    return f"{mismatch/len(extracted)*100:.1f}%"

def _compute_field_missing(data):
    rows = data["rows"]
    if not rows:
        return "0.0%"
    missing = sum(1 for r in rows if not r.get("Date") or not r.get("Time In") or not r.get("Time Out"))
    return f"{missing/len(rows)*100:.1f}%"


def _compute_mean_cer(data):
    rows = data["rows"]
    if not rows:
        return "0.0000"
    cer_vals = []
    for r in rows:
        for rk, pk in [("Time In", "Parsed Time In"), ("Time Out", "Parsed Time Out")]:
            raw = str(r.get(rk, "")).strip().lower()
            parsed = str(r.get(pk, "")).strip().lower()
            if raw and parsed and raw != parsed:
                cer_vals.append(_levenshtein_cer(raw, parsed))
    return f"{sum(cer_vals)/len(cer_vals):.4f}" if cer_vals else "0.0000"


def _write_human_verified(ws, all_data, gt):
    if not gt:
        return
    row = 1
    ws.cell(row=row, column=1, value="Human-Verified Results").font = Font(bold=True, size=14)
    row += 2

    ws.cell(row=row, column=1, value="EXTRACTION COVERAGE & FIELD-LEVEL ACCURACY").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    headers = ["Metric"] + [d["label"] for d in all_data.values()]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    total_gt = len(gt)
    metrics_list = [
        ("Total GT Rows", lambda a: total_gt),
        ("Matched Rows", lambda a: d[a]["gt"]["gt_matched"]),
        ("Missed Rows", lambda a: d[a]["gt"]["gt_not_found"]),
        ("Duplicate Rows", lambda a: len(d[a]["gt"]["duplicates"])),
        ("Extra Rows (Hallucinations)", lambda a: len(d[a]["gt"]["extra"])),
        ("Total Extracted Rows", lambda a: len(d[a]["rows"])),
        ("Date Accuracy", lambda a: d[a]["gt"]["gt_matched"]),  # simplified
        ("Hours Accuracy (±15min)", lambda a: f"{d[a]['gt']['gt_h']*100:.1f}%"),
        ("Time-In Accuracy (±30min)", lambda a: f"{d[a]['gt']['gt_ti']*100:.1f}%"),
        ("Time-Out Accuracy (±30min)", lambda a: f"{d[a]['gt']['gt_to']*100:.1f}%"),
        ("Fully Correct", lambda a: d[a]["gt"]["fully_correct_count"]),
        ("False Accepts", lambda a: d[a]["gt"]["false_accepts"]),
        ("Missed Accepts", lambda a: d[a]["gt"]["missed_accepts"]),
    ]

    for label, fn in metrics_list:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        for i, (aid, d_) in enumerate(all_data.items()):
            fill = FILLS.get(aid, None)
            _style_cell(ws, row, i + 2, fill)
            ws.cell(row=row, column=i + 2, value=fn(aid))
        row += 1

    # Per-row detail
    row += 1
    ws.cell(row=row, column=1, value="PER-ROW DETAILED COMPARISON").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    gt_by_source = defaultdict(list)
    for g in gt:
        gt_by_source[g.get("source_file", "")].append(g)

    detail_headers = ["Source", "GT Date", "GT Time-In", "GT Time-Out", "GT Hours"]
    for aid, dd in all_data.items():
        short = _short_label(dd["label"])
        detail_headers.append(f"Hours ({short})")
        detail_headers.append(f"Match ({short})")
        detail_headers.append(f"Status ({short})")

    for c, h in enumerate(detail_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    for source in sorted(gt_by_source.keys()):
        for gt_row in gt_by_source[source]:
            gt_date = gt_row["_date"] or str(gt_row.get("date", ""))
            gt_hours = gt_row["_hours"]
            ws.cell(row=row, column=1, value=_anon_source(source)).border = THIN_BORDER
            ws.cell(row=row, column=2, value=gt_date).border = THIN_BORDER
            ws.cell(row=row, column=3, value=str(gt_row.get("time_in", "") or "")).border = THIN_BORDER
            ws.cell(row=row, column=4, value=str(gt_row.get("time_out", "") or "")).border = THIN_BORDER
            ws.cell(row=row, column=5, value=gt_hours).border = THIN_BORDER

            col = 6
            for aid, dd in all_data.items():
                fill_color = FILLS.get(aid, None)
                gt_key = (source, gt_row["_date"])
                matched = None
                for m in dd["gt"]["matched_list"]:
                    if m.get("_source") == source and _parse_date(m.get("Date")) == gt_row["_date"]:
                        matched = m
                        break

                if matched and not matched.get("_not_extracted"):
                    hours = matched.get("_hours", "")
                    fc = matched.get("_fully_correct", False)
                    status = matched.get("_status", "")
                    match_text = "✓" if fc else "✗"
                    match_fill = MATCH_FILL if fc else MISMATCH_FILL
                else:
                    hours = ""
                    match_text = "✗"
                    match_fill = MISMATCH_FILL
                    status = "not extracted"

                _style_cell(ws, row, col, fill_color)
                ws.cell(row=row, column=col, value=hours)
                col += 1
                _style_cell(ws, row, col, match_fill)
                ws.cell(row=row, column=col, value=match_text)
                col += 1
                sf = MATCH_FILL if status == "accepted" else MISMATCH_FILL
                _style_cell(ws, row, col, sf)
                ws.cell(row=row, column=col, value=status)
                col += 1

            row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    for i in range(5, len(detail_headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20


def _write_kpi_dashboard(ws, all_data, gt):
    if not gt:
        return
    row = 1
    ws.cell(row=row, column=1, value="KPI Dashboard").font = Font(bold=True, size=14)
    row += 2

    headers = ["Metric"] + [d["label"] for d in all_data.values()]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    total_gt = len(gt)
    kpis = [
        ("Total GT Rows", lambda a: total_gt),
        ("Hours Match (±0.25h)", lambda a: f"{d[a]['gt']['gt_h']*100:.1f}%"),
        ("Time-In Match (±30min)", lambda a: f"{d[a]['gt']['gt_ti']*100:.1f}%"),
        ("Time-Out Match (±30min)", lambda a: f"{d[a]['gt']['gt_to']*100:.1f}%"),
        ("Fully Correct (all 3)", lambda a: f"{d[a]['gt']['gt_fc']*100:.1f}%"),
        ("Rows Matched", lambda a: d[a]["gt"]["gt_matched"]),
        ("Rows Not Found", lambda a: d[a]["gt"]["gt_not_found"]),
        ("False Accepts", lambda a: d[a]["gt"]["false_accepts"]),
        ("Missed Accepts", lambda a: d[a]["gt"]["missed_accepts"]),
    ]

    for label, fn in kpis:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        for i, (aid, dd) in enumerate(all_data.items()):
            fill = FILLS.get(aid, None)
            _style_cell(ws, row, i + 2, fill)
            ws.cell(row=row, column=i + 2, value=fn(aid))
        row += 1

    ws.column_dimensions["A"].width = 30
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20


def _write_per_row_detail(ws, all_data, gt):
    if not gt:
        return
    row = 1
    ws.cell(row=row, column=1, value="Per-Row Detail").font = Font(bold=True, size=14)
    row += 2

    headers = ["Source", "GT Date", "GT Hours"]
    for aid, dd in all_data.items():
        short = _short_label(dd["label"])
        headers.append(f"Hours ({short})")
        headers.append(f"Winner ({short})")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    gt_by_source = defaultdict(list)
    for g in gt:
        gt_by_source[g.get("source_file", "")].append(g)

    for source in sorted(gt_by_source.keys()):
        for gt_row in gt_by_source[source]:
            gt_date = gt_row["_date"]
            gt_hours = gt_row["_hours"]
            ws.cell(row=row, column=1, value=_anon_source(source)).border = THIN_BORDER
            ws.cell(row=row, column=2, value=gt_date).border = THIN_BORDER
            ws.cell(row=row, column=3, value=gt_hours).border = THIN_BORDER

            col = 4
            best_score = 999999
            best_aid = None
            for aid, dd in all_data.items():
                gt_key = (source, gt_date)
                found = None
                for ar in dd["rows"]:
                    if ar["_source"] == source and ar["_date"] == gt_date:
                        found = ar
                        break

                fill_color = FILLS.get(aid, None)
                if found:
                    hours = found.get("_hours", "")
                    h_val = _parse_float(hours)
                    score = abs(h_val - float(gt_hours)) if (h_val is not None and gt_hours is not None) else 999
                    winner = "✓" if score == min(best_score, score) else ""
                    if score < best_score:
                        best_score = score
                        best_aid = aid
                else:
                    hours = ""
                    score = 999
                    winner = ""

                _style_cell(ws, row, col, fill_color)
                ws.cell(row=row, column=col, value=hours)
                col += 1
                if winner:
                    _style_cell(ws, row, col, MATCH_FILL)
                ws.cell(row=row, column=col, value=winner)
                col += 1

            row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    for i in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20


# ── Main ─────────────────────────────────────────────────────────────

# Global dict for sheet writers to access
d = {}

def main():
    global d

    print("Loading per-approach data...")
    all_data = {}
    for aid, label, color in APPROACHES:
        rows, summary = load_approach_data(aid)
        print(f"  {label}: {len(rows)} rows")
        all_data[aid] = {"label": label, "summary": summary, "rows": rows, "_color": color}
    d = all_data

    print("\nLoading ground truth...")
    gt = load_ground_truth()
    print(f"  {len(gt)} GT rows")

    print("\nComputing GT metrics...")
    for aid, dd in all_data.items():
        dd["gt"] = compute_gt_metrics(dd["rows"], gt)
        m = dd["gt"]
        print(f"  {dd['label']}: matched={m['gt_matched']} not_found={m['gt_not_found']} "
              f"h={m['gt_h']*100:.1f}% ti={m['gt_ti']*100:.1f}% to={m['gt_to']*100:.1f}% fc={m['gt_fc']*100:.1f}%")

    # Backup
    if COMBINED_PATH.exists():
        shutil.copy2(COMBINED_PATH, BACKUP_PATH)
        print(f"\nBacked up existing file to {BACKUP_PATH}")

    COMBINED_DIR.mkdir(parents=True, exist_ok=True)

    # Create workbook
    wb = openpyxl.Workbook()

    # Sheet 1: Approach Comparison
    ws1 = wb.active
    ws1.title = "Approach Comparison"
    _write_approach_comparison(ws1, all_data)
    print("\nCreated 'Approach Comparison' sheet")

    # Sheet 2: Human-Verified Results
    if gt:
        ws2 = wb.create_sheet("Human-Verified Results")
        _write_human_verified(ws2, all_data, gt)
        print("Created 'Human-Verified Results' sheet")

    # Sheet 3: IEEE Paper Results
    if gt:
        ws3 = wb.create_sheet("IEEE Paper Results")
        _write_ieee_paper(ws3, all_data, gt)
        print("Created 'IEEE Paper Results' sheet")

    # Sheet 4: Per-Row Detail
    if gt:
        ws4 = wb.create_sheet("Per-Row Detail")
        _write_per_row_detail(ws4, all_data, gt)
        print("Created 'Per-Row Detail' sheet")

    wb.save(COMBINED_PATH)
    print(f"\nSaved: {COMBINED_PATH}")
    print(f"Backup: {BACKUP_PATH}")


def _write_ieee_paper(ws, all_data, gt):
    row = 1
    ws.cell(row=row, column=1, value="IEEE Paper Results — Comparative OCR Evaluation").font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True, size=10)
    row += 2

    # Section A: Performance
    ws.cell(row=row, column=1, value="A. PERFORMANCE METRICS").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    headers = ["Metric"] + [d_[ "label"] for d_ in all_data.values()]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    perf_metrics = [
        ("Total Processing Time (s)", lambda a: f"{d[a]['summary'].get('Total Processing Time (s)', 0):.2f}"),
        ("Pages Processed", lambda a: str(d[a]['summary'].get('Number of Pages', 0))),
        ("Rows Extracted", lambda a: str(d[a]['summary'].get('Total Rows Extracted', 0))),
        ("VLM Fallbacks", lambda a: str(d[a]['summary'].get('VLM Fallbacks Triggered', 0))),
        ("Mean CER", lambda a: _compute_mean_cer(d[a])),
    ]
    for label, fn in perf_metrics:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        for i, (aid, dd) in enumerate(all_data.items()):
            fill = FILLS.get(aid, None)
            _style_cell(ws, row, i + 2, fill)
            ws.cell(row=row, column=i + 2, value=fn(aid))
        row += 1

    row += 1

    # Section B: Accuracy
    ws.cell(row=row, column=1, value="B. ACCURACY METRICS (vs Ground Truth)").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    acc_metrics = [
        ("GT Hours Accuracy (±15min)", lambda a: f"{d[a]['gt']['gt_h']*100:.1f}%" if d[a]['gt']['gt_h'] > 0 else "N/A"),
        ("GT Time-In Accuracy (±30min)", lambda a: f"{d[a]['gt']['gt_ti']*100:.1f}%" if d[a]['gt']['gt_ti'] > 0 else "N/A"),
        ("GT Time-Out Accuracy (±30min)", lambda a: f"{d[a]['gt']['gt_to']*100:.1f}%" if d[a]['gt']['gt_to'] > 0 else "N/A"),
        ("GT Fully Correct Rate", lambda a: f"{d[a]['gt']['gt_fc']*100:.1f}%" if d[a]['gt']['gt_fc'] > 0 else "N/A"),
        ("Rows Matched", lambda a: str(d[a]['gt']['gt_matched'])),
        ("Rows Not Found", lambda a: str(d[a]['gt']['gt_not_found'])),
    ]
    for label, fn in acc_metrics:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        for i, (aid, dd) in enumerate(all_data.items()):
            fill = FILLS.get(aid, None)
            _style_cell(ws, row, i + 2, fill)
            ws.cell(row=row, column=i + 2, value=fn(aid))
        row += 1

    row += 1

    # Section C: Quality
    ws.cell(row=row, column=1, value="C. QUALITY & PRIVACY METRICS").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    qual_metrics = [
        ("Hours Mismatch Rate", lambda a: _compute_hours_mismatch(d[a])),
        ("Field Missing Rate", lambda a: _compute_field_missing(d[a])),
        ("False Accepts", lambda a: str(d[a]['gt']['false_accepts'])),
        ("Missed Accepts", lambda a: str(d[a]['gt']['missed_accepts'])),
        ("Duplicate Extractions", lambda a: str(len(d[a]['gt']['duplicates']))),
        ("Extra Rows (not in GT)", lambda a: str(len(d[a]['gt']['extra']))),
    ]
    for label, fn in qual_metrics:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        for i, (aid, dd) in enumerate(all_data.items()):
            fill = FILLS.get(aid, None)
            _style_cell(ws, row, i + 2, fill)
            ws.cell(row=row, column=i + 2, value=fn(aid))
        row += 1

    ws.column_dimensions["A"].width = 38
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22


if __name__ == "__main__":
    main()
