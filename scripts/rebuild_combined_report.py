"""Rebuild combined benchmark report from all approach outputs.

Reads each approach's Row-Level benchmark data using column NAMES
(never positional indices). Produces a clean, correct combined report.

Output: output/combined/benchmark_combined_v2.xlsx

Usage:
    uv run python scripts/rebuild_combined_report.py
"""

import glob
import os
import re
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "output")
COMBINED_DIR = os.path.join(OUTPUT_DIR, "combined")
BENCH_OUTPUT = os.path.join(COMBINED_DIR, "benchmark_combined_v2.xlsx")
GROUND_TRUTH_PATH = os.path.join(PROJECT_ROOT, "ground_truth.xlsx")

HOURS_TOLERANCE = 0.25
TIME_TOLERANCE_MIN = 30

APPROACHES = [
    ("ocr_only", "OCR Only", "E8E8E8"),
    ("ppocr_grid", "OCR+VLM", "E2EFDA"),
    ("vlm_full_page", "VLM Full", "D6E4F0"),
    ("layout_guided_vlm_local", "Layout Local", "FFF2CC"),
    ("layout_guided_vlm_cloud", "Layout Cloud", "FCE4EC"),
]

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MISMATCH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PARTIAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

FILLS = {}
for _, _, color in APPROACHES:
    FILLS[color] = PatternFill(start_color=color, end_color=color, fill_type="solid")

ANON_MAP = {
    "patient_a_week1": "File 1 (Week 1)",
    "patient_b_week2": "File 2 (Week 2)",
    "patient_c_week3": "File 3 (Week 3)",
}


# ── Helpers ──────────────────────────────────────────────────────────

def _short_label(label, max_len=16):
    if len(label) <= max_len:
        return label
    suffix = label.split("(")[-1].rstrip(")") if "(" in label else label.split()[-1]
    prefix_max = max_len - len(suffix) - 3
    prefix = label[:prefix_max].rsplit(" ", 1)[0] if " " in label[:prefix_max] else label[:prefix_max]
    return f"{prefix} ({suffix})"


def _parse_time(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.hour * 60 + val.minute
    if isinstance(val, (int, float)):
        v = int(val)
        hour, minute = v // 100, v % 100
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return hour * 60 + minute
        return None
    val = str(val).strip()
    if not val:
        return None
    match = re.match(r"(\d{1,2}):?(\d{2})?\s*(AM|PM|am|pm|Am|Pm)?", val)
    if match:
        hour = int(match.group(1))
        minute = int(match.group(2) or 0)
        ampm = match.group(3)
        if ampm and ampm.lower() == "pm" and hour != 12:
            hour += 12
        elif ampm and ampm.lower() == "am" and hour == 12:
            hour = 0
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return hour * 60 + minute
    try:
        v = int(val.replace(":", ""))
        hour, minute = v // 100, v % 100
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return hour * 60 + minute
    except (ValueError, TypeError):
        pass
    return None


def _parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    val = str(val).strip()
    if not val:
        return None
    match = re.match(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", val)
    if match:
        month, day, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
        if year < 100:
            year += 2000
        return f"{year:04d}-{month:02d}-{day:02d}"
    match = re.match(r"(\d{4})-(\d{2})-(\d{2})", val)
    if match:
        return val
    return val


def _compute_hours(time_in_min, time_out_min):
    if time_in_min is None or time_out_min is None:
        return None
    diff = time_out_min - time_in_min
    if diff < 0:
        diff += 24 * 60
    return diff / 60.0


def _anon_source(source):
    for key, label in ANON_MAP.items():
        if key in str(source):
            return label
    return str(source)


def style_cell(ws, row, col, fill=None, alignment=None):
    cell = ws.cell(row=row, column=col)
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    return cell


# ── Load approach data ───────────────────────────────────────────────

def load_approach_data(approach_id):
    """Load all Row-Level rows from all benchmark files for an approach.

    Uses column NAMES (never positional indices).
    Also loads merged_results.xlsx to fill in missing hours.
    Returns list of dicts + summary metrics dict.
    """
    bench_files = sorted(
        f for f in glob.glob(os.path.join(OUTPUT_DIR, approach_id, "benchmark_*.xlsx"))
    )
    all_rows = []
    summary = {}

    # Load merged_results.xlsx to fill missing hours
    merged_path = os.path.join(OUTPUT_DIR, approach_id, "merged_results.xlsx")
    merged_hours = {}  # (source, date) -> hours
    if os.path.exists(merged_path):
        try:
            mwb = openpyxl.load_workbook(merged_path, read_only=True, data_only=True)
            mws = mwb.active
            mheader = None
            for mrow in mws.iter_rows(values_only=True):
                if mheader is None:
                    mheader = [str(c).strip() if c else "" for c in mrow]
                    continue
                mrec = dict(zip(mheader, mrow))
                source = str(mrec.get("Source File", ""))
                date = _parse_date(mrec.get("Date", ""))
                hours = mrec.get("Total Hours") or mrec.get("Calculated Hours")
                if date:
                    merged_hours[(source, date)] = hours
            mwb.close()
        except Exception:
            pass

    for path in bench_files:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue

        # Load summary metrics
        if "Run Summary" in wb.sheetnames:
            ws = wb["Run Summary"]
            for row in ws.iter_rows(min_row=1, values_only=False):
                key = row[0].value
                val = row[1].value
                if key:
                    summary[str(key)] = val

        # Load row-level data using column names
        if "Row-Level" in wb.sheetnames:
            ws = wb["Row-Level"]
            header = None
            for row in ws.iter_rows(values_only=True):
                if header is None:
                    header = [str(c).strip() if c else "" for c in row]
                    continue
                if not any(row):
                    continue
                record = dict(zip(header, row))
                # Normalize key fields
                record["_date"] = _parse_date(record.get("Parsed Date", ""))
                record["_status"] = str(record.get("Status", "")).strip().lower()
                record["_time_in"] = record.get("Parsed Time In", "")
                record["_time_out"] = record.get("Parsed Time Out", "")
                record["_source"] = str(record.get("Source File", ""))

                # Fill missing hours from merged_results.xlsx
                hours = record.get("Parsed Hours")
                if hours is None or str(hours).strip() == "":
                    key = (record["_source"], record["_date"])
                    hours = merged_hours.get(key)
                record["_hours"] = hours

                all_rows.append(record)

        wb.close()

    # Compute aggregate metrics from actual row data (not just last file's Run Summary)
    summary["Total Rows Extracted"] = len(all_rows)
    accepted = sum(1 for r in all_rows if r["_status"] == "accepted")
    flagged = sum(1 for r in all_rows if r["_status"] == "flagged")
    failed = sum(1 for r in all_rows if r["_status"] == "failed")
    summary["Accepted Rows"] = accepted
    summary["Flagged Rows"] = flagged
    summary["Failed Rows"] = failed

    # Compute average confidence from actual data
    confs = []
    for r in all_rows:
        for key in ["Date Confidence", "Time In Confidence", "Time Out Confidence", "Hours Confidence"]:
            v = r.get(key)
            if v is not None:
                try:
                    confs.append(float(v))
                except (ValueError, TypeError):
                    pass
    if confs:
        summary["Mean Overall Confidence"] = round(sum(confs) / len(confs), 4)
        summary["Min Overall Confidence"] = round(min(confs), 4)

    return all_rows, summary


def load_ground_truth():
    if not os.path.exists(GROUND_TRUTH_PATH):
        return None
    wb = openpyxl.load_workbook(GROUND_TRUTH_PATH, read_only=True, data_only=True)
    ws = wb.active
    header = None
    rows = []
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [str(c).strip() if c else "" for c in row]
            continue
        if not any(row):
            continue
        record = dict(zip(header, row))
        record["_date"] = _parse_date(record.get("date", ""))
        record["_time_in_min"] = _parse_time(record.get("time_in"))
        record["_time_out_min"] = _parse_time(record.get("time_out"))
        gt_hours = record.get("total_hours")
        if gt_hours is None:
            gt_hours = _compute_hours(record["_time_in_min"], record["_time_out_min"])
        record["_hours"] = gt_hours
        rows.append(record)
    wb.close()
    return rows


# ── Sheet 1: Approach Comparison ─────────────────────────────────────

def create_approach_comparison(ws, row, all_data):
    """Create the summary + row-level comparison sheet."""
    ws.cell(row=row, column=1, value="Approach Comparison: Handwritten Timesheet OCR").font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True, size=10)
    row += 2

    # Summary metrics
    ws.cell(row=row, column=1, value="SUMMARY METRICS").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    headers = ["Metric"] + [d["label"] for d in all_data.values()]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    metrics_list = [
        ("Total Processing Time (s)", "Total Processing Time (s)"),
        ("Pages Processed", "Number of Pages"),
        ("Rows Extracted", "Total Rows Extracted"),
        ("Accepted Rows", "Accepted Rows"),
        ("Flagged Rows", "Flagged Rows"),
        ("Failed Rows", "Failed Rows"),
        ("Mean Confidence", "Mean Overall Confidence"),
        ("Min Confidence", "Min Overall Confidence"),
        ("VLM Fallbacks Triggered", "VLM Fallbacks Triggered"),
        ("Hours Mismatch Rate", "Hours Mismatch Rate"),
        ("Field Missing Rate", "Field Missing Rate"),
        ("Mean CER", "Mean Character Error Rate"),
    ]

    for label, key in metrics_list:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        for i, (folder, d) in enumerate(all_data.items()):
            val = d["summary"].get(key, "N/A")
            fill = FILLS.get(d.get("_color", ""), None)
            style_cell(ws, row, i + 2, fill)
            ws.cell(row=row, column=i + 2, value=val)
        row += 1

    row += 1

    # Row-level comparison
    ws.cell(row=row, column=1, value="ROW-LEVEL COMPARISON (by date)").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    # Build per-approach lookup by (source, date)
    approach_by_date = {}
    for folder, d in all_data.items():
        by_date = {}
        for r in d["rows"]:
            date_key = r["_date"] or ""
            source_key = r["_source"]
            by_date[(source_key, date_key)] = r
        approach_by_date[folder] = by_date

    # Collect all unique (source, date) pairs
    all_keys = set()
    for by_date in approach_by_date.values():
        all_keys.update(by_date.keys())
    all_keys = sorted(all_keys, key=lambda x: (x[0], x[1]))

    comp_headers = ["Source", "Date"]
    for folder, d in all_data.items():
        short = _short_label(d["label"])
        comp_headers.append(f"Hours ({short})")
        comp_headers.append(f"Status ({short})")

    for c, h in enumerate(comp_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    for source, date_val in all_keys:
        ws.cell(row=row, column=1, value=_anon_source(source)).border = THIN_BORDER
        ws.cell(row=row, column=2, value=date_val).border = THIN_BORDER

        col = 3
        for folder, d in all_data.items():
            r = approach_by_date[folder].get((source, date_val))
            fill_color = FILLS.get(d.get("_color", ""), None)

            if r:
                hours = str(r["_hours"]) if r["_hours"] is not None else ""
                status = r["_status"] or ""
            else:
                hours = ""
                status = "not extracted"

            style_cell(ws, row, col, fill_color)
            ws.cell(row=row, column=col, value=hours)
            col += 1

            status_fill = fill_color
            if status == "accepted":
                status_fill = MATCH_FILL
            elif status == "not extracted":
                status_fill = MISMATCH_FILL
            style_cell(ws, row, col, status_fill)
            ws.cell(row=row, column=col, value=status)
            col += 1

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    for i in range(3, len(comp_headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22

    return row


# ── Sheet 2: Human-Verified Results ──────────────────────────────────

def create_ground_truth_comparison(ws, row, all_data, ground_truth):
    """Create ground truth comparison sheet."""
    ws.cell(row=row, column=1, value="Human-Verified Results").font = Font(bold=True, size=14)
    row += 2

    # Compare each approach to ground truth
    all_results = {}
    for approach_id in [a[0] for a in APPROACHES]:
        approach_rows = all_data[approach_id]["rows"]
        all_results[approach_id] = _compare_approach_to_gt(ground_truth, approach_rows)

    # Coverage & accuracy metrics
    ws.cell(row=row, column=1, value="EXTRACTION COVERAGE & FIELD-LEVEL ACCURACY").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    gt_headers = ["Metric"] + [d["label"] for d in all_data.values()]
    for c, h in enumerate(gt_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    total_gt = len(ground_truth)
    metrics_list = [
        ("Total GT Rows", lambda a: total_gt),
        ("Matched Rows", lambda a: all_results[a]["matched_count"]),
        ("Missed Rows", lambda a: all_results[a]["missed_count"]),
        ("Duplicate Rows", lambda a: all_results[a]["duplicate_count"]),
        ("Extra Rows (Hallucinations)", lambda a: all_results[a]["extra_count"]),
        ("Total Extracted Rows", lambda a: all_results[a]["total_approach_rows"]),
        ("Date Accuracy", lambda a: all_results[a]["date_accuracy"]),
        ("Hours Accuracy (±15min)", lambda a: all_results[a]["hours_accuracy"]),
        ("Time-In Accuracy (±30min)", lambda a: all_results[a]["time_in_accuracy"]),
        ("Time-Out Accuracy (±30min)", lambda a: all_results[a]["time_out_accuracy"]),
        ("Fully Correct", lambda a: all_results[a]["fully_correct_count"]),
        ("Fully + Partially Correct", lambda a: all_results[a]["fully_correct_count"] + all_results[a]["partial_count"]),
        ("False Accepts (accepted but wrong)", lambda a: all_results[a]["false_accepts"]),
        ("Missed Accepts (flagged but correct)", lambda a: all_results[a]["missed_accepts"]),
    ]

    for label, fn in metrics_list:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        for i, (folder, d) in enumerate(all_data.items()):
            val = fn(folder)
            fill = FILLS.get(d.get("_color", ""), None)
            style_cell(ws, row, i + 2, fill)
            ws.cell(row=row, column=i + 2, value=val)
        row += 1

    row += 1

    # Per-row detail
    ws.cell(row=row, column=1, value="PER-ROW DETAILED COMPARISON").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    detail_headers = ["Source", "GT Date", "GT Time-In", "GT Time-Out", "GT Hours"]
    for folder, d in all_data.items():
        short = _short_label(d["label"])
        detail_headers.append(f"Hours ({short})")
        detail_headers.append(f"Match ({short})")
        detail_headers.append(f"Status ({short})")

    for c, h in enumerate(detail_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    # Group GT rows by source file
    gt_by_source = {}
    for gt_row in ground_truth:
        source = gt_row.get("source_file", "")
        gt_by_source.setdefault(source, []).append(gt_row)

    for source in sorted(gt_by_source.keys()):
        for gt_row in gt_by_source[source]:
            gt_date = gt_row["_date"] or str(gt_row.get("date", ""))
            gt_hours = gt_row["_hours"]
            gt_time_in = str(gt_row.get("time_in", "") or "")
            gt_time_out = str(gt_row.get("time_out", "") or "")

            ws.cell(row=row, column=1, value=_anon_source(source)).border = THIN_BORDER
            ws.cell(row=row, column=2, value=gt_date).border = THIN_BORDER
            ws.cell(row=row, column=3, value=gt_time_in).border = THIN_BORDER
            ws.cell(row=row, column=4, value=gt_time_out).border = THIN_BORDER
            ws.cell(row=row, column=5, value=gt_hours).border = THIN_BORDER

            col = 6
            for folder, d in all_data.items():
                fill_color = FILLS.get(d.get("_color", ""), None)
                matched = all_results[folder].get("matched_by_key", {}).get(gt_date)

                if matched:
                    hours = matched.get("_hours", "")
                    is_fully = matched.get("_fully_correct", False)
                    is_partial = matched.get("_partially_correct", False)
                    status = matched.get("_status", "")

                    match_text = "✓" if is_fully else ("~" if is_partial else "✗")
                    match_fill = MATCH_FILL if is_fully else (PARTIAL_FILL if is_partial else MISMATCH_FILL)
                else:
                    hours = ""
                    match_text = "✗"
                    match_fill = MISMATCH_FILL
                    status = "not extracted"

                style_cell(ws, row, col, fill_color)
                ws.cell(row=row, column=col, value=hours)
                col += 1

                style_cell(ws, row, col, match_fill)
                ws.cell(row=row, column=col, value=match_text)
                col += 1

                status_fill = fill_color
                if status == "accepted":
                    status_fill = MATCH_FILL
                elif status == "not extracted":
                    status_fill = MISMATCH_FILL
                style_cell(ws, row, col, status_fill)
                ws.cell(row=row, column=col, value=status)
                col += 1

            row += 1

    # Duplicate rows section
    has_duplicates = any(all_results[a]["duplicates"] for a in all_results)
    if has_duplicates:
        row += 1
        ws.cell(row=row, column=1, value="DUPLICATE ROWS (extra extractions for same GT date)").font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).fill = SECTION_FILL
        row += 1
        dup_headers = ["Approach", "Date", "Hours", "Status", "Employee"]
        for c, h in enumerate(dup_headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = THIN_BORDER
        row += 1

        for folder, d in all_data.items():
            for dup in all_results[folder].get("duplicates", []):
                ws.cell(row=row, column=1, value=d["label"]).border = THIN_BORDER
                ws.cell(row=row, column=2, value=dup.get("_date", "")).border = THIN_BORDER
                ws.cell(row=row, column=3, value=dup.get("_hours", "")).border = THIN_BORDER
                ws.cell(row=row, column=4, value=dup.get("_status", "")).border = THIN_BORDER
                ws.cell(row=row, column=5, value=dup.get("Employee Name", "")).border = THIN_BORDER
                row += 1

    # Extra rows section
    has_extras = any(all_results[a]["extra"] for a in all_results)
    if has_extras:
        row += 1
        ws.cell(row=row, column=1, value="EXTRA ROWS (dates not in ground truth)").font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).fill = SECTION_FILL
        row += 1
        extra_headers = ["Approach", "Date", "Hours", "Status", "Employee"]
        for c, h in enumerate(extra_headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = THIN_BORDER
        row += 1

        for folder, d in all_data.items():
            for extra in all_results[folder].get("extra", []):
                ws.cell(row=row, column=1, value=d["label"]).border = THIN_BORDER
                ws.cell(row=row, column=2, value=extra.get("_date", "")).border = THIN_BORDER
                ws.cell(row=row, column=3, value=extra.get("_hours", "")).border = THIN_BORDER
                ws.cell(row=row, column=4, value=extra.get("_status", "")).border = THIN_BORDER
                ws.cell(row=row, column=5, value=extra.get("Employee Name", "")).border = THIN_BORDER
                row += 1

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    for i in range(3, len(detail_headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    return row


def _compare_approach_to_gt(ground_truth, approach_rows):
    """Compare approach output against ground truth.

    Matches by (source_file, date). For each GT row, finds the BEST
    matching approach row (closest hours). Returns structured results.
    """
    gt_by_key = {}
    for gt_row in ground_truth:
        source = str(gt_row.get("source_file", ""))
        date = gt_row["_date"]
        key = (source, date)
        gt_by_key[key] = gt_row

    # Normalize approach rows: compute (source, date) key
    approach_by_key = {}
    for ar in approach_rows:
        source = ar["_source"]
        date = ar["_date"]
        key = (source, date)
        approach_by_key.setdefault(key, []).append(ar)

    matched_list = []
    matched_by_key = {}
    duplicates = []
    extra_list = []
    total_approach_rows = len(approach_rows)

    for gt_key, gt_row in gt_by_key.items():
        gt_hours = gt_row["_hours"]
        gt_time_in_min = gt_row["_time_in_min"]
        gt_time_out_min = gt_row["_time_out_min"]
        gt_hours_calc = _compute_hours(gt_time_in_min, gt_time_out_min)
        gt_hours_val = gt_hours_calc if gt_hours_calc is not None else gt_hours

        candidates = approach_by_key.get(gt_key, [])
        if not candidates:
            matched_list.append({
                "_date": gt_row["_date"],
                "_hours": None,
                "_status": "not extracted",
                "_fully_correct": False,
                "_partially_correct": False,
                "_not_extracted": True,
            })
            continue

        # Score each candidate
        scored = []
        for ar in candidates:
            ext_hours = ar["_hours"]
            ext_time_in_min = _parse_time(ar["_time_in"])
            ext_time_out_min = _parse_time(ar["_time_out"])

            h_dist = 999
            if ext_hours is not None and gt_hours_val is not None:
                try:
                    h_dist = abs(float(ext_hours) - float(gt_hours_val))
                except (ValueError, TypeError):
                    h_dist = 999

            ti_dist = 9999
            if ext_time_in_min is not None and gt_time_in_min is not None:
                ti_dist = abs(ext_time_in_min - gt_time_in_min)

            to_dist = 9999
            if ext_time_out_min is not None and gt_time_out_min is not None:
                to_dist = abs(ext_time_out_min - gt_time_out_min)

            score = (h_dist * 1000) + ti_dist + (to_dist * 0.01)
            scored.append((score, ar, ext_hours, ext_time_in_min, ext_time_out_min))

        scored.sort(key=lambda x: x[0])
        _, best, best_hours, best_ti, best_to = scored[0]

        # Mark remaining as duplicates
        for _, dup, _, _, _ in scored[1:]:
            duplicates.append(dup)

        # Check field correctness
        hours_ok = False
        if best_hours is not None and gt_hours_val is not None:
            try:
                hours_ok = abs(float(best_hours) - float(gt_hours_val)) <= HOURS_TOLERANCE
            except (ValueError, TypeError):
                pass

        time_in_ok = False
        if best_ti is not None and gt_time_in_min is not None:
            time_in_ok = abs(best_ti - gt_time_in_min) <= TIME_TOLERANCE_MIN

        time_out_ok = False
        if best_to is not None and gt_time_out_min is not None:
            time_out_ok = abs(best_to - gt_time_out_min) <= TIME_TOLERANCE_MIN

        fields_correct = sum([hours_ok, time_in_ok, time_out_ok])
        fully_correct = (fields_correct == 3)
        partially_correct = (0 < fields_correct < 3)
        not_extracted = False

        matched_entry = {
            **best,
            "_fully_correct": fully_correct,
            "_partially_correct": partially_correct,
            "_not_extracted": not_extracted,
            "_hours_ok": hours_ok,
            "_time_in_ok": time_in_ok,
            "_time_out_ok": time_out_ok,
        }
        matched_list.append(matched_entry)
        matched_by_key[gt_row["_date"]] = matched_entry

    # Find extra rows (approach rows with dates not in GT)
    gt_dates_all = set()
    for gt_row in ground_truth:
        gt_dates_all.add((gt_row.get("source_file", ""), gt_row["_date"]))

    used_keys = set()
    for gt_key in gt_by_key:
        used_keys.add(gt_key)

    for ar_key, ar_list in approach_by_key.items():
        if ar_key not in gt_dates_all:
            extra_list.extend(ar_list)

    # Compute accuracy metrics
    matched_count = sum(1 for m in matched_list if not m.get("_not_extracted", True))
    missed_count = sum(1 for m in matched_list if m.get("_not_extracted", False))
    date_correct = sum(1 for m in matched_list if not m.get("_not_extracted", True))
    hours_correct = sum(1 for m in matched_list if m.get("_hours_ok", False))
    time_in_correct_count = sum(1 for m in matched_list if m.get("_time_in_ok", False))
    time_out_correct_count = sum(1 for m in matched_list if m.get("_time_out_ok", False))
    fully_correct_count = sum(1 for m in matched_list if m.get("_fully_correct"))
    partial_count = sum(1 for m in matched_list if m.get("_partially_correct"))

    def _pct(n, total):
        return f"{n}/{total} ({n/total*100:.1f}%)" if total > 0 else "N/A"

    # False accepts: rows marked "accepted" but not fully correct
    false_accepts = 0
    missed_accepts = 0
    for m in matched_list:
        if m.get("_not_extracted"):
            continue
        if m.get("_status") == "accepted" and not m.get("_fully_correct"):
            false_accepts += 1
        if m.get("_status") != "accepted" and m.get("_fully_correct"):
            missed_accepts += 1

    return {
        "matched": matched_list,
        "matched_by_key": matched_by_key,
        "duplicates": duplicates,
        "extra": extra_list,
        "matched_count": matched_count,
        "missed_count": missed_count,
        "duplicate_count": len(duplicates),
        "extra_count": len(extra_list),
        "total_approach_rows": total_approach_rows,
        "date_accuracy": _pct(date_correct, len(matched_list)),
        "hours_accuracy": _pct(hours_correct, len(matched_list)),
        "time_in_accuracy": _pct(time_in_correct_count, matched_count),
        "time_out_accuracy": _pct(time_out_correct_count, matched_count),
        "fully_correct_count": fully_correct_count,
        "partial_count": partial_count,
        "false_accepts": false_accepts,
        "missed_accepts": missed_accepts,
    }


# ── Sheet 3: Time Comparison ─────────────────────────────────────────

def create_time_comparison(ws, row, all_data, ground_truth):
    """Create time-in/time-out comparison sheet."""
    ws.cell(row=row, column=1, value="Time Comparison").font = Font(bold=True, size=14)
    row += 2

    # Time-In section
    ws.cell(row=row, column=1, value="TIME-IN COMPARISON").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    time_headers = ["Source", "GT Date", "GT Time-In"]
    for folder, d in all_data.items():
        short = _short_label(d["label"])
        time_headers.append(f"Time-In ({short})")
        time_headers.append(f"Correct ({short})")

    for c, h in enumerate(time_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    gt_by_source = {}
    for gt_row in ground_truth:
        source = gt_row.get("source_file", "")
        gt_by_source.setdefault(source, []).append(gt_row)

    time_in_correct_total = {a[0]: 0 for a in APPROACHES}
    time_out_correct_total = {a[0]: 0 for a in APPROACHES}

    for source in sorted(gt_by_source.keys()):
        for gt_row in gt_by_source[source]:
            gt_date = gt_row["_date"] or str(gt_row.get("date", ""))
            gt_time_in = str(gt_row.get("time_in", "") or "")
            gt_time_in_min = _parse_time(gt_row.get("time_in"))

            ws.cell(row=row, column=1, value=_anon_source(source)).border = THIN_BORDER
            ws.cell(row=row, column=2, value=gt_date).border = THIN_BORDER
            ws.cell(row=row, column=3, value=gt_time_in).border = THIN_BORDER

            col = 4
            for folder, d in all_data.items():
                fill_color = FILLS.get(d.get("_color", ""), None)

                # Find matching approach row
                match = None
                for ar in all_data[folder]["rows"]:
                    if ar["_date"] == gt_date and ar["_source"] == source:
                        match = ar
                        break

                if match:
                    ext_time_in = str(match.get("Parsed Time In", "") or "")
                    ext_time_in_min = _parse_time(match.get("Parsed Time In"))
                    correct = False
                    if ext_time_in_min is not None and gt_time_in_min is not None:
                        if abs(ext_time_in_min - gt_time_in_min) <= TIME_TOLERANCE_MIN:
                            correct = True
                            time_in_correct_total[folder] += 1
                else:
                    ext_time_in = ""
                    correct = False

                style_cell(ws, row, col, fill_color)
                ws.cell(row=row, column=col, value=ext_time_in)
                col += 1

                check_fill = MATCH_FILL if correct else MISMATCH_FILL
                style_cell(ws, row, col, check_fill)
                ws.cell(row=row, column=col, value="✓" if correct else "✗")
                col += 1

            row += 1

    row += 1

    # Time-Out section
    ws.cell(row=row, column=1, value="TIME-OUT COMPARISON").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    for c, h in enumerate(time_headers, 1):
        h_label = h.replace("Time-In", "Time-Out")
        cell = ws.cell(row=row, column=c, value=h_label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    for source in sorted(gt_by_source.keys()):
        for gt_row in gt_by_source[source]:
            gt_date = gt_row["_date"] or str(gt_row.get("date", ""))
            gt_time_out = str(gt_row.get("time_out", "") or "")
            gt_time_out_min = _parse_time(gt_row.get("time_out"))

            ws.cell(row=row, column=1, value=_anon_source(source)).border = THIN_BORDER
            ws.cell(row=row, column=2, value=gt_date).border = THIN_BORDER
            ws.cell(row=row, column=3, value=gt_time_out).border = THIN_BORDER

            col = 4
            for folder, d in all_data.items():
                fill_color = FILLS.get(d.get("_color", ""), None)

                match = None
                for ar in all_data[folder]["rows"]:
                    if ar["_date"] == gt_date and ar["_source"] == source:
                        match = ar
                        break

                if match:
                    ext_time_out = str(match.get("Parsed Time Out", "") or "")
                    ext_time_out_min = _parse_time(match.get("Parsed Time Out"))
                    correct = False
                    if ext_time_out_min is not None and gt_time_out_min is not None:
                        if abs(ext_time_out_min - gt_time_out_min) <= TIME_TOLERANCE_MIN:
                            correct = True
                            time_out_correct_total[folder] += 1
                else:
                    ext_time_out = ""
                    correct = False

                style_cell(ws, row, col, fill_color)
                ws.cell(row=row, column=col, value=ext_time_out)
                col += 1

                check_fill = MATCH_FILL if correct else MISMATCH_FILL
                style_cell(ws, row, col, check_fill)
                ws.cell(row=row, column=col, value="✓" if correct else "✗")
                col += 1

            row += 1

    row += 1

    # Summary
    ws.cell(row=row, column=1, value="CORRECTNESS SUMMARY").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    row += 1

    total_gt = len(ground_truth)
    summary_headers = ["Metric"] + [d["label"] for d in all_data.values()]
    for c, h in enumerate(summary_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    ws.cell(row=row, column=1, value="Total GT Rows").border = THIN_BORDER
    for i, (folder, d) in enumerate(all_data.items()):
        style_cell(ws, row, i + 2, FILLS.get(d.get("_color", ""), None))
        ws.cell(row=row, column=i + 2, value=total_gt)
    row += 1

    ws.cell(row=row, column=1, value="Time-In Correct").border = THIN_BORDER
    for i, (folder, d) in enumerate(all_data.items()):
        n = time_in_correct_total[folder]
        style_cell(ws, row, i + 2, FILLS.get(d.get("_color", ""), None))
        ws.cell(row=row, column=i + 2, value=f"{n}/{total_gt} ({n/total_gt*100:.1f}%)")
    row += 1

    ws.cell(row=row, column=1, value="Time-Out Correct").border = THIN_BORDER
    for i, (folder, d) in enumerate(all_data.items()):
        n = time_out_correct_total[folder]
        style_cell(ws, row, i + 2, FILLS.get(d.get("_color", ""), None))
        ws.cell(row=row, column=i + 2, value=f"{n}/{total_gt} ({n/total_gt*100:.1f}%)")
    row += 1

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    for i in range(3, len(time_headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20


# ── Main ─────────────────────────────────────────────────────────────

def main():
    print("Loading approach data...")
    all_data = {}
    for folder, label, color in APPROACHES:
        rows, summary = load_approach_data(folder)
        if rows or summary:
            all_data[folder] = {"label": label, "summary": summary, "rows": rows, "_color": color}
            print(f"  {label}: {len(rows)} rows")
        else:
            print(f"  {label}: NO DATA")

    if not all_data:
        print("ERROR: No approach data found. Run approaches first.")
        sys.exit(1)

    print("\nLoading ground truth...")
    ground_truth = load_ground_truth()
    if ground_truth:
        print(f"  Loaded {len(ground_truth)} ground truth rows")
    else:
        print("  No ground truth file found — skipping ground truth sheets")

    # Create output
    os.makedirs(COMBINED_DIR, exist_ok=True)
    wb = openpyxl.Workbook()

    # Sheet 1: Approach Comparison
    ws1 = wb.active
    ws1.title = "Approach Comparison"
    create_approach_comparison(ws1, 1, all_data)
    print(f"\nCreated 'Approach Comparison' sheet")

    # Sheet 2: Human-Verified Results
    if ground_truth:
        ws2 = wb.create_sheet("Human-Verified Results")
        create_ground_truth_comparison(ws2, 1, all_data, ground_truth)
        print("Created 'Human-Verified Results' sheet")

        # Sheet 3: Time Comparison
        ws3 = wb.create_sheet("Time Comparison")
        create_time_comparison(ws3, 1, all_data, ground_truth)
        print("Created 'Time Comparison' sheet")

    wb.save(BENCH_OUTPUT)
    print(f"\nSaved: {BENCH_OUTPUT}")


if __name__ == "__main__":
    main()
