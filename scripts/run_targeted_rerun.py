#!/usr/bin/env python3
"""Run targeted re-run on 14 problem files and collect results."""

import csv
import json
import subprocess
import sys
import time
from pathlib import Path
from datetime import datetime
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
INPUT_DIR = PROJECT_ROOT / "input"
OUTPUT_DIR = PROJECT_ROOT / "output"
RERUN_DIR = OUTPUT_DIR / "band_crop_rerun"

# Mapping of user-provided filenames to actual input filenames
TARGET_FILES = [
    ("n.rivera_timesheets_030426_031026.pdf", "N.Rivera Timesheets -030426-031026.pdf"),
    ("l.moran_timesheet_123125_010626.pdf", "L.Moran Timesheet 123125-010626.pdf"),
    ("s.hanton_timesheet___012826_020326.pdf", "S.Hanton Timesheet - 012826-020326 c.pdf"),
    ("p.derricott_timesheet_010726_011326.pdf", "P.Derricott Timesheet - 010726-011326.pdf"),
    ("s.hanton_timesheet_021126_021726.pdf", "S.Hanton Timesheet - 021126-021726.pdf"),
    ("n.rivera_timesheets_021826_022426.pdf", "N.Rivera Timesheets -021826-022426.pdf"),
    ("r.elliott_timesheets_112625_120225.pdf", "R.Elliott Timesheets 112625-120225.pdf"),
    ("p.derricott_timesheet___012826_020326.pdf", "P.Derricott Timesheet - 012826-020326.pdf"),
    ("l.moran_timesheet_122425_123025.pdf", "L.Moran Timesheet 122425-123025.pdf"),
    ("h.leal_123125_010626.pdf", "H.Leal 123125-010626.pdf"),
    ("k.drewry_timesheets_022526_030326.pdf", "K.Drewry Timesheets 022526-030326.pdf"),
    ("j.jackson_timesheets_121725_122325.pdf", "J.Jackson Timesheets 121725-122325.pdf"),
    ("s.bussa_timesheet_020426_021026.pdf", "S.Bussa Timesheet - 020426-021026.pdf"),
    ("r.elliott_timesheets_122425_123025.pdf", "R.Elliott Timesheets 122425-123025.pdf"),
]

def run_extraction(filename: str) -> dict:
    """Run extraction on a single file and return results."""
    print(f"\n{'='*60}")
    print(f"Processing: {filename}")
    print(f"{'='*60}")
    
    start_time = time.time()
    
    cmd = [
        sys.executable, "-m", "src.main",
        "--file", f"input/{filename}",
        "--verbose"
    ]
    
    result = subprocess.run(
        cmd,
        cwd=PROJECT_ROOT,
        capture_output=True,
        text=True,
        timeout=600  # 10 min timeout per file
    )
    
    elapsed = time.time() - start_time
    
    # Check if successful
    if result.returncode != 0:
        print(f"ERROR: {result.stderr}")
        return {
            "filename": filename,
            "status": "error",
            "error": result.stderr,
            "elapsed": elapsed
        }
    
    # Read the merged results
    results_path = OUTPUT_DIR / "merged_results.xlsx"
    if not results_path.exists():
        return {
            "filename": filename,
            "status": "error",
            "error": "merged_results.xlsx not found",
            "elapsed": elapsed
        }
    
    wb = openpyxl.load_workbook(results_path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
    
    rows = []
    for i in range(2, ws.max_row + 1):
        row = {}
        for j, h in enumerate(headers, 1):
            row[h] = ws.cell(i, j).value
        rows.append(row)
    
    return {
        "filename": filename,
        "status": "extracted",
        "rows": rows,
        "total_rows": len(rows),
        "elapsed": elapsed
    }


def main():
    RERUN_DIR.mkdir(parents=True, exist_ok=True)
    
    all_results = []
    
    for i, (user_name, actual_filename) in enumerate(TARGET_FILES, 1):
        print(f"\n[{i}/{len(TARGET_FILES)}] Running {actual_filename}...")
        
        result = run_extraction(actual_filename)
        all_results.append(result)
        
        # Rate limit between files
        if i < len(TARGET_FILES):
            print("Waiting 5 seconds before next file...")
            time.sleep(5)
    
    # Save raw results
    with open(RERUN_DIR / 'targeted_rerun_raw_results.json', 'w') as f:
        json.dump(all_results, f, indent=2, default=str)
    
    # Print summary
    print("\n" + "="*60)
    print("EXTRACTION COMPLETE")
    print("="*60)
    
    total_extracted = sum(1 for r in all_results if r.get('status') == 'extracted')
    total_errors = sum(1 for r in all_results if r.get('status') == 'error')
    total_rows = sum(r.get('total_rows', 0) for r in all_results)
    
    print(f"Files processed: {len(TARGET_FILES)}")
    print(f"Successful: {total_extracted}")
    print(f"Errors: {total_errors}")
    print(f"Total rows extracted: {total_rows}")
    
    # Create CSV output
    with open(RERUN_DIR / 'targeted_rerun_results.csv', 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['source_file', 'page_number', 'employee_name', 'date', 'time_in', 'time_out', 'total_hours', 'status'])
        
        for r in all_results:
            fn = r.get('filename', '')
            if r.get('status') == 'extracted':
                for row in r.get('rows', []):
                    writer.writerow([
                        row.get('Source File', ''),
                        row.get('Page', ''),
                        row.get('Employee Name', '') or '',
                        row.get('Date', ''),
                        row.get('Time In', ''),
                        row.get('Time Out', ''),
                        row.get('Total Hours', ''),
                        'extracted'
                    ])
            else:
                writer.writerow([fn, '', '', '', '', '', '', r.get('status', 'error')])
    
    print(f"\nResults saved to {RERUN_DIR}/")
    print("Next: Run ground truth comparison...")


if __name__ == "__main__":
    main()