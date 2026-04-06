"""Compare extraction results against ground truth data.

Reads ground truth from output/ground_truth.xlsx, compares against
all approach outputs, and adds a 'Human-Verified Results' sheet to
benchmark_combined.xlsx.

This script computes field-level accuracy (Date, Time In, Time Out, Hours)
and confusion-matrix-style metrics where the pipeline's internal validation
status (accepted/flagged) is treated as a binary prediction of correctness.

Usage:
    uv run python scripts/compare_ground_truth.py
"""

import glob
import os
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GROUND_TRUTH_PATH = "ground_truth.xlsx"
COMBINED_PATH = "output/combined/benchmark_combined.xlsx"
NAME_DB_PATH = "output/name_mapping.db"

APPROACHES = [
    ("ocr_only", "OCR Only"),
    ("ppocr_grid", "OCR + VLM Fallback"),
    ("vlm_full_page", "VLM Full Page"),
    ("layout_guided_vlm_local", "Layout-Guided VLM (Local)"),
    ("layout_guided_vlm_cloud", "Layout-Guided VLM (Cloud)"),
]

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
BEST_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MISMATCH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PARTIAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HOURS_TOLERANCE = 0.25  # ±15 minutes
TIME_TOLERANCE_MIN = 30  # ±30 minutes


def parse_time(val):
    """Parse time value to minutes since midnight."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.hour * 60 + val.minute
    if isinstance(val, (int, float)):
        v = int(val)
        hour = v // 100
        minute = v % 100
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return hour * 60 + minute
        return None
    val = str(val).strip()
    if not val:
        return None
    match = re.match(r"(\d{1,2}):?(\d{2})?\s*(AM|PM|am|pm|Am|Pm)?", val)
    if match:
        hour = int(match.group(1))
        minute = int(match.group(2) or 0)
        ampm = match.group(3)
        if ampm and ampm.lower() == "pm" and hour != 12:
            hour += 12
        elif ampm and ampm.lower() == "am" and hour == 12:
            hour = 0
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return hour * 60 + minute
    try:
        v = int(val.replace(":", ""))
        hour = v // 100
        minute = v % 100
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return hour * 60 + minute
    except (ValueError, TypeError):
        pass
    return None


def parse_date(val):
    """Parse date value to YYYY-MM-DD string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    val = str(val).strip()
    if not val:
        return None
    match = re.match(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", val)
    if match:
        month = int(match.group(1))
        day = int(match.group(2))
        year = int(match.group(3))
        if year < 100:
            year += 2000
        return f"{year:04d}-{month:02d}-{day:02d}"
    match = re.match(r"(\d{4})-(\d{2})-(\d{2})", val)
    if match:
        return val
    return val


def compute_hours_from_times(time_in_minutes, time_out_minutes):
    """Compute hours between two time values. Handles overnight shifts."""
    if time_in_minutes is None or time_out_minutes is None:
        return None
    diff = time_out_minutes - time_in_minutes
    if diff < 0:
        diff += 24 * 60
    return diff / 60.0


def load_ground_truth():
    """Load ground truth data from Excel."""
    if not os.path.exists(GROUND_TRUTH_PATH):
        print(f"ERROR: Ground truth file not found: {GROUND_TRUTH_PATH}")
        print("Please fill in output/ground_truth.xlsx first.")
        return None
    wb = openpyxl.load_workbook(GROUND_TRUTH_PATH, read_only=True)
    ws = wb.active
    rows = []
    header = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [str(c).strip() if c else "" for c in row]
            continue
        if not any(row):
            continue
        record = dict(zip(header, row))
        if record.get("source_file") and record.get("date"):
            rows.append(record)
    wb.close()
    print(f"Loaded {len(rows)} ground truth rows")
    return rows


def load_name_mapping():
    """Load name mapping from SQLite DB."""
    mapping = {}
    if not os.path.exists(NAME_DB_PATH):
        return mapping
    import sqlite3

    conn = sqlite3.connect(NAME_DB_PATH)
    conn.row_factory = sqlite3.Row
    for table, label in [("patients", "patient"), ("employees", "employee")]:
        rows = conn.execute(f"SELECT * FROM {table}").fetchall()
        for row in rows:
            mapping[row["anonymized_id"]] = {
                "real_name": row["real_name"],
                "label": label,
                "source_files": (row["source_files"] or "").split(","),
            }
    conn.close()
    return mapping


def load_approach_data(approach_id):
    """Load all benchmark data for an approach."""
    bench_files = sorted(
        f for f in glob.glob(f"output/{approach_id}/benchmark_*.xlsx")
        if "combined" not in f
    )
    all_rows = []
    for path in bench_files:
        wb = openpyxl.load_workbook(path, read_only=True)
        if "Row-Level" in wb.sheetnames:
            ws = wb["Row-Level"]
            header = None
            for row in ws.iter_rows(values_only=True):
                if header is None:
                    header = [str(c).strip() if c else "" for c in row]
                    continue
                if not any(row):
                    continue
                record = dict(zip(header, row))
                record["Date"] = record.get("Parsed Date", "")
                record["Hours"] = record.get("Parsed Hours", record.get("Written Hours"))
                record["Status"] = record.get("Status", "")
                record["Time In"] = record.get("Parsed Time In", "")
                record["Time Out"] = record.get("Parsed Time Out", "")
                record["Employee"] = record.get("Employee Name", "")
                all_rows.append(record)
        wb.close()

    # Also load merged results to fill in missing data
    merged_path = f"output/{approach_id}/merged_results.xlsx"
    if os.path.exists(merged_path):
        wb = openpyxl.load_workbook(merged_path, read_only=True)
        ws = wb.active
        header = None
        merged_by_date = {}
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c).strip() if c else "" for c in row]
                continue
            if not any(row):
                continue
            record = dict(zip(header, row))
            date = parse_date(record.get("Date", ""))
            if date:
                merged_by_date[date] = record
        wb.close()
        for record in all_rows:
            date = parse_date(record.get("Date", ""))
            if date and date in merged_by_date:
                mr = merged_by_date[date]
                if record["Hours"] is None:
                    record["Hours"] = mr.get("Total Hours")
                if record["Status"] in ("", None):
                    record["Status"] = mr.get("Status", "")
                if not record["Time In"]:
                    record["Time In"] = mr.get("Time In", "")
                if not record["Time Out"]:
                    record["Time Out"] = mr.get("Time Out", "")

    print(f"  {approach_id}: {len(all_rows)} rows loaded")
    return all_rows


def compare_approach(ground_truth, approach_id, name_mapping):
    """Compare approach output against ground truth.

    For each ground truth row, finds the matching extracted row and
    evaluates each field independently:
      - date_correct: exact match
      - time_in_correct: within ±30 min
      - time_out_correct: within ±30 min
      - hours_correct: within ±0.25 hours
      - fully_correct: ALL four fields match
      - partially_correct: at least one field matches but not all
    """
    approach_rows = load_approach_data(approach_id)
    results = []
    for gt_row in ground_truth:
        source_file = gt_row["source_file"]
        gt_date = parse_date(gt_row["date"])
        if gt_date is None:
            gt_date = str(gt_row["date"]).strip()
        gt_hours = gt_row.get("total_hours")
        gt_time_in = gt_row.get("time_in")
        gt_time_out = gt_row.get("time_out")
        gt_employee = gt_row.get("employee_name", "")

        # Find matching row in approach output
        matched = None
        for ar in approach_rows:
            ar_date_raw = ar.get("Date", "")
            ar_date = parse_date(ar_date_raw) if ar_date_raw else str(ar_date_raw).strip()
            if ar_date == gt_date:
                matched = ar
                break

        if matched:
            ext_hours = matched.get("Hours")
            ext_status = matched.get("Status", "")
            ext_time_in = matched.get("Time In", "")
            ext_time_out = matched.get("Time Out", "")
            ext_employee = matched.get("Employee", "")

            # Parse ground truth times
            gt_time_in_min = parse_time(gt_time_in)
            gt_time_out_min = parse_time(gt_time_out)
            gt_hours_calc = compute_hours_from_times(gt_time_in_min, gt_time_out_min)
            gt_hours_val = gt_hours_calc if gt_hours_calc is not None else gt_hours

            # Parse extracted times
            ext_time_in_min = parse_time(ext_time_in)
            ext_time_out_min = parse_time(ext_time_out)

            # Evaluate each field independently
            date_ok = (gt_date is not None and ext_hours is not None)  # matched by date

            hours_ok = False
            if ext_hours is not None and gt_hours_val is not None:
                try:
                    hours_ok = abs(float(ext_hours) - float(gt_hours_val)) <= HOURS_TOLERANCE
                except (ValueError, TypeError):
                    pass

            time_in_ok = False
            if ext_time_in_min is not None and gt_time_in_min is not None:
                time_in_ok = abs(ext_time_in_min - gt_time_in_min) <= TIME_TOLERANCE_MIN

            time_out_ok = False
            if ext_time_out_min is not None and gt_time_out_min is not None:
                time_out_ok = abs(ext_time_out_min - gt_time_out_min) <= TIME_TOLERANCE_MIN

            fields_correct = sum([hours_ok, time_in_ok, time_out_ok])
            fully_correct = (fields_correct == 3)
            partially_correct = (0 < fields_correct < 3)
            not_extracted = False  # row was found

        else:
            ext_hours = None
            ext_status = "not extracted"
            ext_time_in = None
            ext_time_out = None
            ext_employee = None
            date_ok = False
            hours_ok = False
            time_in_ok = False
            time_out_ok = False
            fields_correct = 0
            fully_correct = False
            partially_correct = False
            not_extracted = True

        results.append({
            "source_file": source_file,
            "date": gt_date,
            "gt_hours": gt_hours,
            "gt_time_in": gt_time_in,
            "gt_time_out": gt_time_out,
            "gt_employee": gt_employee,
            "ext_hours": ext_hours,
            "ext_time_in": ext_time_in,
            "ext_time_out": ext_time_out,
            "ext_status": ext_status,
            "ext_employee": ext_employee,
            "date_correct": date_ok,
            "hours_correct": hours_ok,
            "time_in_correct": time_in_ok,
            "time_out_correct": time_out_ok,
            "fields_correct": fields_correct,
            "fully_correct": fully_correct,
            "partially_correct": partially_correct,
            "not_extracted": not_extracted,
        })
    return results


def compute_metrics(all_approach_results):
    """Compute field-level accuracy and pipeline validation quality metrics."""
    metrics = {}
    for approach_id, approach_label, approach_results in all_approach_results:
        total = len(approach_results)
        if total == 0:
            continue

        # Field-level accuracy (per-field correctness rate)
        date_correct = sum(1 for r in approach_results if r["date_correct"])
        hours_correct = sum(1 for r in approach_results if r["hours_correct"])
        time_in_correct = sum(1 for r in approach_results if r["time_in_correct"])
        time_out_correct = sum(1 for r in approach_results if r["time_out_correct"])
        fully_correct = sum(1 for r in approach_results if r["fully_correct"])
        partially_correct = sum(1 for r in approach_results if r["partially_correct"])
        not_extracted = sum(1 for r in approach_results if r["not_extracted"])

        # Pipeline validation quality: does the pipeline's internal
        # "accepted" status correlate with actual correctness?
        accepted = [r for r in approach_results if r["ext_status"].lower() == "accepted"]
        not_accepted = [r for r in approach_results if r["ext_status"].lower() != "accepted"]

        # Of accepted rows, how many are actually fully correct?
        accepted_correct = sum(1 for r in accepted if r["fully_correct"])
        validation_precision = accepted_correct / len(accepted) if accepted else 0.0

        # Of all fully correct rows, how many did the pipeline accept?
        correct_rows = [r for r in approach_results if r["fully_correct"]]
        correct_accepted = sum(1 for r in correct_rows if r["ext_status"].lower() == "accepted")
        validation_recall = correct_accepted / len(correct_rows) if correct_rows else 0.0

        # F1 of the pipeline's internal validation
        if validation_precision + validation_recall > 0:
            validation_f1 = 2 * validation_precision * validation_recall / (validation_precision + validation_recall)
        else:
            validation_f1 = 0.0

        # False acceptance rate: accepted but NOT fully correct
        false_accept = sum(1 for r in accepted if not r["fully_correct"])
        false_accept_rate = false_accept / len(accepted) if accepted else 0.0

        # Missed detection rate: fully correct but flagged/failed
        missed = sum(1 for r in not_accepted if r["fully_correct"])
        missed_rate = missed / len(correct_rows) if correct_rows else 0.0

        metrics[approach_id] = {
            "label": approach_label,
            "total_rows": total,
            "not_extracted": not_extracted,
            "extracted": total - not_extracted,
            # Per-field accuracy
            "date_accuracy": date_correct / total if total else 0.0,
            "hours_accuracy": hours_correct / total if total else 0.0,
            "time_in_accuracy": time_in_correct / total if total else 0.0,
            "time_out_accuracy": time_out_correct / total if total else 0.0,
            # Composite
            "fully_correct": fully_correct,
            "partially_correct": partially_correct,
            "fully_correct_rate": fully_correct / total if total else 0.0,
            "partial_or_full_rate": (fully_correct + partially_correct) / total if total else 0.0,
            # Pipeline validation quality
            "accepted_count": len(accepted),
            "validation_precision": validation_precision,
            "validation_recall": validation_recall,
            "validation_f1": validation_f1,
            "false_accept_rate": false_accept_rate,
            "missed_rate": missed_rate,
        }
    return metrics


def write_human_verified_sheet(wb, metrics, all_approach_results, ground_truth):
    """Add Human-Verified Results sheet with IEEE-quality metrics."""
    ws = wb.create_sheet("Human-Verified Results")

    row = 1
    ws.cell(row=row, column=1, value="Ground Truth Comparison — Extraction Accuracy").font = Font(
        bold=True, size=14
    )
    num_cols = len(metrics) + 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    row += 1
    ws.cell(
        row=row, column=1,
        value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
              f"Tolerance: Hours ±{HOURS_TOLERANCE}h, Time ±{TIME_TOLERANCE_MIN}min",
    ).font = Font(italic=True, size=10)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    row += 2

    # ═══════════════════════════════════════════════════════
    # Section 1: Extraction Coverage & Field-Level Accuracy
    # ═══════════════════════════════════════════════════════
    ws.cell(row=row, column=1, value="EXTRACTION COVERAGE & FIELD-LEVEL ACCURACY").font = Font(
        bold=True, size=12
    )
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    row += 1

    metric_headers = ["Metric"] + [m["label"] for m in metrics.values()]
    for c, h in enumerate(metric_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    section1_rows = [
        ("Total Ground Truth Rows", lambda m: m["total_rows"]),
        ("Rows Extracted", lambda m: m["extracted"]),
        ("Rows Not Extracted", lambda m: m["not_extracted"]),
        ("", lambda m: ""),
        ("Date Accuracy (exact match)", lambda m: f"{m['date_accuracy']:.1%}"),
        ("Hours Accuracy (±0.25h)", lambda m: f"{m['hours_accuracy']:.1%}"),
        ("Time In Accuracy (±30min)", lambda m: f"{m['time_in_accuracy']:.1%}"),
        ("Time Out Accuracy (±30min)", lambda m: f"{m['time_out_accuracy']:.1%}"),
        ("", lambda m: ""),
        ("Fully Correct (all 3 fields)", lambda m: f"{m['fully_correct']} ({m['fully_correct_rate']:.1%})"),
        ("Partial or Full Match", lambda m: f"{m['fully_correct'] + m['partially_correct']} ({m['partial_or_full_rate']:.1%})"),
    ]

    for label, getter in section1_rows:
        if label == "":
            # Separator row
            for c in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.border = THIN_BORDER
            row += 1
            continue
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", wrap_text=True)
        for c, (aid, m) in enumerate(metrics.items(), 2):
            val = getter(m)
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        row += 1

    row += 1

    # ═══════════════════════════════════════════════════════
    # Section 2: Pipeline Validation Quality
    # ═══════════════════════════════════════════════════════
    ws.cell(row=row, column=1, value="PIPELINE VALIDATION QUALITY").font = Font(
        bold=True, size=12
    )
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    row += 1

    metric_headers2 = ["Metric"] + [m["label"] for m in metrics.values()]
    for c, h in enumerate(metric_headers2, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    section2_rows = [
        ("Rows Marked 'Accepted'", lambda m: m["accepted_count"]),
        ("", lambda m: ""),
        ("Validation Precision", lambda m: f"{m['validation_precision']:.1%}"),
        ("  → Of accepted rows, % fully correct", lambda m: ""),
        ("Validation Recall", lambda m: f"{m['validation_recall']:.1%}"),
        ("  → Of correct rows, % accepted", lambda m: ""),
        ("Validation F1", lambda m: f"{m['validation_f1']:.3f}"),
        ("", lambda m: ""),
        ("False Accept Rate", lambda m: f"{m['false_accept_rate']:.1%}"),
        ("  → Accepted but actually wrong", lambda m: ""),
        ("Missed Detection Rate", lambda m: f"{m['missed_rate']:.1%}"),
        ("  → Correct but flagged/failed", lambda m: ""),
    ]

    for label, getter in section2_rows:
        if label.startswith("  →"):
            # Sub-label row
            cell = ws.cell(row=row, column=1, value=label)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left")
            cell.font = Font(italic=True, size=10)
            for c in range(2, num_cols + 1):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1
            continue
        if label == "":
            for c in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.border = THIN_BORDER
            row += 1
            continue
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", wrap_text=True)
        for c, (aid, m) in enumerate(metrics.items(), 2):
            val = getter(m)
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        row += 1

    row += 1

    # ═══════════════════════════════════════════════════════
    # Section 3: Per-Row Detailed Comparison
    # ═══════════════════════════════════════════════════════
    ws.cell(row=row, column=1, value="PER-ROW DETAILED COMPARISON").font = Font(
        bold=True, size=12
    )
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    row += 1

    # Build headers: File | Date | GT: Time In | GT: Time Out | GT: Hours |
    #   For each approach: Hours | ✓/✗ | Status
    detail_headers = [
        "Source File", "Date",
        "GT Time In", "GT Time Out", "GT Hours",
    ]
    for _aid, alabel in APPROACHES:
        detail_headers.append(f"{alabel}: Hours")
        detail_headers.append(f"{alabel}: ✓")
        detail_headers.append(f"{alabel}: Status")

    for c, h in enumerate(detail_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    # Anonymized filename mapping
    anon_map = {
        "patient_a_week1": "File 1 (Week 1)",
        "patient_b_week2": "File 2 (Week 2)",
        "patient_c_week3": "File 3 (Week 3)",
    }

    for gt_row in ground_truth:
        source_file = gt_row["source_file"]
        # Use generic file labels
        anon_name = None
        for key, val in anon_map.items():
            if key in source_file.lower() or source_file.lower() in key:
                anon_name = val
                break
        if anon_name is None:
            anon_name = source_file

        gt_date = parse_date(gt_row["date"])
        if gt_date is None:
            gt_date = str(gt_row["date"]).strip()
        gt_time_in = gt_row.get("time_in", "")
        gt_time_out = gt_row.get("time_out", "")
        gt_hours = gt_row.get("total_hours", "")

        ws.cell(row=row, column=1, value=anon_name).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", wrap_text=True)
        ws.cell(row=row, column=2, value=gt_date).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=gt_time_in).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=gt_time_out).border = THIN_BORDER
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=gt_hours).border = THIN_BORDER
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")

        col = 6
        for aid, alabel in APPROACHES:
            matched = None
            for rid, rlabel, approach_results in all_approach_results:
                if rid != aid:
                    continue
                for r in approach_results:
                    if r["date"] == gt_date:
                        matched = r
                        break
                if matched:
                    break

            if matched and not matched["not_extracted"]:
                # Hours
                cell_h = ws.cell(row=row, column=col, value=matched["ext_hours"])
                cell_h.border = THIN_BORDER
                cell_h.alignment = Alignment(horizontal="center")
                col += 1

                # Correctness (hours only for simplicity in this column)
                correct = matched["hours_correct"]
                cell_c = ws.cell(row=row, column=col, value="✓" if correct else "✗")
                cell_c.border = THIN_BORDER
                cell_c.alignment = Alignment(horizontal="center")
                cell_c.fill = MATCH_FILL if correct else MISMATCH_FILL
                col += 1

                # Status
                status_val = matched["ext_status"]
                cell_s = ws.cell(row=row, column=col, value=status_val)
                cell_s.border = THIN_BORDER
                cell_s.alignment = Alignment(horizontal="center", wrap_text=True)
                if status_val.lower() == "accepted":
                    cell_s.fill = MATCH_FILL
                elif status_val.lower() == "flagged":
                    cell_s.fill = PARTIAL_FILL
                else:
                    cell_s.fill = MISMATCH_FILL
                col += 1
            else:
                # Not extracted
                for _ in range(3):
                    cell = ws.cell(row=row, column=col, value="—")
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal="center")
                    cell.fill = MISMATCH_FILL
                    col += 1
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    for i in range(6, len(detail_headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18


def main():
    print("=" * 60)
    print("Ground Truth Comparison — Extraction Accuracy")
    print("=" * 60)

    ground_truth = load_ground_truth()
    if ground_truth is None:
        return

    name_mapping = load_name_mapping()

    all_approach_results = []
    for approach_id, approach_label in APPROACHES:
        print(f"\nComparing {approach_label}...")
        results = compare_approach(ground_truth, approach_id, name_mapping)
        all_approach_results.append((approach_id, approach_label, results))

    metrics = compute_metrics(all_approach_results)

    # Print summary table
    print("\n" + "=" * 60)
    print("FIELD-LEVEL ACCURACY")
    print("=" * 60)
    print(f"{'Metric':<35}", end="")
    for _, m in metrics.items():
        print(f" {m['label'][:18]:<18}", end="")
    print()
    print("-" * 60)

    labels = [
        ("Rows Extracted", lambda m: str(m["extracted"])),
        ("Rows Not Extracted", lambda m: str(m["not_extracted"])),
        ("Date Accuracy", lambda m: f"{m['date_accuracy']:.1%}"),
        ("Hours Accuracy (±0.25h)", lambda m: f"{m['hours_accuracy']:.1%}"),
        ("Time In Accuracy (±30min)", lambda m: f"{m['time_in_accuracy']:.1%}"),
        ("Time Out Accuracy (±30min)", lambda m: f"{m['time_out_accuracy']:.1%}"),
        ("Fully Correct", lambda m: f"{m['fully_correct']} ({m['fully_correct_rate']:.1%})"),
        ("Partial or Full Match", lambda m: f"{m['fully_correct'] + m['partially_correct']} ({m['partial_or_full_rate']:.1%})"),
    ]
    for label, getter in labels:
        print(f"{label:<35}", end="")
        for _, m in metrics.items():
            print(f" {getter(m):<18}", end="")
        print()

    print("\n" + "=" * 60)
    print("PIPELINE VALIDATION QUALITY")
    print("=" * 60)
    print(f"{'Metric':<35}", end="")
    for _, m in metrics.items():
        print(f" {m['label'][:18]:<18}", end="")
    print()
    print("-" * 60)

    labels2 = [
        ("Rows Accepted", lambda m: str(m["accepted_count"])),
        ("Validation Precision", lambda m: f"{m['validation_precision']:.1%}"),
        ("Validation Recall", lambda m: f"{m['validation_recall']:.1%}"),
        ("Validation F1", lambda m: f"{m['validation_f1']:.3f}"),
        ("False Accept Rate", lambda m: f"{m['false_accept_rate']:.1%}"),
        ("Missed Detection Rate", lambda m: f"{m['missed_rate']:.1%}"),
    ]
    for label, getter in labels2:
        print(f"{label:<35}", end="")
        for _, m in metrics.items():
            print(f" {getter(m):<18}", end="")
        print()

    # Write to combined Excel
    if os.path.exists(COMBINED_PATH):
        wb = openpyxl.load_workbook(COMBINED_PATH)
    else:
        wb = openpyxl.Workbook()

    if "Human-Verified Results" in wb.sheetnames:
        del wb["Human-Verified Results"]

    write_human_verified_sheet(wb, metrics, all_approach_results, ground_truth)
    wb.save(COMBINED_PATH)
    print(f"\nResults saved to {COMBINED_PATH}")


if __name__ == "__main__":
    main()
