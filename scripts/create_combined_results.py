"""Create combined comparison outputs from all 5 approach results.

Generates a single benchmark_combined.xlsx with:
  - Approach Comparison: summary metrics + row-level comparison
  - Human-Verified Results: ground truth comparison (if ground_truth.xlsx exists)
"""

import glob
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

APPROACHES = [
    ("ocr_only", "OCR Only (Baseline)", "E8E8E8"),
    ("ppocr_grid", "OCR + VLM Fallback", "E2EFDA"),
    ("vlm_full_page", "VLM Full Page", "D6E4F0"),
    ("layout_guided_vlm_local", "Layout-Guided VLM (Local)", "FFF2CC"),
    ("layout_guided_vlm_cloud", "Layout-Guided VLM (Cloud)", "FCE4EC"),
]

OUTPUT_DIR = "output/combined"
BENCH_OUTPUT = f"{OUTPUT_DIR}/benchmark_combined.xlsx"

GROUND_TRUTH_PATH = "ground_truth.xlsx"
HOURS_TOLERANCE = 0.25  # ±15 minutes
TIME_TOLERANCE_MIN = 30  # ±30 minutes

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MISMATCH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

FILLS = {}
for _, _, color in APPROACHES:
    FILLS[color] = PatternFill(start_color=color, end_color=color, fill_type="solid")


def _short_label(label, max_len=16):
    """Shorten a label while keeping the last distinguishing word visible."""
    if len(label) <= max_len:
        return label
    suffix = label.split("(")[-1].rstrip(")") if "(" in label else label.split()[-1]
    prefix_max = max_len - len(suffix) - 3
    prefix = (
        label[:prefix_max].rsplit(" ", 1)[0]
        if " " in label[:prefix_max]
        else label[:prefix_max]
    )
    return f"{prefix} ({suffix})"


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, fill=None):
    cell = ws.cell(row=row, column=col)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    if fill:
        cell.fill = fill


def load_benchmark(folder):
    """Load and aggregate all benchmark Excel files from approach output directory."""
    bench_files = sorted(
        f for f in glob.glob(f"output/{folder}/benchmark_*.xlsx") if "combined" not in f
    )
    if not bench_files:
        return None

    # Also load merged results to fill in missing hours
    merged_path = f"output/{folder}/merged_results.xlsx"
    merged_by_date = {}
    if os.path.exists(merged_path):
        mwb = openpyxl.load_workbook(merged_path)
        mws = mwb.active
        mheader = None
        for mrow in mws.iter_rows(min_row=1, values_only=True):
            if mheader is None:
                mheader = [str(c).strip() if c else "" for c in mrow]
                continue
            if not any(mrow):
                continue
            mrec = dict(zip(mheader, mrow))
            date_val = str(mrec.get("Date", "")).strip()
            if date_val:
                merged_by_date[date_val] = mrec
        mwb.close()

    # Aggregate across all patient files
    agg_summary = {}
    all_pages = []
    all_rows = []
    total_time = 0.0

    for path in bench_files:
        wb = openpyxl.load_workbook(path)
        summary = {}
        for r in wb["Run Summary"].iter_rows(min_row=1, values_only=False):
            key = r[0].value
            val = r[1].value
            if key:
                summary[key] = val

        pages = list(wb["Page Details"].iter_rows(min_row=2, values_only=True))
        rows = list(wb["Row-Level"].iter_rows(min_row=2, values_only=True))

        # Fill missing hours from merged results
        if merged_by_date:
            new_rows = []
            for r in rows:
                date_val = str(r[6]) if len(r) > 6 and r[6] else ""
                hours_val = r[18] if len(r) > 18 else None
                if (
                    date_val
                    and (hours_val is None or hours_val == "")
                    and date_val in merged_by_date
                ):
                    mr = merged_by_date[date_val]
                    hours = mr.get("Total Hours") or mr.get("Calculated Hours", "")
                    # Create a new row tuple with hours filled in
                    r_list = list(r)
                    while len(r_list) <= 18:
                        r_list.append(None)
                    r_list[18] = hours
                    new_rows.append(tuple(r_list))
                else:
                    new_rows.append(r)
            rows = new_rows

        total_time += summary.get("Total Processing Time (s)", 0) or 0
        all_pages.extend(pages)
        all_rows.extend(rows)

    # Build aggregated summary
    agg_summary["Total Processing Time (s)"] = round(total_time, 2)
    agg_summary["Number of Pages"] = len(all_pages)
    agg_summary["Total Rows Extracted"] = len(all_rows)

    # Sum integer metrics
    for key in [
        "Accepted Rows",
        "Flagged Rows",
        "Failed Rows",
        "VLM Fallbacks Triggered",
    ]:
        agg_summary[key] = sum(
            (s.get(key, 0) or 0)
            for path in bench_files
            for s in [_load_single_summary(path)]
        )

    # Average percentage metrics
    for key in ["Hours Mismatch Rate", "Field Missing Rate"]:
        vals = []
        for path in bench_files:
            s = _load_single_summary(path)
            v = s.get(key)
            if isinstance(v, str) and "%" in v:
                vals.append(float(v.replace("%", "")))
        if vals:
            agg_summary[key] = f"{sum(vals) / len(vals):.1f}%"

    # Average float metrics
    for key in [
        "Mean Overall Confidence",
        "Min Overall Confidence",
        "Mean Character Error Rate",
    ]:
        vals = [
            s.get(key)
            for path in bench_files
            for s in [_load_single_summary(path)]
            if isinstance(s.get(key), (int, float))
        ]
        if vals:
            agg_summary[key] = round(sum(vals) / len(vals), 4)

    return {"summary": agg_summary, "pages": all_pages, "rows": all_rows}


def _load_single_summary(path):
    """Load summary dict from a single benchmark file."""
    wb = openpyxl.load_workbook(path)
    summary = {}
    for r in wb["Run Summary"].iter_rows(min_row=1, values_only=False):
        key = r[0].value
        val = r[1].value
        if key:
            summary[key] = val
    return summary


def load_merged(folder):
    path = f"output/{folder}/merged_results.xlsx"
    if not os.path.exists(path):
        return None
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    return list(ws.iter_rows(min_row=2, values_only=True))


def _detect_input_file():
    """Detect the input PDF name from benchmark files."""
    for folder, _, _ in APPROACHES:
        bench_files = glob.glob(f"output/{folder}/benchmark_*.xlsx")
        if bench_files:
            fname = os.path.basename(bench_files[0])
            # Extract: benchmark_patient_a_week1.xlsx → patient_a_week1
            stem = fname.replace("benchmark_", "").replace(".xlsx", "")
            return stem.replace("_", " ").title()
    return "Unknown"


def create_benchmark_combined():
    """Create benchmark comparison grouped by approach with IEEE-paper-ready format."""
    data = {}
    for folder, label, color in APPROACHES:
        d = load_benchmark(folder)
        if d:
            d["_color"] = color
            data[folder] = {"label": label, **d}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Approach Comparison"

    input_name = _detect_input_file()
    row = 1
    ws.cell(
        row=row, column=1, value="Approach Comparison: Handwritten Timesheet OCR"
    ).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=f"File: {input_name} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    ).font = Font(italic=True, size=10)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 2

    # SUMMARY METRICS
    ws.cell(row=row, column=1, value="SUMMARY METRICS").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1

    headers = ["Metric"] + [d["label"] for d in data.values()]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    metrics = [
        ("Total Processing Time (s)", "Total Processing Time (s)", True),
        ("Pages Processed", "Number of Pages", False),
        ("Rows Extracted", "Total Rows Extracted", False),
        ("Accepted Rows", "Accepted Rows", False, True),
        ("Flagged Rows", "Flagged Rows", False, False),
        ("Failed Rows", "Failed Rows", False, False),
        ("Mean Confidence", "Mean Overall Confidence", False, True),
        ("Min Confidence", "Min Overall Confidence", False, True),
        ("VLM Fallbacks Triggered", "VLM Fallbacks Triggered", False, False),
        ("Hours Mismatch Rate", "Hours Mismatch Rate", False, False),
        ("Field Missing Rate", "Field Missing Rate", False, False),
        ("Mean CER", "Mean Character Error Rate", False, False),
    ]

    for metric_def in metrics:
        label = metric_def[0]
        key = metric_def[1]

        ws.cell(row=row, column=1, value=label)
        for i, (folder, d) in enumerate(data.items()):
            val = d["summary"].get(key, "N/A")
            ws.cell(row=row, column=i + 2, value=val)
            style_data_cell(ws, row, i + 2, FILLS.get(d.get("_color", ""), None))
        row += 1

    row += 1

    # ROW-LEVEL COMPARISON (dates as rows, Hours+Status per approach as columns)
    ws.cell(row=row, column=1, value="ROW-LEVEL COMPARISON (by date)").font = Font(
        bold=True, size=12
    )
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1

    # Build per-approach row lookups by date only
    approach_by_date = {}
    for folder, d in data.items():
        by_date = {}
        for r in d["rows"]:
            date_val = str(r[6]) if len(r) > 6 and r[6] else ""
            if date_val:
                by_date[date_val] = r
        approach_by_date[folder] = by_date

    all_dates = set()
    for by_date in approach_by_date.values():
        all_dates.update(by_date.keys())
    all_dates = sorted(all_dates)

    # Columns: Date | Hours+Status per approach
    comp_headers = ["Date"]
    for folder, d in data.items():
        short = _short_label(d["label"])
        comp_headers.append(f"Hours ({short})")
        comp_headers.append(f"Status ({short})")

    for c, h in enumerate(comp_headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(comp_headers))
    row += 1

    for date_val in all_dates:
        ws.cell(row=row, column=1, value=date_val)
        style_data_cell(ws, row, 1)

        col = 2
        for folder, d in data.items():
            r = approach_by_date[folder].get(date_val)
            fill_color = FILLS.get(d.get("_color", ""), None)

            if r:
                hours = str(r[18]) if len(r) > 18 and r[18] is not None else ""
                status = str(r[24]) if len(r) > 24 and r[24] else ""
            else:
                hours = ""
                status = "not extracted"

            ws.cell(row=row, column=col, value=hours)
            style_data_cell(ws, row, col, fill_color)
            col += 1

            ws.cell(row=row, column=col, value=status)
            fill = fill_color
            if status == "accepted":
                fill = MATCH_FILL
            elif status == "not extracted":
                fill = MISMATCH_FILL
            style_data_cell(ws, row, col, fill)
            col += 1

        row += 1

    ws.column_dimensions["A"].width = 14
    for i in range(2, len(comp_headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 22

    wb.save(BENCH_OUTPUT)
    print(f"Saved: {BENCH_OUTPUT}")


# ═══════════════════════════════════════════════════════
# Ground Truth Comparison
# ═══════════════════════════════════════════════════════

def _parse_time(val):
    """Parse time value to minutes since midnight."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.hour * 60 + val.minute
    if isinstance(val, (int, float)):
        v = int(val)
        hour = v // 100
        minute = v % 100
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return hour * 60 + minute
        return None
    val = str(val).strip()
    if not val:
        return None
    match = re.match(r"(\d{1,2}):?(\d{2})?\s*(AM|PM|am|pm|Am|Pm)?", val)
    if match:
        hour = int(match.group(1))
        minute = int(match.group(2) or 0)
        ampm = match.group(3)
        if ampm and ampm.lower() == "pm" and hour != 12:
            hour += 12
        elif ampm and ampm.lower() == "am" and hour == 12:
            hour = 0
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return hour * 60 + minute
    try:
        v = int(val.replace(":", ""))
        hour = v // 100
        minute = v % 100
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return hour * 60 + minute
    except (ValueError, TypeError):
        pass
    return None


def _parse_date(val):
    """Parse date value to YYYY-MM-DD string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    val = str(val).strip()
    if not val:
        return None
    match = re.match(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", val)
    if match:
        month = int(match.group(1))
        day = int(match.group(2))
        year = int(match.group(3))
        if year < 100:
            year += 2000
        return f"{year:04d}-{month:02d}-{day:02d}"
    match = re.match(r"(\d{4})-(\d{2})-(\d{2})", val)
    if match:
        return val
    return val


def _compute_hours(time_in_min, time_out_min):
    """Compute hours between two time values. Handles overnight shifts."""
    if time_in_min is None or time_out_min is None:
        return None
    diff = time_out_min - time_in_min
    if diff < 0:
        diff += 24 * 60
    return diff / 60.0


def _load_approach_data_for_gt(approach_id):
    """Load all benchmark data for an approach for ground truth comparison."""
    bench_files = sorted(
        f for f in glob.glob(f"output/{approach_id}/benchmark_*.xlsx")
        if "combined" not in f
    )
    all_rows = []
    for path in bench_files:
        wb = openpyxl.load_workbook(path, read_only=True)
        if "Row-Level" in wb.sheetnames:
            ws = wb["Row-Level"]
            header = None
            for row in ws.iter_rows(values_only=True):
                if header is None:
                    header = [str(c).strip() if c else "" for c in row]
                    continue
                if not any(row):
                    continue
                record = dict(zip(header, row))
                record["Date"] = record.get("Parsed Date", record.get("Date", ""))
                record["Hours"] = record.get("Parsed Hours", record.get("Written Hours"))
                record["Status"] = record.get("Status", "")
                record["Time In"] = record.get("Parsed Time In", "")
                record["Time Out"] = record.get("Parsed Time Out", "")
                record["Employee"] = record.get("Employee Name", "")
                all_rows.append(record)
        wb.close()

    # Also load merged results to fill in missing data
    merged_path = f"output/{approach_id}/merged_results.xlsx"
    if os.path.exists(merged_path):
        wb = openpyxl.load_workbook(merged_path, read_only=True)
        ws = wb.active
        header = None
        merged_by_date = {}
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c).strip() if c else "" for c in row]
                continue
            if not any(row):
                continue
            record = dict(zip(header, row))
            date = _parse_date(record.get("Date", ""))
            if date:
                merged_by_date[date] = record
        wb.close()
        for record in all_rows:
            date = _parse_date(record.get("Date", ""))
            if date and date in merged_by_date:
                mr = merged_by_date[date]
                if record["Hours"] is None:
                    record["Hours"] = mr.get("Total Hours")
                if record["Status"] in ("", None):
                    record["Status"] = mr.get("Status", "")
                if not record["Time In"]:
                    record["Time In"] = mr.get("Time In", "")
                if not record["Time Out"]:
                    record["Time Out"] = mr.get("Time Out", "")

    return all_rows


def _compare_approach_to_gt(ground_truth, approach_id):
    """Compare approach output against ground truth.

    For each GT row, finds the BEST matching approach row (closest hours).
    Additional approach rows with the same GT date are classified as "duplicates"
    (not hallucinations) — these happen when the source timesheet has duplicate entries.
    Approach rows with dates NOT in ground truth are "extra" (hallucinations).

    Returns a dict with:
      - matched: best-matching approach row for each GT row
      - unmatched_gt: GT rows that were NOT found
      - duplicates: approach rows with GT date but not the best match
      - extra: approach rows with dates NOT in ground truth
    """
    approach_rows = _load_approach_data_for_gt(approach_id)

    # Normalize approach row dates
    for ar in approach_rows:
        ar_date_raw = ar.get("Date", "")
        ar["__date"] = _parse_date(ar_date_raw) if ar_date_raw else (str(ar_date_raw).strip() if ar_date_raw else "")

    # Build set of ground truth dates
    gt_dates_set = set()
    for gt_row in ground_truth:
        gt_date = _parse_date(gt_row["date"])
        if gt_date is None:
            gt_date = str(gt_row["date"]).strip()
        gt_dates_set.add(gt_date)

    # Track which approach rows have been claimed
    used_indices = set()
    best_match_indices = set()

    matched = []
    for gt_row in ground_truth:
        source_file = gt_row["source_file"]
        gt_date = _parse_date(gt_row["date"])
        if gt_date is None:
            gt_date = str(gt_row["date"]).strip()
        gt_hours = gt_row.get("total_hours")
        gt_time_in = gt_row.get("time_in")
        gt_time_out = gt_row.get("time_out")
        gt_employee = gt_row.get("employee_name", "")

        # Find GT hours value for picking best match
        gt_time_in_min = _parse_time(gt_time_in)
        gt_time_out_min = _parse_time(gt_time_out)
        gt_hours_calc = _compute_hours(gt_time_in_min, gt_time_out_min)
        gt_hours_val = gt_hours_calc if gt_hours_calc is not None else gt_hours

        # Find ALL approach rows with matching date
        candidates = []
        for i, ar in enumerate(approach_rows):
            if i in used_indices:
                continue
            if ar["__date"] == gt_date:
                ext_time_in = ar.get("Time In", "")
                ext_time_out = ar.get("Time Out", "")
                ext_hours = ar.get("Hours")

                # Parse times
                ext_time_in_min = _parse_time(ext_time_in)
                ext_time_out_min = _parse_time(ext_time_out)

                # Score: PRIMARY = total hours, SECONDARY = time-in/time-out
                # Primary: hours distance
                h_dist = 999
                if ext_hours is not None and gt_hours_val is not None:
                    try:
                        h_dist = abs(float(ext_hours) - float(gt_hours_val))
                    except (ValueError, TypeError):
                        h_dist = 999
                elif ext_hours is not None:
                    h_dist = 0  # has hours but GT doesn't

                # Secondary: time-in distance (minutes)
                ti_dist = 9999
                if ext_time_in_min is not None and gt_time_in_min is not None:
                    ti_dist = abs(ext_time_in_min - gt_time_in_min)

                # Tertiary: time-out distance (minutes)
                to_dist = 9999
                if ext_time_out_min is not None and gt_time_out_min is not None:
                    to_dist = abs(ext_time_out_min - gt_time_out_min)

                # Composite score: hours first, then time-in, then time-out
                score = (h_dist * 1000) + ti_dist + (to_dist * 0.01)
                candidates.append((score, h_dist, ti_dist, to_dist, i, ar))

        if candidates:
            # Pick the candidate with best hours match first, then time
            candidates.sort(key=lambda x: x[0])
            _, _, _, _, best_idx, matched_row = candidates[0]
            used_indices.add(best_idx)
            best_match_indices.add(best_idx)

            # Mark remaining candidates as duplicates
            for _, _, _, _, dup_idx, dup_row in candidates[1:]:
                used_indices.add(dup_idx)

            ext_hours = matched_row.get("Hours")
            ext_status = matched_row.get("Status", "")
            ext_time_in = matched_row.get("Time In", "")
            ext_time_out = matched_row.get("Time Out", "")
            ext_employee = matched_row.get("Employee", "")

            ext_time_in_min = _parse_time(ext_time_in)
            ext_time_out_min = _parse_time(ext_time_out)

            date_ok = (gt_date is not None and ext_hours is not None)

            hours_ok = False
            if ext_hours is not None and gt_hours_val is not None:
                try:
                    hours_ok = abs(float(ext_hours) - float(gt_hours_val)) <= HOURS_TOLERANCE
                except (ValueError, TypeError):
                    pass

            time_in_ok = False
            if ext_time_in_min is not None and gt_time_in_min is not None:
                time_in_ok = abs(ext_time_in_min - gt_time_in_min) <= TIME_TOLERANCE_MIN

            time_out_ok = False
            if ext_time_out_min is not None and gt_time_out_min is not None:
                time_out_ok = abs(ext_time_out_min - gt_time_out_min) <= TIME_TOLERANCE_MIN

            fields_correct = sum([hours_ok, time_in_ok, time_out_ok])
            fully_correct = (fields_correct == 3)
            partially_correct = (0 < fields_correct < 3)
            not_extracted = False
        else:
            ext_hours = None
            ext_status = "not extracted"
            ext_time_in = None
            ext_time_out = None
            ext_employee = None
            date_ok = False
            hours_ok = False
            time_in_ok = False
            time_out_ok = False
            fields_correct = 0
            fully_correct = False
            partially_correct = False
            not_extracted = True

        matched.append({
            "source_file": source_file,
            "date": gt_date,
            "gt_hours": gt_hours,
            "gt_time_in": gt_time_in,
            "gt_time_out": gt_time_out,
            "gt_employee": gt_employee,
            "ext_hours": ext_hours,
            "ext_time_in": ext_time_in,
            "ext_time_out": ext_time_out,
            "ext_status": ext_status,
            "ext_employee": ext_employee,
            "date_correct": date_ok,
            "hours_correct": hours_ok,
            "time_in_correct": time_in_ok,
            "time_out_correct": time_out_ok,
            "fields_correct": fields_correct,
            "fully_correct": fully_correct,
            "partially_correct": partially_correct,
            "not_extracted": not_extracted,
        })

    # Classify remaining approach rows as duplicates or extra
    duplicates = []
    extra = []
    for i, ar in enumerate(approach_rows):
        if i in best_match_indices:
            # Best match for a GT date — already counted in matched list
            continue
        elif i in used_indices:
            # Was a candidate but not the best match — duplicate
            duplicates.append({
                "date": ar["__date"],
                "hours": ar.get("Hours"),
                "status": ar.get("Status", ""),
                "employee": ar.get("Employee", ""),
            })
        elif ar["__date"] and ar["__date"] in gt_dates_set:
            # Has GT date but wasn't matched — duplicate
            duplicates.append({
                "date": ar["__date"],
                "hours": ar.get("Hours"),
                "status": ar.get("Status", ""),
                "employee": ar.get("Employee", ""),
            })
        elif ar["__date"]:
            # Date NOT in ground truth — true hallucination
            extra.append({
                "date": ar["__date"],
                "hours": ar.get("Hours"),
                "status": ar.get("Status", ""),
                "employee": ar.get("Employee", ""),
            })

    unmatched_gt = [r for r in matched if r["not_extracted"]]

    return {
        "matched": matched,
        "unmatched_gt": unmatched_gt,
        "duplicates": duplicates,
        "extra": extra,
        "total_approach_rows": len(approach_rows),
    }


def _compute_gt_metrics(all_approach_results):
    """Compute field-level accuracy and pipeline validation quality metrics.

    all_approach_results: list of (approach_id, label, comparison_dict)
    """
    metrics = {}
    for approach_id, approach_label, comp in all_approach_results:
        matched = comp["matched"]
        duplicates = comp["duplicates"]
        extra = comp["extra"]
        total_approach_rows = comp["total_approach_rows"]
        total_gt = len(matched)
        if total_gt == 0:
            continue

        date_correct = sum(1 for r in matched if r["date_correct"])
        hours_correct = sum(1 for r in matched if r["hours_correct"])
        time_in_correct = sum(1 for r in matched if r["time_in_correct"])
        time_out_correct = sum(1 for r in matched if r["time_out_correct"])
        fully_correct = sum(1 for r in matched if r["fully_correct"])
        partially_correct = sum(1 for r in matched if r["partially_correct"])
        not_extracted = sum(1 for r in matched if r["not_extracted"])

        accepted = [r for r in matched if r["ext_status"].lower() == "accepted"]
        not_accepted = [r for r in matched if r["ext_status"].lower() != "accepted"]

        accepted_correct = sum(1 for r in accepted if r["fully_correct"])
        validation_precision = accepted_correct / len(accepted) if accepted else 0.0

        correct_rows = [r for r in matched if r["fully_correct"]]
        correct_accepted = sum(1 for r in correct_rows if r["ext_status"].lower() == "accepted")
        validation_recall = correct_accepted / len(correct_rows) if correct_rows else 0.0

        if validation_precision + validation_recall > 0:
            validation_f1 = 2 * validation_precision * validation_recall / (validation_precision + validation_recall)
        else:
            validation_f1 = 0.0

        false_accept = sum(1 for r in accepted if not r["fully_correct"])
        false_accept_rate = false_accept / len(accepted) if accepted else 0.0

        missed = sum(1 for r in not_accepted if r["fully_correct"])
        missed_rate = missed / len(correct_rows) if correct_rows else 0.0

        metrics[approach_id] = {
            "label": approach_label,
            "total_gt_rows": total_gt,
            "gt_rows_matched": total_gt - not_extracted,
            "gt_rows_missed": not_extracted,
            "duplicate_rows": len(duplicates),
            "duplicate_details": duplicates,
            "extra_rows": len(extra),
            "extra_details": extra,
            "total_approach_rows": total_approach_rows,
            "date_accuracy": date_correct / total_gt if total_gt else 0.0,
            "hours_accuracy": hours_correct / total_gt if total_gt else 0.0,
            "time_in_accuracy": time_in_correct / total_gt if total_gt else 0.0,
            "time_out_accuracy": time_out_correct / total_gt if total_gt else 0.0,
            "fully_correct": fully_correct,
            "partially_correct": partially_correct,
            "fully_correct_rate": fully_correct / total_gt if total_gt else 0.0,
            "partial_or_full_rate": (fully_correct + partially_correct) / total_gt if total_gt else 0.0,
            "accepted_count": len(accepted),
            "validation_precision": validation_precision,
            "validation_recall": validation_recall,
            "validation_f1": validation_f1,
            "false_accept_rate": false_accept_rate,
            "missed_rate": missed_rate,
        }
    return metrics


def add_ground_truth_comparison():
    """Add Human-Verified Results sheet to benchmark_combined.xlsx if ground truth exists."""
    if not os.path.exists(GROUND_TRUTH_PATH):
        print(f"Ground truth file not found: {GROUND_TRUTH_PATH} — skipping ground truth comparison")
        return

    print("Adding ground truth comparison...")

    # Load ground truth
    wb_gt = openpyxl.load_workbook(GROUND_TRUTH_PATH, read_only=True)
    ws_gt = wb_gt.active
    ground_truth = []
    header = None
    for row in ws_gt.iter_rows(values_only=True):
        if header is None:
            header = [str(c).strip() if c else "" for c in row]
            continue
        if not any(row):
            continue
        record = dict(zip(header, row))
        if record.get("source_file") and record.get("date"):
            ground_truth.append(record)
    wb_gt.close()
    print(f"  Loaded {len(ground_truth)} ground truth rows")

    if not ground_truth:
        print("  No ground truth rows found — skipping")
        return

    # Compare each approach
    all_results = []
    for approach_id, approach_label, _ in APPROACHES:
        results = _compare_approach_to_gt(ground_truth, approach_id)
        all_results.append((approach_id, approach_label, results))
        print(f"  {approach_id}: {len(results)} comparison rows")

    metrics = _compute_gt_metrics(all_results)

    # Open the existing benchmark_combined.xlsx and add the new sheet
    wb = openpyxl.load_workbook(BENCH_OUTPUT)
    ws = wb.create_sheet("Human-Verified Results")

    # Write the Human-Verified Results sheet
    row = 1
    ws.cell(row=row, column=1, value="Ground Truth Comparison — Extraction Accuracy").font = Font(
        bold=True, size=14
    )
    num_cols = len(metrics) + 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    row += 1
    ws.cell(
        row=row, column=1,
        value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
              f"Tolerance: Hours ±{HOURS_TOLERANCE}h, Time ±{TIME_TOLERANCE_MIN}min",
    ).font = Font(italic=True, size=10)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    row += 2

    # Section 1: Extraction Coverage & Field-Level Accuracy
    ws.cell(row=row, column=1, value="EXTRACTION COVERAGE & FIELD-LEVEL ACCURACY").font = Font(
        bold=True, size=12
    )
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    row += 1

    metric_headers = ["Metric"] + [m["label"] for m in metrics.values()]
    for c, h in enumerate(metric_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    section1_rows = [
        ("Total Ground Truth Rows", lambda m: m["total_gt_rows"]),
        ("GT Rows Matched (best fit)", lambda m: m["gt_rows_matched"]),
        ("GT Rows Missed (not found)", lambda m: m["gt_rows_missed"]),
        ("Duplicate Rows (same date, not best match)", lambda m: m["duplicate_rows"]),
        ("Extra Rows (hallucinated)", lambda m: m["extra_rows"]),
        ("Total Rows Approach Extracted", lambda m: m["total_approach_rows"]),
        ("", lambda m: ""),
        ("Date Accuracy (exact match)", lambda m: f"{m['date_accuracy']:.1%}"),
        ("Hours Accuracy (±0.25h)", lambda m: f"{m['hours_accuracy']:.1%}"),
        ("Time In Accuracy (±30min)", lambda m: f"{m['time_in_accuracy']:.1%}"),
        ("Time Out Accuracy (±30min)", lambda m: f"{m['time_out_accuracy']:.1%}"),
        ("", lambda m: ""),
        ("Fully Correct (all 3 fields)", lambda m: f"{m['fully_correct']} ({m['fully_correct_rate']:.1%})"),
        ("Partial or Full Match", lambda m: f"{m['fully_correct'] + m['partially_correct']} ({m['partial_or_full_rate']:.1%})"),
    ]

    for label, getter in section1_rows:
        if label == "":
            for c in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.border = THIN_BORDER
            row += 1
            continue
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", wrap_text=True)
        for c, (aid, m) in enumerate(metrics.items(), 2):
            val = getter(m)
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        row += 1

    row += 1

    # Section 2: Pipeline Validation Quality
    ws.cell(row=row, column=1, value="PIPELINE VALIDATION QUALITY").font = Font(
        bold=True, size=12
    )
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    row += 1

    metric_headers2 = ["Metric"] + [m["label"] for m in metrics.values()]
    for c, h in enumerate(metric_headers2, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    section2_rows = [
        ("Rows Marked 'Accepted'", lambda m: m["accepted_count"]),
        ("", lambda m: ""),
        ("Validation Precision", lambda m: f"{m['validation_precision']:.1%}"),
        ("  → Of accepted rows, % fully correct", lambda m: ""),
        ("Validation Recall", lambda m: f"{m['validation_recall']:.1%}"),
        ("  → Of correct rows, % accepted", lambda m: ""),
        ("Validation F1", lambda m: f"{m['validation_f1']:.3f}"),
        ("", lambda m: ""),
        ("False Accept Rate", lambda m: f"{m['false_accept_rate']:.1%}"),
        ("  → Accepted but actually wrong", lambda m: ""),
        ("Missed Detection Rate", lambda m: f"{m['missed_rate']:.1%}"),
        ("  → Correct but flagged/failed", lambda m: ""),
    ]

    for label, getter in section2_rows:
        if label.startswith("  →"):
            cell = ws.cell(row=row, column=1, value=label)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left")
            cell.font = Font(italic=True, size=10)
            for c in range(2, num_cols + 1):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1
            continue
        if label == "":
            for c in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.border = THIN_BORDER
            row += 1
            continue
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", wrap_text=True)
        for c, (aid, m) in enumerate(metrics.items(), 2):
            val = getter(m)
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        row += 1

    row += 1

    # Section 3: Per-Row Detailed Comparison
    ws.cell(row=row, column=1, value="PER-ROW DETAILED COMPARISON").font = Font(
        bold=True, size=12
    )
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    row += 1

    detail_headers = [
        "Source File", "Date",
        "GT Time In", "GT Time Out", "GT Hours",
    ]
    for _aid, alabel, _ in APPROACHES:
        detail_headers.append(f"{alabel}: Hours")
        detail_headers.append(f"{alabel}: ✓")
        detail_headers.append(f"{alabel}: Status")

    for c, h in enumerate(detail_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    anon_map = {
        "patient_a_week1": "File 1 (Week 1)",
        "patient_b_week2": "File 2 (Week 2)",
        "patient_c_week3": "File 3 (Week 3)",
    }

    for gt_row in ground_truth:
        source_file = gt_row["source_file"]
        anon_name = None
        for key, val in anon_map.items():
            if key in source_file.lower() or source_file.lower() in key:
                anon_name = val
                break
        if anon_name is None:
            anon_name = source_file

        gt_date = _parse_date(gt_row["date"])
        if gt_date is None:
            gt_date = str(gt_row["date"]).strip()
        gt_time_in = gt_row.get("time_in", "")
        gt_time_out = gt_row.get("time_out", "")
        gt_hours = gt_row.get("total_hours", "")

        ws.cell(row=row, column=1, value=anon_name).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", wrap_text=True)
        ws.cell(row=row, column=2, value=gt_date).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=gt_time_in).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=gt_time_out).border = THIN_BORDER
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=gt_hours).border = THIN_BORDER
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")

        col = 6
        for aid, alabel, _ in APPROACHES:
            matched_row = None
            for rid, rlabel, comp_dict in all_results:
                if rid == aid:
                    for r in comp_dict["matched"]:
                        if r["date"] == gt_date and r["source_file"] == source_file:
                            matched_row = r
                            break
                    break

            if matched_row:
                hours_val = matched_row["ext_hours"] if matched_row["ext_hours"] is not None else ""
                status_val = matched_row["ext_status"]
                correct = matched_row["fully_correct"]
                partial = matched_row["partially_correct"]
            else:
                hours_val = ""
                status_val = "not extracted"
                correct = False
                partial = False

            ws.cell(row=row, column=col, value=hours_val).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
            col += 1

            if correct:
                check = "✓"
                fill = MATCH_FILL
            elif partial:
                check = "△"
                fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                check = "✗"
                fill = MISMATCH_FILL if status_val == "not extracted" else None
            cell = ws.cell(row=row, column=col, value=check)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if fill:
                cell.fill = fill
            col += 1

            cell = ws.cell(row=row, column=col, value=status_val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if status_val == "accepted":
                cell.fill = MATCH_FILL
            elif status_val == "not extracted":
                cell.fill = MISMATCH_FILL
            col += 1

        row += 1

    # Section 4: Duplicate Rows (same date as GT but not the best match)
    row += 1
    has_dupes = any(comp["duplicates"] for _, _, comp in all_results)
    if has_dupes:
        ws.cell(row=row, column=1, value="DUPLICATE ROWS (Same Date as GT — Source Timesheet Had Duplicate Entries)").font = Font(
            bold=True, size=12
        )
        ws.cell(row=row, column=1).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1

        dupe_headers = ["Approach", "Date", "Hours", "Status", "Employee"]
        for c, h in enumerate(dupe_headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = THIN_BORDER
        row += 1

        for aid, alabel, comp_dict in all_results:
            dupes = comp_dict["duplicates"]
            if not dupes:
                continue
            for du in dupes:
                ws.cell(row=row, column=1, value=alabel).border = THIN_BORDER
                ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", wrap_text=True)
                ws.cell(row=row, column=2, value=du["date"]).border = THIN_BORDER
                ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=3, value=du["hours"] if du["hours"] else "").border = THIN_BORDER
                ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=4, value=du["status"]).border = THIN_BORDER
                ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=4).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                ws.cell(row=row, column=5, value=du["employee"] if du["employee"] else "").border = THIN_BORDER
                ws.cell(row=row, column=5).alignment = Alignment(horizontal="left")
                row += 1

    # Section 5: Extra/Hallucinated Rows (approach extracted but not in ground truth)
    row += 1
    has_extra = any(comp["extra"] for _, _, comp in all_results)
    if has_extra:
        ws.cell(row=row, column=1, value="EXTRA ROWS (Extracted by Approach — Not in Ground Truth)").font = Font(
            bold=True, size=12
        )
        ws.cell(row=row, column=1).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1

        extra_headers = ["Approach", "Date", "Hours", "Status", "Employee"]
        for c, h in enumerate(extra_headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = THIN_BORDER
        row += 1

        for aid, alabel, comp_dict in all_results:
            extras = comp_dict["extra"]
            if not extras:
                continue
            for ex in extras:
                ws.cell(row=row, column=1, value=alabel).border = THIN_BORDER
                ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", wrap_text=True)
                ws.cell(row=row, column=2, value=ex["date"]).border = THIN_BORDER
                ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=3, value=ex["hours"] if ex["hours"] else "").border = THIN_BORDER
                ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=4, value=ex["status"]).border = THIN_BORDER
                ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=4).fill = MISMATCH_FILL
                ws.cell(row=row, column=5, value=ex["employee"] if ex["employee"] else "").border = THIN_BORDER
                ws.cell(row=row, column=5).alignment = Alignment(horizontal="left")
                row += 1

    # Set column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    for i in range(6, col + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18

    wb.save(BENCH_OUTPUT)
    print(f"  Added 'Human-Verified Results' sheet to {BENCH_OUTPUT}")

    # Add the Time Comparison sheet
    add_time_comparison_sheet(wb, ground_truth, all_results)


def add_time_comparison_sheet(wb, ground_truth, all_results):
    """Add a 'Time Comparison' sheet with Time-In, Time-Out, and correctness tables."""
    ws = wb.create_sheet("Time Comparison")

    HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    MISMATCH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    PARTIAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    row = 1
    ws.cell(row=row, column=1, value="Time-In / Time-Out Comparison vs Ground Truth").font = Font(bold=True, size=14)
    num_cols = len(APPROACHES) + 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    row += 1
    ws.cell(
        row=row, column=1,
        value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
              f"Tolerance: Time ±{TIME_TOLERANCE_MIN}min",
    ).font = Font(italic=True, size=10)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    row += 2

    # Build a lookup: for each approach, map (source_file, date) → matched row data
    approach_lookup = {}
    for approach_id, approach_label, comp in all_results:
        approach_lookup[approach_id] = {}
        for m in comp["matched"]:
            key = (m["source_file"], m["date"])
            approach_lookup[approach_id][key] = m

    # ═══════════════════════════════════════════
    # Table 1: Time-In Comparison
    # ═══════════════════════════════════════════
    ws.cell(row=row, column=1, value="TIME-IN COMPARISON").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    row += 1

    ti_headers = ["Source File", "Date", "Ground Truth"]
    for _aid, alabel, _ in APPROACHES:
        ti_headers.append(alabel)
    for c, h in enumerate(ti_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    anon_map = {
        "patient_a_week1": "File 1 (Week 1)",
        "patient_b_week2": "File 2 (Week 2)",
        "patient_c_week3": "File 3 (Week 3)",
    }

    for gt_row in ground_truth:
        source_file = gt_row["source_file"]
        anon_name = None
        for key, val in anon_map.items():
            if key in source_file.lower() or source_file.lower() in key:
                anon_name = val
                break
        if anon_name is None:
            anon_name = source_file

        gt_date = _parse_date(gt_row["date"])
        if gt_date is None:
            gt_date = str(gt_row["date"]).strip()
        gt_time_in = gt_row.get("time_in", "")

        ws.cell(row=row, column=1, value=anon_name).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", wrap_text=True)
        ws.cell(row=row, column=2, value=gt_date).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=gt_time_in).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

        col = 4
        for approach_id, approach_label, _ in APPROACHES:
            m = approach_lookup[approach_id].get((source_file, gt_date))
            if m and m["ext_time_in"]:
                ext_val = m["ext_time_in"]
                fill = MATCH_FILL if m["time_in_correct"] else MISMATCH_FILL
            else:
                ext_val = "—"
                fill = MISMATCH_FILL
            cell = ws.cell(row=row, column=col, value=ext_val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            cell.fill = fill
            col += 1

        row += 1

    row += 1

    # ═══════════════════════════════════════════
    # Table 2: Time-Out Comparison
    # ═══════════════════════════════════════════
    ws.cell(row=row, column=1, value="TIME-OUT COMPARISON").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    row += 1

    to_headers = ["Source File", "Date", "Ground Truth"]
    for _aid, alabel, _ in APPROACHES:
        to_headers.append(alabel)
    for c, h in enumerate(to_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    for gt_row in ground_truth:
        source_file = gt_row["source_file"]
        anon_name = None
        for key, val in anon_map.items():
            if key in source_file.lower() or source_file.lower() in key:
                anon_name = val
                break
        if anon_name is None:
            anon_name = source_file

        gt_date = _parse_date(gt_row["date"])
        if gt_date is None:
            gt_date = str(gt_row["date"]).strip()
        gt_time_out = gt_row.get("time_out", "")

        ws.cell(row=row, column=1, value=anon_name).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", wrap_text=True)
        ws.cell(row=row, column=2, value=gt_date).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=gt_time_out).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

        col = 4
        for approach_id, approach_label, _ in APPROACHES:
            m = approach_lookup[approach_id].get((source_file, gt_date))
            if m and m["ext_time_out"]:
                ext_val = m["ext_time_out"]
                fill = MATCH_FILL if m["time_out_correct"] else MISMATCH_FILL
            else:
                ext_val = "—"
                fill = MISMATCH_FILL
            cell = ws.cell(row=row, column=col, value=ext_val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            cell.fill = fill
            col += 1

        row += 1

    row += 1

    # ═══════════════════════════════════════════
    # Table 3: Correctness Summary
    # ═══════════════════════════════════════════
    ws.cell(row=row, column=1, value="CORRECTNESS SUMMARY (within ±30min tolerance)").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    row += 1

    summary_headers = ["Metric"]
    for _aid, alabel, _ in APPROACHES:
        summary_headers.append(alabel)
    for c, h in enumerate(summary_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    # Compute correctness counts
    correct_ti = {}
    correct_to = {}
    for approach_id, approach_label, comp in all_results:
        matched = comp["matched"]
        correct_ti[approach_id] = sum(1 for m in matched if m["time_in_correct"])
        correct_to[approach_id] = sum(1 for m in matched if m["time_out_correct"])
        total = len(matched)
        if total == 0:
            correct_ti[approach_id] = 0
            correct_to[approach_id] = 0

    total_gt = len(ground_truth)
    summary_rows = [
        ("Total GT Rows", lambda aid: total_gt),
        ("Time-In Correct", lambda aid: f"{correct_ti[aid]} / {total_gt} ({correct_ti[aid]/total_gt:.0%})" if total_gt else "N/A"),
        ("Time-Out Correct", lambda aid: f"{correct_to[aid]} / {total_gt} ({correct_to[aid]/total_gt:.0%})" if total_gt else "N/A"),
    ]

    for label, getter in summary_rows:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", wrap_text=True)
        for c, (aid, alabel, _) in enumerate(APPROACHES, 2):
            val = getter(aid)
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    for i in range(4, num_cols + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 22

    wb.save(BENCH_OUTPUT)
    print(f"  Added 'Time Comparison' sheet to {BENCH_OUTPUT}")


def copy_debug_images():
    """Copy debug images to combined folder."""
    import shutil

    for folder, label, _ in APPROACHES:
        debug_dir = f"output/{folder}/debug"
        if os.path.exists(debug_dir):
            for f in os.listdir(debug_dir):
                src = os.path.join(debug_dir, f)
                dst = os.path.join(OUTPUT_DIR, "debug", f"{folder}_{f}")
                if os.path.isfile(src):
                    shutil.copy2(src, dst)
                    print(f"Copied: {dst}")

    # Also copy shared debug images from output/debug/ (for approaches that don't have their own debug dir)
    shared_debug = "output/debug"
    if os.path.exists(shared_debug):
        for f in os.listdir(shared_debug):
            if f.endswith(".png"):
                src = os.path.join(shared_debug, f)
                dst = os.path.join(OUTPUT_DIR, "debug", f)
                if os.path.isfile(src):
                    shutil.copy2(src, dst)
                    print(f"Copied shared: {dst}")


if __name__ == "__main__":
    os.makedirs(f"{OUTPUT_DIR}/debug", exist_ok=True)
    create_benchmark_combined()
    add_ground_truth_comparison()
    copy_debug_images()
    print("Done.")
