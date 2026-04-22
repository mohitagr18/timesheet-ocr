"""Re-run band_crop_vlm_cloud for specific files that had missed pages.

This script:
  1. Removes existing rows for the target files from merged_results.xlsx
     (so the re-run appends cleanly without duplicates)
  2. Removes stale per-file report/benchmark files for the target files
  3. Runs the pipeline for ONLY these files
  4. Regenerates the combined benchmark report

Usage:
    uv run python scripts/rerun_band_crop_partial.py
    uv run python scripts/rerun_band_crop_partial.py --dry-run   # preview only
"""

from __future__ import annotations

import argparse
import json
import logging
import shutil
import subprocess
import sys
import time
from pathlib import Path

import yaml

PROJECT_ROOT = Path(__file__).resolve().parent.parent
CONFIG_PATH = PROJECT_ROOT / "config.yaml"
INPUT_DIR = PROJECT_ROOT / "input"
OUTPUT_DIR = PROJECT_ROOT / "output"
APPROACH_ID = "band_crop_vlm_cloud"
APPROACH_DIR = OUTPUT_DIR / APPROACH_ID

# The 8 input files whose pages were missed by PP-DocLayoutV3
TARGET_FILES = [
    "H.Leal 123125-010626.pdf",
    "J.Jackson Timesheets 121025-121625.pdf",
    "J.Jackson Timesheets 121725-122325.pdf",
    "L.Moran Timesheet 122425-123025.pdf",
    "L.Moran Timesheet 123125-010626.pdf",
    "N.Rivera Timesheets -021826-022426.pdf",
    "R.Elliott Timesheets 122425-123025.pdf",
    "S.Bussa Timesheet - 020426-021026 c.pdf",
]


def setup_logging():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)-7s | %(message)s",
        datefmt="%H:%M:%S",
    )
    return logging.getLogger(__name__)


def get_anonymized_names():
    """Map real filenames to their anonymized counterparts."""
    from src.phi import PhiAnonymizer

    supported = {".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp"}
    all_files = sorted(
        f.name
        for f in INPUT_DIR.iterdir()
        if f.suffix.lower() in supported and not f.name.startswith(".")
    )
    anonymizer = PhiAnonymizer(all_files)
    return {real: anonymizer.anonymize_filename(real) for real in TARGET_FILES}


def strip_rows_from_merged(anon_names: set[str], dry_run: bool, logger):
    """Remove rows for target files from merged_results.xlsx."""
    merged = APPROACH_DIR / "merged_results.xlsx"
    if not merged.exists():
        logger.warning(f"merged_results.xlsx not found at {merged}")
        return

    from openpyxl import load_workbook

    wb = load_workbook(merged)
    ws = wb.active

    rows_to_delete = []
    for row_idx in range(ws.max_row, 1, -1):  # iterate bottom-up for safe deletion
        source = str(ws.cell(row=row_idx, column=1).value or "")
        if source in anon_names:
            rows_to_delete.append(row_idx)

    if not rows_to_delete:
        logger.info("No existing rows to remove from merged_results.xlsx")
        return

    logger.info(
        f"{'Would remove' if dry_run else 'Removing'} {len(rows_to_delete)} rows "
        f"from merged_results.xlsx for {len(anon_names)} target files"
    )

    if not dry_run:
        # Back up first
        backup = APPROACH_DIR / "merged_results_backup_before_rerun.xlsx"
        shutil.copy2(merged, backup)
        logger.info(f"Backed up to {backup.name}")

        for row_idx in rows_to_delete:  # already sorted bottom-up
            ws.delete_rows(row_idx)
        wb.save(merged)
        logger.info(f"Saved cleaned merged_results.xlsx ({ws.max_row - 1} rows remaining)")

    wb.close()


def remove_stale_per_file_outputs(anon_names: dict[str, str], dry_run: bool, logger):
    """Remove report/benchmark files for target files so pipeline regenerates them."""
    for real_name, anon_name in anon_names.items():
        stem = Path(anon_name).stem
        patterns = [
            f"{stem}_report.json",
            f"benchmark_{stem}.xlsx",
            f"{stem}_results.csv",
            f"{stem}_results.json",
            f"{stem}_review.json",
        ]
        for pat in patterns:
            f = APPROACH_DIR / pat
            if f.exists():
                logger.info(f"  {'Would delete' if dry_run else 'Deleting'}: {f.name}")
                if not dry_run:
                    f.unlink()


def run_pipeline(target_paths: list[Path], logger):
    """Run band_crop_vlm_cloud pipeline for only the target files."""
    # Update config to band_crop mode
    with open(CONFIG_PATH) as f:
        config = yaml.safe_load(f)
    original_mode = config.get("extraction_mode")
    config["extraction_mode"] = APPROACH_ID
    with open(CONFIG_PATH, "w") as f:
        yaml.dump(config, f, default_flow_style=False, sort_keys=False)

    try:
        from src.config import load_config
        from src.pipeline import Pipeline

        app_config = load_config(str(CONFIG_PATH))
        app_config.paths.output_dir = str(APPROACH_DIR)

        pipeline = Pipeline(app_config)
        results = pipeline.process_directory(
            INPUT_DIR,
            generate_combined=False,
            files_to_process=target_paths,
        )

        # Print summary
        for r in results:
            if isinstance(r, dict):
                src = r["source_file"]
                rows = r["total_rows"]
                accepted = r["accepted_count"]
                t = r["processing_time_seconds"]
            else:
                src = r.source_file
                rows = r.total_rows
                accepted = r.accepted_count
                t = r.processing_time_seconds
            logger.info(f"  ✓ {src}: {rows} rows ({accepted} accepted) in {t:.1f}s")

        pipeline.cleanup()
        return results

    finally:
        # Restore original extraction mode
        config["extraction_mode"] = original_mode
        with open(CONFIG_PATH, "w") as f:
            yaml.dump(config, f, default_flow_style=False, sort_keys=False)
        logger.info(f"Restored config extraction_mode to '{original_mode}'")


def regenerate_combined_report(logger):
    """Re-run build_combined_report.py."""
    logger.info("Regenerating combined benchmark report...")
    result = subprocess.run(
        [sys.executable, "scripts/build_combined_report.py"],
        cwd=PROJECT_ROOT,
        capture_output=True,
        text=True,
    )
    # Show just the metrics lines
    for line in result.stdout.splitlines():
        if "matched=" in line or "Saved:" in line or "rows" in line.lower():
            logger.info(f"  {line.strip()}")


def main():
    parser = argparse.ArgumentParser(
        description="Re-run band_crop for files with missed pages"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview what would happen without making changes",
    )
    args = parser.parse_args()

    logger = setup_logging()
    logger.info(f"{'DRY RUN — ' if args.dry_run else ''}Re-running band_crop for {len(TARGET_FILES)} files")

    # 1. Map real → anonymized names
    anon_map = get_anonymized_names()
    anon_names = set(anon_map.values())

    logger.info("\nTarget files:")
    for real, anon in anon_map.items():
        logger.info(f"  {real}  →  {anon}")

    # 2. Verify all input files exist
    target_paths = []
    for real_name in TARGET_FILES:
        p = INPUT_DIR / real_name
        if not p.exists():
            logger.error(f"Input file not found: {p}")
            return 1
        target_paths.append(p)

    # 3. Strip existing rows from merged_results.xlsx
    logger.info("\n--- Step 1: Clean existing rows from merged_results.xlsx ---")
    strip_rows_from_merged(anon_names, args.dry_run, logger)

    # 4. Remove stale per-file outputs
    logger.info("\n--- Step 2: Remove stale per-file outputs ---")
    remove_stale_per_file_outputs(anon_map, args.dry_run, logger)

    if args.dry_run:
        logger.info("\n--- DRY RUN complete. No changes made. ---")
        return 0

    # 5. Run pipeline
    logger.info(f"\n--- Step 3: Running band_crop pipeline for {len(target_paths)} files ---")
    start = time.time()
    results = run_pipeline(target_paths, logger)
    elapsed = time.time() - start
    logger.info(f"Pipeline completed in {elapsed:.1f}s")

    # 6. Regenerate combined report
    logger.info("\n--- Step 4: Regenerating combined report ---")
    regenerate_combined_report(logger)

    logger.info("\n✅ Done!")
    return 0


if __name__ == "__main__":
    sys.exit(main())
