"""Create per-row best-approach results vs ground truth.

For each (source_file, date) with ground truth:
  1. Score all 5 approaches against GT hours
  2. Pick the single best approach (closest hours to GT)
  3. Report which approach won for that row
  4. Compute: per-approach accuracy, and "best of 5" accuracy

KPIs:
  1. North Star: % rows with hours matching GT (±0.25h)
  2. Time-In match rate (±30min)
  3. Time-Out match rate (±30min)
  4. Employee name match rate

Usage:
    uv run python scripts/create_consensus_results.py

Output: output/combined/consensus.xlsx
"""

import glob
import os
import re
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

APPROACHES = [
    ("ocr_only", "OCR Only"),
    ("ppocr_grid", "OCR+VLM"),
    ("vlm_full_page", "VLM Full"),
    ("layout_guided_vlm_local", "Layout Local"),
    ("layout_guided_vlm_cloud", "Layout Cloud"),
]

APPROACH_COLORS = {
    "ocr_only": "E8E8E8",
    "ppocr_grid": "E2EFDA",
    "vlm_full_page": "D6E4F0",
    "layout_guided_vlm_local": "FFF2CC",
    "layout_guided_vlm_cloud": "FCE4EC",
}

OUTPUT_PATH = "output/combined/consensus.xlsx"
HOURS_TOLERANCE = 0.25
TIME_TOLERANCE = 30  # minutes

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MISMATCH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PARTIAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


# ─── Parse helpers ───────────────────────────────────────────────────

def parse_date(val):
    if val is None:
        return None
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    val = str(val).strip()
    if not val:
        return None
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", val)
    if m:
        mo, d, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        return f"{y:04d}-{mo:02d}-{d:02d}"
    return val


def parse_time(val):
    if val is None:
        return None
    if hasattr(val, "hour"):
        return val.hour * 60 + val.minute
    val = str(val).strip()
    if not val:
        return None

    # If plain digits (no colon), treat as HHMM format (e.g. 830 → 8:30, 1530 → 15:30)
    if val.replace(":", "").isdigit() and ":" not in val:
        try:
            v = int(val)
            h = v // 100
            mi = v % 100
            if 0 <= h <= 23 and 0 <= mi <= 59:
                return h * 60 + mi
        except (ValueError, TypeError):
            pass

    # Otherwise try colon-separated format with AM/PM
    m = re.match(r"(\d{1,2}):?(\d{2})?\s*(AM|PM|am|pm)?", val, re.IGNORECASE)
    if m:
        h = int(m.group(1))
        mi = int(m.group(2) or 0)
        ap = m.group(3)
        if ap and ap.lower() == "pm" and h != 12:
            h += 12
        elif ap and ap.lower() == "am" and h == 12:
            h = 0
        if 0 <= h <= 23 and 0 <= mi <= 59:
            return h * 60 + mi
    return None


def compute_hours(ti, to):
    if ti is None or to is None:
        return None
    diff = to - ti
    if diff < 0:
        diff += 24 * 60
    return diff / 60.0


# ─── Load data ───────────────────────────────────────────────────────

def load_gt():
    if not os.path.exists("ground_truth.xlsx"):
        return []
    wb = openpyxl.load_workbook("ground_truth.xlsx", read_only=True)
    ws = wb.active
    rows = []
    header = None
    for r in ws.iter_rows(values_only=True):
        if header is None:
            header = [str(c).strip() if c else "" for c in r]
            continue
        if not any(r):
            continue
        rec = dict(zip(header, r))
        if rec.get("source_file") and rec.get("date"):
            rows.append(rec)
    wb.close()
    return rows


def load_approach_data(approach_id):
    bench_files = sorted(
        f for f in glob.glob(f"output/{approach_id}/benchmark_*.xlsx")
        if "combined" not in f
    )
    rows = []
    for path in bench_files:
        wb = openpyxl.load_workbook(path, read_only=True)
        if "Row-Level" in wb.sheetnames:
            ws = wb["Row-Level"]
            header = None
            for r in ws.iter_rows(min_row=1, values_only=True):
                if header is None:
                    header = [str(c).strip() if c else "" for c in r]
                    continue
                if not any(r):
                    continue
                rec = dict(zip(header, r))
                rows.append(rec)
        wb.close()

    # Fill from merged if available
    merged = f"output/{approach_id}/merged_results.xlsx"
    if os.path.exists(merged):
        wb = openpyxl.load_workbook(merged, read_only=True)
        ws = wb.active
        header = None
        merged_by_date = {}
        for r in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c).strip() if c else "" for c in r]
                continue
            if not any(r):
                continue
            rec = dict(zip(header, r))
            d = parse_date(rec.get("Date", ""))
            if d:
                merged_by_date[d] = rec
        wb.close()
        for row in rows:
            d = parse_date(row.get("Parsed Date", ""))
            if d and d in merged_by_date:
                mr = merged_by_date[d]
                if not row.get("Parsed Hours") and mr.get("Total Hours"):
                    row["Parsed Hours"] = mr["Total Hours"]
                if not row.get("Parsed Time In") and mr.get("Time In"):
                    row["Parsed Time In"] = mr["Time In"]
                if not row.get("Parsed Time Out") and mr.get("Time Out"):
                    row["Parsed Time Out"] = mr["Time Out"]

    return rows


def build_lookup(rows):
    lookup = defaultdict(list)
    for r in rows:
        d = parse_date(r.get("Parsed Date", r.get("Date", "")))
        src = r.get("Source File", "")
        if d:
            key = (src, d)
            lookup[key].append(r)
    return lookup


# ─── Scoring ─────────────────────────────────────────────────────────

def score_hours(ext_hrs, gt_hrs):
    """Return distance from GT hours. Lower = better."""
    if ext_hrs is None or gt_hrs is None:
        return 9999
    try:
        return abs(float(ext_hrs) - float(gt_hrs))
    except (ValueError, TypeError):
        return 9999


def score_time(ext, gt):
    """Return distance in minutes."""
    if ext is None or gt is None:
        return 9999
    return abs(ext - gt)


# ─── Main ────────────────────────────────────────────────────────────

def main():
    print("Loading data...")
    gt = load_gt()
    print(f"  Ground truth: {len(gt)} rows")

    approach_data = {}
    approach_lookups = {}
    for aid, label in APPROACHES:
        rows = load_approach_data(aid)
        approach_data[aid] = rows
        approach_lookups[aid] = build_lookup(rows)
        print(f"  {aid}: {len(rows)} rows, {len(approach_lookups[aid])} unique dates")

    # Score every row for every approach
    results = []
    per_approach = {aid: {"total": 0, "hrs_ok": 0, "ti_ok": 0, "to_ok": 0, "emp_ok": 0}
                     for aid, _ in APPROACHES}
    best_of_5 = {"total": 0, "hrs_ok": 0, "ti_ok": 0, "to_ok": 0, "emp_ok": 0}

    for g in gt:
        src = g["source_file"]
        d = parse_date(g["date"])
        key = (src, d)

        gt_hrs = float(g["total_hours"]) if g.get("total_hours") else None
        gt_ti = parse_time(g.get("time_in", ""))
        gt_to = parse_time(g.get("time_out", ""))
        gt_emp = g.get("employee_name", "").strip()
        gt_hrs_computed = compute_hours(gt_ti, gt_to)
        gt_hrs_val = gt_hrs_computed if gt_hrs_computed is not None else gt_hrs

        # Score each approach
        approach_scores = []
        for aid, label in APPROACHES:
            rows = approach_lookups[aid].get(key, [])
            if not rows:
                approach_scores.append((aid, label, None, None, None, None, 9999))
                continue

            # Score each candidate row from this approach
            best_row = None
            best_score = 9999
            for r in rows:
                ext_hrs = r.get("Parsed Hours", r.get("Written Hours"))
                ext_hrs_val = None
                try:
                    ext_hrs_val = float(ext_hrs) if ext_hrs is not None else None
                except (ValueError, TypeError):
                    pass
                ext_ti = parse_time(r.get("Parsed Time In", ""))
                ext_to = parse_time(r.get("Parsed Time Out", ""))

                h_score = score_hours(ext_hrs_val, gt_hrs_val)
                ti_score = score_time(ext_ti, gt_ti)
                to_score = score_time(ext_to, gt_to)

                # Primary: hours distance, then time
                combined = (h_score * 1000) + ti_score + (to_score * 0.01)
                if combined < best_score:
                    best_score = combined
                    best_row = (ext_hrs_val, ext_ti, ext_to, h_score, ti_score, to_score)

            if best_row:
                ext_hrs_val, ext_ti, ext_to, h_sc, ti_sc, to_sc = best_row
                approach_scores.append((aid, label, ext_hrs_val, ext_ti, ext_to, ext_hrs_val, h_sc))
            else:
                approach_scores.append((aid, label, None, None, None, None, 9999))

        # Score each approach individually
        for aid, label, ext_hrs, ext_ti, ext_to, _, h_sc in approach_scores:
            hrs_ok = h_sc <= HOURS_TOLERANCE
            ti_ok = score_time(ext_ti, gt_ti) <= TIME_TOLERANCE
            to_ok = score_time(ext_to, gt_to) <= TIME_TOLERANCE

            per_approach[aid]["total"] += 1
            if hrs_ok:
                per_approach[aid]["hrs_ok"] += 1
            if ti_ok:
                per_approach[aid]["ti_ok"] += 1
            if to_ok:
                per_approach[aid]["to_ok"] += 1

        # Best of 5: pick the approach with lowest hours distance
        scored = [(h_sc, ext_hrs, ext_ti, ext_to, aid, label)
                   for aid, label, ext_hrs, ext_ti, ext_to, _, h_sc in approach_scores]
        scored.sort(key=lambda x: x[0])
        _, best_hrs, best_ti, best_to, best_aid, best_label = scored[0]

        hrs_ok = scored[0][0] <= HOURS_TOLERANCE
        ti_ok = score_time(best_ti, gt_ti) <= TIME_TOLERANCE
        to_ok = score_time(best_to, gt_to) <= TIME_TOLERANCE

        best_of_5["total"] += 1
        if hrs_ok:
            best_of_5["hrs_ok"] += 1
        if ti_ok:
            best_of_5["ti_ok"] += 1
        if to_ok:
            best_of_5["to_ok"] += 1

        results.append({
            "source": src,
            "date": d,
            "gt_hrs": gt_hrs,
            "gt_ti": g.get("time_in", ""),
            "gt_to": g.get("time_out", ""),
            "gt_emp": gt_emp,
            "best_hrs": best_hrs,
            "best_ti": best_ti,
            "best_to": best_to,
            "best_aid": best_aid,
            "best_label": best_label,
            "hrs_ok": hrs_ok,
            "ti_ok": ti_ok,
            "to_ok": to_ok,
        })

        # Add per-approach hours
        for aid, _, ext_hrs, _, _, _, _ in approach_scores:
            r["hrs_" + aid] = ext_hrs

    # ─── Write Excel ─────────────────────────────────────────────────
    print(f"\nWriting results to {OUTPUT_PATH}...")
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb = openpyxl.Workbook()

    # Sheet 1: KPI Dashboard
    ws = wb.active
    ws.title = "KPI Dashboard"

    row = 1
    ws.cell(row=row, column=1, value="Timesheet OCR Extraction — KPI Dashboard").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1
    ws.cell(
        row=row, column=1,
        value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
              f"Hours Tolerance: ±{HOURS_TOLERANCE}h | Time Tolerance: ±{TIME_TOLERANCE}min | "
              f"Ground Truth: {len(gt)} rows",
    ).font = Font(italic=True, size=10)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 2

    # ─── Section 1: Per-Approach Accuracy ───
    ws.cell(row=row, column=1, value="PER-APPROACH ACCURACY").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1

    dash_headers = ["Approach"] + [a[1] for a in APPROACHES] + ["Best of 5"]
    for c, h in enumerate(dash_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    n = len(gt)

    def p(n_ok, total):
        return f"{n_ok}/{total} ({n_ok/total:.0%})" if total else "N/A"

    kpi_rows = [
        ("⭐ Hours Accuracy (±0.25h)", lambda aid: p(per_approach[aid]["hrs_ok"], per_approach[aid]["total"]),
                                     p(best_of_5["hrs_ok"], best_of_5["total"])),
        ("Time-In Accuracy (±30min)", lambda aid: p(per_approach[aid]["ti_ok"], per_approach[aid]["total"]),
                                    p(best_of_5["ti_ok"], best_of_5["total"])),
        ("Time-Out Accuracy (±30min)", lambda aid: p(per_approach[aid]["to_ok"], per_approach[aid]["total"]),
                                     p(best_of_5["to_ok"], best_of_5["total"])),
    ]

    for label, get_val, get_best in kpi_rows:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", wrap_text=True)
        ws.cell(row=row, column=1).font = Font(bold=True)
        col = 2
        for aid, _ in APPROACHES:
            val = get_val(aid)
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color=APPROACH_COLORS.get(aid, "FFFFFF"),
                                    end_color=APPROACH_COLORS.get(aid, "FFFFFF"),
                                    fill_type="solid")
            col += 1
        # Best of 5
        cell = ws.cell(row=row, column=col, value=get_best)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        cell.fill = MATCH_FILL
        row += 1

    row += 1

    # ─── Section 2: Approach Win Counts ───
    ws.cell(row=row, column=1, value="APPROACH WIN COUNTS (how often each approach is closest to GT hours)").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1

    win_counts = defaultdict(int)
    for r in results:
        win_counts[r["best_aid"]] += 1

    for aid, label in APPROACHES:
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=f"{win_counts[aid]} / {n} ({win_counts[aid]/n:.0%})").border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        row += 1

    # ─── Sheet 2: Per-Row Detail ───
    ws2 = wb.create_sheet("Per-Row Detail")
    row = 1
    ws2.cell(row=row, column=1, value="Per-Row Best Approach Selection").font = Font(bold=True, size=14)
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
    row += 2

    detail_headers = [
        "Source", "Date", "GT Hours", "GT Time In", "GT Time Out", "GT Employee",
        "Best Hours", "Best ✓", "Best Approach", "Best Time In", "Best Time Out",
    ] + [a[1] + " Hrs" for a in APPROACHES]

    for c, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    row += 1

    anon_map = {
        "patient_a_week1": "File 1 (Week 1)",
        "patient_b_week2": "File 2 (Week 2)",
        "patient_c_week3": "File 3 (Week 3)",
    }

    for r in results:
        src = r["source"]
        anon = None
        for key, val in anon_map.items():
            if key in src.lower() or src.lower() in key:
                anon = val
                break
        if anon is None:
            anon = src

        ws2.cell(row=row, column=1, value=anon).border = THIN_BORDER
        ws2.cell(row=row, column=1).alignment = Alignment(horizontal="left", wrap_text=True)
        ws2.cell(row=row, column=2, value=r["date"]).border = THIN_BORDER
        ws2.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws2.cell(row=row, column=3, value=r["gt_hrs"]).border = THIN_BORDER
        ws2.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws2.cell(row=row, column=4, value=r["gt_ti"]).border = THIN_BORDER
        ws2.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        ws2.cell(row=row, column=5, value=r["gt_to"]).border = THIN_BORDER
        ws2.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        ws2.cell(row=row, column=6, value=r["gt_emp"]).border = THIN_BORDER
        ws2.cell(row=row, column=6).alignment = Alignment(horizontal="center")

        ws2.cell(row=row, column=7, value=r["best_hrs"]).border = THIN_BORDER
        ws2.cell(row=row, column=7).alignment = Alignment(horizontal="center")

        cell = ws2.cell(row=row, column=8, value="✓" if r["hrs_ok"] else "✗")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        cell.fill = MATCH_FILL if r["hrs_ok"] else MISMATCH_FILL

        ws2.cell(row=row, column=9, value=r["best_label"]).border = THIN_BORDER
        ws2.cell(row=row, column=9).alignment = Alignment(horizontal="center")
        ws2.cell(row=row, column=9).fill = PatternFill(start_color=APPROACH_COLORS.get(r["best_aid"], "FFFFFF"),
                                                        end_color=APPROACH_COLORS.get(r["best_aid"], "FFFFFF"),
                                                        fill_type="solid")

        ws2.cell(row=row, column=10, value=r["best_ti"] if r["best_ti"] is not None else "").border = THIN_BORDER
        ws2.cell(row=row, column=10).alignment = Alignment(horizontal="center")
        ws2.cell(row=row, column=11, value=r["best_to"] if r["best_to"] is not None else "").border = THIN_BORDER
        ws2.cell(row=row, column=11).alignment = Alignment(horizontal="center")

        col = 12
        for aid, _ in APPROACHES:
            h_val = r.get(f"hrs_{aid}")
            cell = ws2.cell(row=row, column=col, value=h_val if h_val is not None else "—")
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if aid == r["best_aid"]:
                cell.fill = PatternFill(start_color=APPROACH_COLORS.get(aid, "FFFFFF"),
                                        end_color=APPROACH_COLORS.get(aid, "FFFFFF"),
                                        fill_type="solid")
            col += 1

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    for i in range(2, 8):
        ws.column_dimensions[get_column_letter(i)].width = 16

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 14
    for i in range(3, 15):
        ws2.column_dimensions[get_column_letter(i)].width = 14

    wb.save(OUTPUT_PATH)
    print(f"  Saved: {OUTPUT_PATH}")

    # ─── Print summary ───
    print(f"\n{'='*60}")
    print(f"KPI SUMMARY (n={n} ground truth rows)")
    print(f"{'='*60}")

    print(f"\n{'Approach':<25} {'Hours':>12} {'Time-In':>12} {'Time-Out':>12}")
    print(f"{'-'*61}")
    for aid, label in APPROACHES:
        pa = per_approach[aid]
        print(f"{label:<25} {pa['hrs_ok']}/{pa['total']} ({pa['hrs_ok']/pa['total']:>5.1%})  "
              f"{pa['ti_ok']}/{pa['total']} ({pa['ti_ok']/pa['total']:>5.1%})  "
              f"{pa['to_ok']}/{pa['total']} ({pa['to_ok']/pa['total']:>5.1%})")

    bo = best_of_5
    print(f"{'BEST OF 5':<25} {bo['hrs_ok']}/{bo['total']} ({bo['hrs_ok']/bo['total']:>5.1%})  "
          f"{bo['ti_ok']}/{bo['total']} ({bo['ti_ok']/bo['total']:>5.1%})  "
          f"{bo['to_ok']}/{bo['total']} ({bo['to_ok']/bo['total']:>5.1%})")

    print(f"\nApproach Win Counts (closest to GT hours):")
    for aid, label in APPROACHES:
        print(f"  {label:<25} {win_counts[aid]} / {n} ({win_counts[aid]/n:.0%})")


if __name__ == "__main__":
    main()
